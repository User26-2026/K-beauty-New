"""Пересчет прайса с потолком по российским оптовикам.

Прежний прайс строился от кыргызской цены (Дордой). Для российского
оптового покупателя это не альтернатива: туда надо ехать, растаможивать и
маркировать самому. Реальный потолок задают KEAUTY и KoreaTrade — у них
покупатель берет тот же товар с документами.

Новая цена = минимальный опт конкурента минус запас, но не ниже
себестоимости с минимальной наценкой. Где конкурента нет, цена остается
прежней: сравнивать не с чем, гадать не будем.

Запуск:
    python3 tools/rebuild_pricelist.py
    python3 tools/rebuild_pricelist.py --discount 15 --floor 20
"""

import argparse
import os
import re

import openpyxl
import pandas as pd

PRICES = "outputs/prices_normalized.xlsx"
OUT = "outputs/pricelist_rebuilt.xlsx"
RU_SUPPLIERS = ("KEAUTY", "KoreaTrade")


def read_current(path):
    """Действующий прайс: штрихкод, наименование, остаток, цена."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for number, name, barcode, stock, price in ws.iter_rows(min_row=6, max_col=5, values_only=True):
        code = re.sub(r"\D", "", str(barcode or ""))
        if len(code) < 8 or price is None:
            continue
        rows.append({"Штрихкод": code, "Товар": str(name).strip(),
                     "Остаток, шт": stock, "Цена сейчас, руб": float(price)})
    wb.close()
    return pd.DataFrame(rows)


def read_costs(path):
    """Себестоимость по последнему приходу из файла остатков."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for article, name, unit, cost, stock, total, shelf, *rest in ws.iter_rows(
            min_row=5, max_col=8, values_only=True):
        if not name or cost is None:
            continue
        try:
            cost = float(cost)
        except (TypeError, ValueError):
            continue
        rows.append({"Товар": str(name).strip(), "Себестоимость, руб": cost,
                     "Срок годности": shelf})
    wb.close()
    return pd.DataFrame(rows).drop_duplicates("Товар")


def ru_competitors():
    """Минимальная оптовая цена по каждому штрихкоду у KEAUTY и KoreaTrade."""
    df = pd.read_excel(PRICES, dtype={"Штрихкод": str})
    ru = df[df["Поставщик"].isin(RU_SUPPLIERS) & df["Закупка, KRW"].notna()].copy()
    ru["Штрихкод"] = ru["Штрихкод"].fillna("").str.replace(r"\D", "", regex=True)
    ru = ru[ru["Штрихкод"].str.len() >= 8]
    wide = (ru.sort_values("Закупка, KRW")
              .drop_duplicates(["Штрихкод", "Поставщик"])
              .pivot(index="Штрихкод", columns="Поставщик", values="Закупка, KRW"))
    for supplier in RU_SUPPLIERS:
        if supplier not in wide.columns:
            wide[supplier] = None
    wide = wide.rename(columns={s: f"{s}, руб" for s in RU_SUPPLIERS})
    wide["Потолок, руб"] = wide.min(axis=1)
    return wide


def main(price_path, stock_path, discount, floor):
    current = read_current(price_path)
    costs = read_costs(stock_path)
    table = current.merge(costs, on="Товар", how="left").merge(
        ru_competitors(), on="Штрихкод", how="left")

    has_ceiling = table["Потолок, руб"].notna()
    proposed = (table["Потолок, руб"] * (1 - discount / 100)).round(0)
    # Ниже себестоимости с минимальной наценкой не опускаемся.
    minimum = (table["Себестоимость, руб"] * (1 + floor / 100)).round(0)
    table["Новая цена, руб"] = table["Цена сейчас, руб"]
    table.loc[has_ceiling, "Новая цена, руб"] = (
        pd.concat([proposed[has_ceiling], minimum[has_ceiling]], axis=1).max(axis=1))
    # Цену не снижаем: если конкурент дешевле нас, остаемся на своей.
    table["Новая цена, руб"] = table[["Новая цена, руб", "Цена сейчас, руб"]].max(axis=1)

    table["Наценка сейчас, %"] = ((table["Цена сейчас, руб"] / table["Себестоимость, руб"] - 1) * 100).round(1)
    table["Наценка новая, %"] = ((table["Новая цена, руб"] / table["Себестоимость, руб"] - 1) * 100).round(1)
    table["Рост цены, руб"] = (table["Новая цена, руб"] - table["Цена сейчас, руб"]).round(0)
    table["Рост на остаток, руб"] = (table["Рост цены, руб"] * table["Остаток, шт"]).round(0)
    table["Ниже конкурента на, %"] = (
        (1 - table["Новая цена, руб"] / table["Потолок, руб"]) * 100).round(1)
    # Конкурент дешевле нашей себестоимости — конкурировать нечем. По сводке
    # причина в разнесении логистики по стоимости: тяжелые дешевые позиции
    # (тонеры 300 мл, шампуни) получают завышенную себестоимость.
    table["Конкурент ниже себестоимости"] = (
        table["Потолок, руб"] < table["Себестоимость, руб"]).fillna(False)
    table["Основание"] = "конкурента нет, цена без изменений"
    table.loc[has_ceiling, "Основание"] = f"потолок РФ минус {discount:.0f}%"
    table.loc[has_ceiling & (minimum > proposed), "Основание"] = f"себестоимость плюс {floor:.0f}%"

    table = table.sort_values("Рост на остаток, руб", ascending=False)
    table.to_excel(OUT, index=False)

    rival = table[table["Потолок, руб"].notna()]
    print(f"Позиций в прайсе: {len(table)}   из них есть у российских оптовиков: {len(rival)}")
    print(f"Правило: потолок минус {discount:.0f}%, но не ниже себестоимости плюс {floor:.0f}%\n")
    print(f"{'Выручка по прайсу сейчас':<34} {(table['Цена сейчас, руб'] * table['Остаток, шт']).sum():>16,.0f} руб")
    print(f"{'Выручка по новому прайсу':<34} {(table['Новая цена, руб'] * table['Остаток, шт']).sum():>16,.0f} руб")
    print(f"{'Прирост':<34} {table['Рост на остаток, руб'].sum():>16,.0f} руб")
    print(f"\nПо {len(rival)} позициям с конкурентом наценка: было "
          f"{rival['Наценка сейчас, %'].median():.1f}%, стало "
          f"{rival['Наценка новая, %'].median():.1f}% (медиана)")
    print(f"Мы дешевле конкурента на {rival['Ниже конкурента на, %'].median():.1f}% (медиана)")
    print("\nОснования новой цены:")
    print(table["Основание"].value_counts().to_string())
    below = int(table["Конкурент ниже себестоимости"].sum())
    if below:
        print(f"\nУ {below} позиций конкурент дешевле нашей себестоимости — "
              f"конкурировать нечем, цена оставлена по правилу минимальной наценки.")
    print("\nТоп-12 по приросту:")
    cols = ["Товар", "Остаток, шт", "Себестоимость, руб", "Цена сейчас, руб",
            "Потолок, руб", "Новая цена, руб", "Рост на остаток, руб"]
    print(table[cols].head(12).to_string(index=False))
    print(f"\nСохранено: {OUT}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--price", default="", help="файл действующего прайса")
    parser.add_argument("--stock", default="", help="файл остатков с себестоимостью")
    parser.add_argument("--discount", type=float, default=12.0, help="запас к цене конкурента, %%")
    parser.add_argument("--floor", type=float, default=20.0, help="минимальная наценка к себестоимости, %%")
    ns = parser.parse_args()
    main(ns.price, ns.stock, ns.discount, ns.floor)
