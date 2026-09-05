"""Сравнение цен по названию — для прайсов без штрихкода.

Обычно товары сводим по штрихкоду. Но часть прайсов приходит без него:
у Классика в прайсе LANEIGE только номер, название и цена. Тогда сводим по
названию: чистим его от бренда, скобок и фасовки, сравниваем набор слов и
объем. Совпадение по одному набору слов, но с разным объемом — это разные
товары, их не смешиваем.

Запуск:
    python3 tools/compare_by_name.py LANEIGE
"""

import argparse
import os
import re

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = "outputs/prices_normalized.xlsx"
OUT_DIR = "outputs"
from rates import IMPORT_MULTIPLIER as IMPORT, KRW_RUB as RATE

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BEST_FILL = PatternFill("solid", fgColor="E2EFDA")

# Слова, которые ничего не различают внутри одного бренда.
NOISE = {"THE", "AND", "FOR", "NEW", "SET", "EA", "PCS", "ML", "GR", "KIT"}
# В прайсах рядом с обычной ценой встречаются лоты и товар с истекающим
# сроком. Такие строки сравнивать с обычной ценой нельзя.
SPECIAL = r"EXP\s*DATE|FULL[-\s]?LOT|\bLOT\b|UNITS|\bMOQ\b|СРОК"
UNIT = r"(?:ML|L|G|KG|EA|PCS|매)"


def volume(text):
    """Фасовка из строки: 20g, 60ML, 25ml*10ea -> нормализованный вид."""
    if not isinstance(text, str):
        return ""
    text = text.upper().replace(" ", "")
    pack = re.search(rf"(\d+(?:\.\d+)?){UNIT}?[X*](\d+)EA", text)
    if pack:
        return f"{pack.group(1)}X{pack.group(2)}"
    single = re.search(rf"(\d+(?:\.\d+)?)\s*{UNIT}\b", text)
    return single.group(1) if single else ""


def words(name):
    """Слова названия без бренда, фасовки и мусора."""
    text = re.sub(r"[^A-Z0-9]+", " ", str(name).upper())
    parts = []
    for part in text.split():
        if part in NOISE or re.fullmatch(rf"\d+(?:\.\d+)?{UNIT}?", part):
            continue
        parts.append(part)
    return parts


def keys(df, brand):
    """Ключ товара: набор слов названия плюс фасовка."""
    brand_words = set(words(brand))
    tokens, packs = [], []
    for _, row in df.iterrows():
        rest = [w for w in words(row["Название EN"]) if w not in brand_words]
        tokens.append(" ".join(sorted(set(rest))))
        packs.append(volume(row.get("Объем")) or volume(row["Название EN"]))
    return pd.Series(tokens, index=df.index), pd.Series(packs, index=df.index)


def load(brand, country):
    df = pd.read_excel(SRC, dtype={"Штрихкод": str})
    df = df[df["Страна"] == country]
    mask = (df["Бренд"].fillna("").astype(str).str.upper() == brand.upper()) | \
           df["Название EN"].fillna("").astype(str).str.upper().str.contains(brand.upper())
    rows = df[mask].copy()
    rows = rows[rows["Закупка, KRW"].notna() & (rows["Закупка, KRW"] > 0)]
    rows["Слова"], rows["Фасовка"] = keys(rows, brand)
    return rows[rows["Слова"] != ""]


def pair_up(rows, suppliers):
    """Сначала сводим по словам и фасовке, потом по одним словам.

    Второй проход нужен там, где фасовку пишет только один поставщик.
    Берем его, только если с каждой стороны остался ровно один кандидат:
    иначе непонятно, какой объем с каким сравнивать.
    """
    rows = rows.copy()
    rows["Ключ"] = rows["Слова"] + " | " + rows["Фасовка"]

    counts = rows.groupby("Ключ")["Поставщик"].nunique()
    matched = set(counts[counts > 1].index)

    rest = rows[~rows["Ключ"].isin(matched)]
    for name, group in rest.groupby("Слова"):
        sizes = group.groupby("Поставщик").size()
        if len(sizes) < 2 or (sizes > 1).any():
            continue
        # Разные объемы — это разные товары. Пробник 8 мл и банка 60 мл
        # называются одинаково, и по цене их сравнивать нельзя.
        known = {pack for pack in group["Фасовка"] if pack}
        if len(known) > 1:
            continue
        rows.loc[group.index, "Ключ"] = name + " | ~"
    return rows


def main(brand, country):
    rows = load(brand, country)
    if rows.empty:
        raise SystemExit(f"Бренд {brand} не найден в стране {country}")
    suppliers = sorted(rows["Поставщик"].unique())
    rows = pair_up(rows, suppliers)

    best = rows.sort_values("Закупка, KRW").drop_duplicates(["Ключ", "Поставщик"])
    prices = best.pivot(index="Ключ", columns="Поставщик", values="Закупка, KRW")
    info = (best.sort_values("Название EN", key=lambda s: s.astype(str).str.len(), ascending=False)
                .drop_duplicates("Ключ").set_index("Ключ")[["Название EN", "Объем", "Фасовка",
                                                            "MSRP, KRW", "Штрихкод"]])
    table = info.join(prices)

    enough = (prices.notna().sum(axis=1) > 1).reindex(table.index, fill_value=False)
    shared = table[enough].copy()
    columns = [s for s in suppliers if s in shared.columns]
    if shared.empty:
        raise SystemExit(f"Общих позиций по названию у поставщиков нет")

    shared["Дешевле у"] = shared[columns].idxmin(axis=1)
    shared["Мин, KRW"] = shared[columns].min(axis=1)
    shared["Макс, KRW"] = shared[columns].max(axis=1)
    shared["Дороже на, %"] = ((shared["Макс, KRW"] / shared["Мин, KRW"] - 1) * 100).round(1)
    shared["Экономия, руб"] = ((shared["Макс, KRW"] - shared["Мин, KRW"]) * RATE * IMPORT).round(0)
    shared["Себестоимость мин, руб"] = (shared["Мин, KRW"] * RATE * IMPORT).round(0)
    special = shared["Название EN"].fillna("").astype(str).str.upper().str.contains(SPECIAL)
    shared["Пометка"] = ""
    shared.loc[special, "Пометка"] = "лот или истекающий срок, цена особая"
    shared.loc[shared["Дороже на, %"] >= 30, "Пометка"] = "разница 30%+, сверить"
    shared = shared.sort_values("Дороже на, %", ascending=False)

    rest = table[~enough].copy()
    rest["Есть только у"] = rest[columns].apply(
        lambda row: row.dropna().index[0] if row.notna().any() else "", axis=1)

    total = [{"Показатель": "Сопоставлено по названию", "Значение": len(shared)}]
    for supplier in columns:
        column = shared[supplier]
        over = (column / shared["Мин, KRW"] - 1) * 100
        total.append({"Показатель": f"{supplier}: позиций всего",
                      "Значение": int(rows[rows["Поставщик"] == supplier]["Ключ"].nunique())})
        total.append({"Показатель": f"{supplier}: дешевле всех",
                      "Значение": int((shared["Дешевле у"] == supplier).sum())})
        total.append({"Показатель": f"{supplier}: медиана переплаты, %",
                      "Значение": round(over.median(), 1)})
    basket = shared[columns].dropna()
    if len(columns) == 2 and not basket.empty:
        first, second = columns
        cheap, rival = (first, second) if basket[first].sum() <= basket[second].sum() else (second, first)
        gap = (basket[rival].sum() / basket[cheap].sum() - 1) * 100
        total.append({"Показатель": f"Корзина {first}, KRW", "Значение": round(basket[first].sum())})
        total.append({"Показатель": f"Корзина {second}, KRW", "Значение": round(basket[second].sum())})
        total.append({"Показатель": "Корзина дешевле у", "Значение": cheap})
        total.append({"Показатель": f"{rival} дороже по корзине, %", "Значение": round(gap, 1)})
    summary = pd.DataFrame(total)

    slug = re.sub(r"[^a-z0-9]+", "_", brand.lower()).strip("_")
    out_path = os.path.join(OUT_DIR, f"brand_{slug}_{country.lower()}_by_name.xlsx")
    with pd.ExcelWriter(out_path) as writer:
        summary.to_excel(writer, sheet_name="ИТОГО", index=False)
        shared.to_excel(writer, sheet_name="СРАВНЕНИЕ")
        rest.to_excel(writer, sheet_name="ТОЛЬКО У ОДНОГО")

        book = writer.book
        for name in book.sheetnames:
            sheet = book[name]
            sheet.freeze_panes = "B2"
            titles = [str(c.value or "") for c in sheet[1]]
            for index, title in enumerate(titles, start=1):
                cell = sheet.cell(row=1, column=index)
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                letter = get_column_letter(index)
                sheet.column_dimensions[letter].width = (
                    60 if title in ("Ключ", "Название EN", "Показатель") else 15)
                fmt = "0.0" if "%" in title else ("# ##0" if ("KRW" in title or "руб" in title) else None)
                if fmt:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = fmt
            if "Дешевле у" in titles:
                winner = titles.index("Дешевле у")
                for row in sheet.iter_rows(min_row=2):
                    if row[winner].value in titles:
                        row[titles.index(row[winner].value)].fill = BEST_FILL

        sheet = book["СРАВНЕНИЕ"]
        titles = [str(c.value or "") for c in sheet[1]]
        letter = get_column_letter(titles.index("Дороже на, %") + 1)
        sheet.conditional_formatting.add(
            f"{letter}2:{letter}{len(shared) + 1}",
            ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                           mid_type="num", mid_value=10, mid_color="FFEB84",
                           end_type="num", end_value=30, end_color="F8696B"))

    print(f"{brand}, {country}: поставщики {', '.join(suppliers)}")
    print(summary.to_string(index=False))
    print(f"\nФайл: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Сравнение по названию, без штрихкода")
    parser.add_argument("brand")
    parser.add_argument("--country", default="KR")
    args = parser.parse_args()
    main(args.brand, args.country)
