from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
UNIT_FILE = ROOT / "data/unit_economics/Celimax_unit_model.xlsx"


def load_constants() -> dict:
    wb = openpyxl.load_workbook(UNIT_FILE, data_only=True)
    ws = wb["Юнит"]
    return {
        "handling": float(ws["B1"].value),
        "ads_default": float(ws["B2"].value),
        "storage_per_liter_day": float(ws["B6"].value),
        "storage_days": float(ws["B7"].value),
        "tax": float(ws["B8"].value),
        "reverse_logistics_base": float(ws["S3"].value),
        "commission": float(ws["S6"].value),
        "acquiring": float(ws["S7"].value),
        "spp": 0.30,
        "logistics": 90.0,
    }


def guess_buyout(name: str) -> float:
    low = name.lower()
    if any(word in low for word in ["пенк", "cleanser", "мицелляр", "гель для душа"]):
        return 96.0
    return 92.19


def calc(
    price_spp: float,
    cost: float,
    orders: float = 0.0,
    name: str = "",
    drr: float = 0.10,
    buyout: float | None = None,
    volume_liter: float = 1.0,
) -> dict:
    c = load_constants()
    buyout_percent = buyout if buyout is not None else guess_buyout(name)
    price_wb = round(price_spp / (1 - c["spp"]))
    price_spp_calc = price_wb * (1 - c["spp"])
    tax = price_spp_calc * c["tax"]
    commission = price_wb * c["commission"]
    acquiring = price_spp_calc * c["acquiring"]
    reverse = c["reverse_logistics_base"] * ((100 - buyout_percent) / 100)
    storage = c["storage_per_liter_day"] * c["storage_days"] * volume_liter
    ads = price_wb * drr

    margin_before_ads = (
        price_wb
        - c["handling"]
        - tax
        - c["logistics"]
        - commission
        - acquiring
        - reverse
        - storage
        - cost
    )
    margin_after_ads = margin_before_ads - ads
    roi_after_ads = margin_after_ads / cost * 100 if cost else None
    roi_before_ads = margin_before_ads / cost * 100 if cost else None

    return {
        "name": name,
        "price_spp": round(price_spp, 2),
        "price_wb_before_spp": price_wb,
        "cost": round(cost, 2),
        "drr_pct": round(drr * 100, 2),
        "buyout_pct": round(buyout_percent, 2),
        "handling": round(c["handling"], 2),
        "tax": round(tax, 2),
        "logistics": round(c["logistics"], 2),
        "commission": round(commission, 2),
        "acquiring": round(acquiring, 2),
        "reverse_logistics": round(reverse, 2),
        "storage": round(storage, 2),
        "ads": round(ads, 2),
        "margin_before_drr": round(margin_before_ads, 2),
        "roi_before_drr_pct": round(roi_before_ads, 1) if roi_before_ads is not None else None,
        "margin_after_drr": round(margin_after_ads, 2),
        "roi_after_drr_pct": round(roi_after_ads, 1) if roi_after_ads is not None else None,
        "profit_month_after_drr": round(margin_after_ads * orders, 2),
        "orders": orders,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="WB unit economics calculator")
    parser.add_argument("--price-spp", type=float, required=True, help="Цена с СПП")
    parser.add_argument("--cost", type=float, required=True, help="Себестоимость")
    parser.add_argument("--orders", type=float, default=0.0, help="Заказы в месяц")
    parser.add_argument("--name", default="", help="Название товара для определения выкупа")
    parser.add_argument("--drr", type=float, default=0.10, help="ДРР, например 0.10")
    parser.add_argument("--buyout", type=float, default=None, help="Процент выкупа, если нужно задать вручную")
    parser.add_argument("--volume-liter", type=float, default=1.0, help="Объем для хранения")
    args = parser.parse_args()

    result = calc(
        price_spp=args.price_spp,
        cost=args.cost,
        orders=args.orders,
        name=args.name,
        drr=args.drr,
        buyout=args.buyout,
        volume_liter=args.volume_liter,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
