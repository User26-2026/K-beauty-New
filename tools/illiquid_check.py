"""Что на складе действительно неликвид, а что просто перестало продаваться.

Считать неликвид по выкупам нельзя: продажи падают и когда товар никому
не нужен, и когда он пропал с полки — сгорел склад, кончился остаток,
карточка вылетела из выдачи. Это разные болезни и лечатся по-разному.

Поэтому смотрим воронку по двум периодам подряд и разделяем три случая:
покупатель заходит в карточку и не берет (спроса нет), покупатель заходил
и брал, а теперь конверсия рухнула (сбой), и покупатель просто перестал
заходить (карточка потеряла трафик).

Запас в месяцах считаем по лучшему из двух периодов: так провал последних
месяцев не превращает ходовой товар в мертвый.

Запуск:
    python3 tools/illiquid_check.py
"""

import argparse

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import name_match
from stock_by_brand import KNOWN, brand_from_name, clean_brand, read_plan, read_stock

FUNNEL = "workspace/wb_audit/sales_funnel_2026-03-01_2026-06-09.xlsx"
COSTS = "workspace/wb_audit/sales_profit_with_cogs_2026-04-01_2026-05-24.csv"
PERIOD_DAYS = 101          # столько дней в каждом из двух периодов отчета
MIN_CLICKS = 500           # меньше — трафика мало, судить о спросе рано
DEAD_CONVERSION = 1.0      # процент заказов от переходов, ниже которого спроса нет
OUT = "outputs/illiquid_check.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
FILLS = {
    "неликвид: смотрят и не берут": PatternFill("solid", fgColor="FFC7CE"),
    "затоварен: спрос есть, запаса на годы": PatternFill("solid", fgColor="FCE4D6"),
    "сбой: спрос был, конверсия упала": PatternFill("solid", fgColor="FFF2CC"),
    "сбой: карточка потеряла трафик": PatternFill("solid", fgColor="FFF2CC"),
    "ходовой, запас кончается": PatternFill("solid", fgColor="C6EFCE"),
}

NOW = {
    "clicks": "Переходы в карточку",
    "orders": "Заказали товаров, шт",
    "stock_wb": "Остатки «Склад WB», шт",
}
WAS = {
    "clicks": "Переходы в карточку (предыдущий период)",
    "orders": "Заказали товаров, шт (предыдущий период)",
}


def read_funnel(path):
    funnel = pd.read_excel(path, sheet_name="Товары", header=1)
    funnel["Артикул WB"] = pd.to_numeric(funnel["Артикул WB"], errors="coerce").astype("Int64")
    columns = list(NOW.values()) + list(WAS.values())
    for column in columns:
        funnel[column] = pd.to_numeric(funnel[column], errors="coerce").fillna(0)
    return funnel[["Артикул WB", "Бренд"] + columns].rename(columns={"Бренд": "Бренд из отчета"})


def verdict(row):
    """Один диагноз на позицию — по трафику, конверсии и запасу."""
    clicks_now, clicks_was = row["Переходов сейчас"], row["Переходов раньше"]
    conv_now, conv_was = row["Конверсия сейчас, %"], row["Конверсия раньше, %"]
    cover = row["Запас, месяцев"]

    if max(clicks_now, clicks_was) < MIN_CLICKS:
        return "мало трафика, судить рано"

    # Спроса нет: покупатель заходит в карточку и не покупает, и так было
    # в оба периода. Наличие тут ни при чем — товар просто не нужен.
    weak_now = pd.isna(conv_now) or conv_now < DEAD_CONVERSION
    weak_was = pd.isna(conv_was) or conv_was < DEAD_CONVERSION
    if weak_now and weak_was:
        return "неликвид: смотрят и не берут"

    # Раньше брали, сейчас нет — это сбой: пропало наличие, выросла цена,
    # уехали склады. Товар не виноват.
    if conv_was and conv_was >= 2 and conv_now is not None and conv_now < conv_was / 2:
        return "сбой: спрос был, конверсия упала"
    if clicks_was >= MIN_CLICKS and clicks_now < clicks_was * 0.4:
        return "сбой: карточка потеряла трафик"

    if pd.isna(cover):
        return "мало трафика, судить рано"
    if cover > 12:
        return "затоварен: спрос есть, запаса на годы"
    if cover > 6:
        return "избыток запаса"
    if cover >= 2:
        return "нормальный запас"
    return "ходовой, запас кончается"


def main():
    plan = read_plan("data/shipments/2026-09_plan_raspredeleniya.xlsx")
    stock = read_stock("data/stock_costs/Остатки_31.08.2026.xlsx")
    funnel = read_funnel(FUNNEL)

    table = plan.merge(funnel, on="Артикул WB", how="left")
    pairs = name_match.match(table["Наименование"], stock["Товар в остатках"])
    table["_pair"] = table.index.map(pairs.get)
    table = table.merge(stock, left_on="_pair", right_index=True, how="left").drop(columns="_pair")

    # Себестоимость берем из файла остатков, а где товара там нет — из
    # отчета WB: иначе позиция выпадает из денег и картина занижается.
    from_report = (pd.read_csv(COSTS, usecols=["Артикул WB", "cost"])
                   .assign(**{"Артикул WB": lambda frame: pd.to_numeric(
                       frame["Артикул WB"], errors="coerce").astype("Int64")})
                   .groupby("Артикул WB", as_index=False)["cost"].max())
    table = table.merge(from_report, on="Артикул WB", how="left")
    table["Себестоимость, руб"] = pd.to_numeric(
        table["Себестоимость из остатков"], errors="coerce").fillna(
        pd.to_numeric(table["cost"], errors="coerce"))
    table["Остаток, руб"] = (table["Склад, шт"] * table["Себестоимость, руб"]).round(0)
    table["Бренд"] = table["Бренд из отчета"].map(clean_brand)
    table.loc[~table["Бренд"].isin(KNOWN), "Бренд"] = \
        table.loc[~table["Бренд"].isin(KNOWN), "Наименование"].map(brand_from_name)

    table["Переходов сейчас"] = table[NOW["clicks"]]
    table["Переходов раньше"] = table[WAS["clicks"]]
    table["Заказов сейчас, шт"] = table[NOW["orders"]]
    table["Заказов раньше, шт"] = table[WAS["orders"]]
    table["Остаток на складе WB, шт"] = table[NOW["stock_wb"]]

    table["Конверсия сейчас, %"] = (table["Заказов сейчас, шт"] /
                                    table["Переходов сейчас"].replace(0, np.nan) * 100).round(2)
    table["Конверсия раньше, %"] = (table["Заказов раньше, шт"] /
                                    table["Переходов раньше"].replace(0, np.nan) * 100).round(2)

    # Потенциал спроса — по лучшему периоду: провал последних месяцев не
    # должен записывать ходовой товар в мертвый.
    per_month = pd.concat([table["Заказов сейчас, шт"], table["Заказов раньше, шт"]],
                          axis=1).max(axis=1) * 30 / PERIOD_DAYS
    table["Спрос, шт/мес"] = per_month.round(1)
    table["Запас, месяцев"] = (table["Склад, шт"] /
                               table["Спрос, шт/мес"].replace(0, np.nan)).round(1)
    table["Вердикт"] = table.apply(verdict, axis=1)
    table["Приход в пути, шт"] = (pd.to_numeric(table["Машина"], errors="coerce").fillna(0)
                                  + pd.to_numeric(table["Контейнер"], errors="coerce").fillna(0))

    columns = ["Бренд", "Наименование", "Артикул WB", "Склад, шт", "Остаток, руб",
               "Остаток на складе WB, шт", "Переходов раньше", "Переходов сейчас",
               "Заказов раньше, шт", "Заказов сейчас, шт", "Конверсия раньше, %",
               "Конверсия сейчас, %", "Спрос, шт/мес", "Запас, месяцев", "Вердикт",
               "Приход в пути, шт"]
    items = table[columns].sort_values("Остаток, руб", ascending=False)

    by_verdict = (items.groupby("Вердикт")
                  .agg(**{"Позиций": ("Наименование", "size"),
                          "Штук": ("Склад, шт", "sum"),
                          "Деньги, руб": ("Остаток, руб", "sum"),
                          "Спрос, шт/мес": ("Спрос, шт/мес", "sum")})
                  .reset_index().sort_values("Деньги, руб", ascending=False))
    by_verdict["Доля денег, %"] = (by_verdict["Деньги, руб"] /
                                   by_verdict["Деньги, руб"].sum() * 100).round(1)

    by_brand = (items.groupby("Бренд")
                .agg(**{"Позиций": ("Наименование", "size"),
                        "Деньги, руб": ("Остаток, руб", "sum")})
                .reset_index())
    dead = items[items["Вердикт"] == "неликвид: смотрят и не берут"]
    over = items[items["Вердикт"] == "затоварен: спрос есть, запаса на годы"]
    broken = items[items["Вердикт"].str.startswith("сбой")]
    by_brand["Неликвид, руб"] = by_brand["Бренд"].map(
        dead.groupby("Бренд")["Остаток, руб"].sum()).fillna(0)
    by_brand["Затоварено, руб"] = by_brand["Бренд"].map(
        over.groupby("Бренд")["Остаток, руб"].sum()).fillna(0)
    by_brand["Сбой продаж, руб"] = by_brand["Бренд"].map(
        broken.groupby("Бренд")["Остаток, руб"].sum()).fillna(0)
    by_brand = by_brand.sort_values("Деньги, руб", ascending=False)

    total = pd.DataFrame([
        {"Показатель": "Позиций", "Значение": len(items)},
        {"Показатель": "Остаток по себестоимости, руб", "Значение": round(items["Остаток, руб"].sum())},
        {"Показатель": "НЕЛИКВИД: позиций", "Значение": len(dead)},
        {"Показатель": "НЕЛИКВИД: денег, руб", "Значение": round(dead["Остаток, руб"].sum())},
        {"Показатель": "ЗАТОВАРЕНО: позиций", "Значение": len(over)},
        {"Показатель": "ЗАТОВАРЕНО: денег, руб", "Значение": round(over["Остаток, руб"].sum())},
        {"Показатель": "СБОЙ ПРОДАЖ: позиций", "Значение": len(broken)},
        {"Показатель": "СБОЙ ПРОДАЖ: денег, руб", "Значение": round(broken["Остаток, руб"].sum())},
        {"Показатель": "Позиций с нулем на складе WB",
         "Значение": int((items["Остаток на складе WB, шт"] == 0).sum())},
        {"Показатель": "Длина периода отчета, дней", "Значение": PERIOD_DAYS},
    ])

    with pd.ExcelWriter(OUT) as writer:
        total.to_excel(writer, sheet_name="ИТОГО", index=False)
        by_verdict.to_excel(writer, sheet_name="ПО ВЕРДИКТАМ", index=False)
        by_brand.to_excel(writer, sheet_name="ПО БРЕНДАМ", index=False)
        dead.to_excel(writer, sheet_name="НЕЛИКВИД", index=False)
        over.to_excel(writer, sheet_name="ЗАТОВАРЕНО", index=False)
        broken.to_excel(writer, sheet_name="СБОЙ ПРОДАЖ", index=False)
        items.to_excel(writer, sheet_name="ВСЕ ПОЗИЦИИ", index=False)

        for name in writer.book.sheetnames:
            sheet = writer.book[name]
            sheet.freeze_panes = "B2"
            titles = [str(c.value or "") for c in sheet[1]]
            for index, title in enumerate(titles, start=1):
                cell = sheet.cell(row=1, column=index)
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                sheet.column_dimensions[get_column_letter(index)].width = (
                    58 if title in ("Наименование", "Показатель") else
                    36 if title == "Вердикт" else 22 if title == "Бренд" else 13)
                if "руб" in title or "шт" in title:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = "# ##0"
                if "%" in title:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = "0.00"
            if "Вердикт" in titles:
                column = titles.index("Вердикт")
                for row in sheet.iter_rows(min_row=2):
                    fill = FILLS.get(row[column].value)
                    if fill:
                        row[column].fill = fill

    print(total.to_string(index=False))
    print("\nПо вердиктам:")
    print(by_verdict.to_string(index=False))
    print(f"\nФайл: {OUT}")


if __name__ == "__main__":
    argparse.ArgumentParser(description="Неликвид против сбоя продаж").parse_args()
    main()
