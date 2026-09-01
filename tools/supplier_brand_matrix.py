"""Где какой бренд закупать: все поставщики против всех брендов.

Сравниваем только внутри одной страны и только по штрихкоду. Цена — за
штуку, фасовку выравниваем по штрихкоду. Проценты считаем от самого
дешевого предложения по позиции: 0% у того, кто дешевле всех.

Запуск:
    python3 tools/supplier_brand_matrix.py            # Корея
    python3 tools/supplier_brand_matrix.py --country RU
"""

import argparse
import os
import re

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import brand_names
from price_unit import unify_packs

SRC = "outputs/prices_normalized.xlsx"
OUT_DIR = "outputs"
RATE = 0.058          # рублей за вону
IMPORT = 1.4          # логистика, пошлина и приемка
MIN_COVERAGE = 0.5    # лидером бренда считаем только того, у кого есть половина позиций

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
LEADER_FILL = PatternFill("solid", fgColor="C6EFCE")
ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")

# Слова, которые есть в названии любого товара и о совпадении ничего не говорят.
NOISE = {"ML", "GR", "EA", "PCS", "SET", "KIT", "NEW", "THE", "AND", "FOR",
         "SKIN", "CARE", "CREAM", "TONER", "SERUM", "MASK", "PACK"}


def load(country):
    df = pd.read_excel(SRC, dtype={"Штрихкод": str})
    df = df[df["Страна"] == country].copy()
    df["Штрихкод"] = df["Штрихкод"].fillna("").astype(str).str.replace(r"\D", "", regex=True)
    df = df[df["Штрихкод"].str.len() >= 8]
    df = df[df["Закупка, KRW"].notna() & (df["Закупка, KRW"] > 0)]
    df["Бренд в прайсе"] = df["Бренд"]
    df["Бренд"] = brand_names.resolve(df).fillna("БЕЗ БРЕНДА")
    return df


def tokens(name):
    """Значимые слова названия: без объемов, фасовки и общих слов."""
    parts = re.split(r"[^A-Z0-9]+", str(name).upper())
    return {p for p in parts if len(p) >= 3 and p not in NOISE
            and not p.isdigit() and not re.fullmatch(r"\d+[A-Z]{1,4}", p)}


def split_conflicts(df):
    """Разводит товары, которым поставщики дали один штрихкод.

    Так вскрывается копипаста в прайсе: у Papa Cosmetic в списке SKIN1004
    стоят штрихкоды из списка ROUND LAB. Оставляем ту группу строк, где
    товар подтвержден большим числом поставщиков, остальные выносим на
    сверку — иначе из-за одного кривого прайса теряется весь бренд.
    """
    dropped, report = [], []
    for ean, group in df.groupby("Штрихкод"):
        if group["Поставщик"].nunique() < 2:
            continue
        named = {index: tokens(name) for index, name in group["Название EN"].items()}
        clusters = []
        for index, words in named.items():
            # Одно слово — не улика: у части поставщиков название обрезано.
            if len(words) < 2:
                continue
            for cluster in clusters:
                if any(words & named[other] for other in cluster):
                    cluster.append(index)
                    break
            else:
                clusters.append([index])
        if len(clusters) < 2:
            continue
        # Побеждает та версия товара, которую подтверждает больше поставщиков.
        clusters.sort(key=lambda rows: (df.loc[rows, "Поставщик"].nunique(), len(rows)),
                      reverse=True)
        main = set(clusters[0])
        strangers = [index for cluster in clusters[1:] for index in cluster]
        dropped.extend(strangers)
        report.append({
            "Штрихкод": ean,
            "Оставили товар": df.loc[clusters[0][0], "Название EN"],
            "Подтверждают": ", ".join(sorted(df.loc[clusters[0], "Поставщик"].unique())),
            "Убрали товар": df.loc[strangers[0], "Название EN"],
            "Чей прайс": ", ".join(sorted(df.loc[strangers, "Поставщик"].unique())),
            "Бренд убранного": ", ".join(sorted(set(
                df.loc[strangers, "Бренд в прайсе"].dropna().astype(str)))),
        })
    return df.drop(index=dropped), pd.DataFrame(report)


def price_table(df, column):
    """Одна строка на штрихкод, колонка на поставщика."""
    best = df.sort_values(column).drop_duplicates(["Штрихкод", "Поставщик"])
    prices = best.pivot(index="Штрихкод", columns="Поставщик", values=column)
    info = (best.sort_values("Название EN", key=lambda s: s.astype(str).str.len(), ascending=False)
                .drop_duplicates("Штрихкод")
                .set_index("Штрихкод")[["Бренд", "Название EN", "Объем"]])
    return info.join(prices), prices


def brand_stats(prices, info, suppliers):
    """По каждому бренду: у кого дешевле и на сколько процентов дороже остальные."""
    comparable = prices[prices.notna().sum(axis=1) > 1]
    rows, matrix, baskets = [], [], []
    for brand, index in info.loc[comparable.index].groupby("Бренд").groups.items():
        block = comparable.loc[index]
        cheapest = block.min(axis=1)
        record = {"Бренд": brand, "Позиций в сравнении": len(block)}
        overpays, coverage, loss = {}, {}, {}
        for supplier in suppliers:
            column = block.get(supplier)
            if column is None or column.notna().sum() == 0:
                continue
            present = column.notna()
            over = (column[present] / cheapest[present] - 1) * 100
            overpays[supplier] = round(over.median(), 1)
            coverage[supplier] = present.sum() / len(block)
            loss[supplier] = float((column[present] - cheapest[present]).sum())

        # Лидер — кто дешевле по медиане, но только если у него есть половина позиций.
        wide = {s: v for s, v in overpays.items() if coverage[s] >= MIN_COVERAGE}
        pool = wide or overpays
        leader = min(pool, key=lambda s: (pool[s], -coverage[s]))
        rivals = {s: v for s, v in overpays.items() if s != leader}
        second = min(rivals, key=rivals.get) if rivals else None

        # Медиана может совпасть у двоих. Тогда решает корзина: считаем ее
        # по позициям, которые есть и у лидера, и у соперника.
        basket_gap = None
        if second:
            both = block[leader].notna() & block[second].notna()
            if both.any() and block.loc[both, leader].sum():
                basket_gap = round(
                    (block.loc[both, second].sum() / block.loc[both, leader].sum() - 1) * 100, 1)

        record.update({
            "Поставщиков": len(overpays),
            "Дешевле всех": leader,
            "Покрытие лидера, %": round(coverage[leader] * 100),
            "Второй по цене": second,
            "Второй дороже, % (медиана)": rivals[second] if second else None,
            "Второй дороже, % (корзина)": basket_gap,
            "Средняя переплата у остальных, %": (
                round(sum(rivals.values()) / len(rivals), 1) if rivals else 0.0),
            "Переплата, если брать не у лидера, руб": round(
                max(loss.values()) * RATE * IMPORT) if loss else 0,
        })
        rows.append(record)
        matrix.append({"Бренд": brand, **overpays})
        baskets.append({"Бренд": brand, **{s: round(v * RATE * IMPORT) for s, v in loss.items()}})

    order = ["Бренд", "Позиций в сравнении", "Поставщиков", "Дешевле всех",
             "Покрытие лидера, %", "Второй по цене", "Второй дороже, % (медиана)",
             "Второй дороже, % (корзина)",
             "Средняя переплата у остальных, %", "Переплата, если брать не у лидера, руб"]
    brands = pd.DataFrame(rows)[order].sort_values("Позиций в сравнении", ascending=False)
    percent = (pd.DataFrame(matrix).set_index("Бренд")
                 .reindex(index=brands["Бренд"], columns=suppliers).reset_index())
    money = (pd.DataFrame(baskets).set_index("Бренд")
               .reindex(index=brands["Бренд"], columns=suppliers).reset_index())
    return brands, percent, money


def supplier_stats(df, prices, info, brands, suppliers):
    """Свод по поставщикам: охват, победы и медиана переплаты."""
    comparable = prices[prices.notna().sum(axis=1) > 1]
    cheapest = comparable.min(axis=1)
    leaders = brands.groupby("Дешевле всех")["Бренд"].count()
    rows = []
    for supplier in suppliers:
        column = comparable.get(supplier)
        present = column.notna() if column is not None else pd.Series(dtype=bool)
        over = ((column[present] / cheapest[present] - 1) * 100) if present.any() else pd.Series(dtype=float)
        own = df[df["Поставщик"] == supplier]
        rows.append({
            "Поставщик": supplier,
            "Брендов": own["Бренд"].nunique(),
            "SKU со штрихкодом": own["Штрихкод"].nunique(),
            "Позиций в сравнении": int(present.sum()),
            "Дешевле всех, позиций": int((over == 0).sum()),
            "Дешевле всех, % позиций": round((over == 0).mean() * 100, 1) if len(over) else 0.0,
            "Медиана переплаты, %": round(over.median(), 1) if len(over) else None,
            "Лидер по брендам, шт": int(leaders.get(supplier, 0)),
        })
    return pd.DataFrame(rows).sort_values("Медиана переплаты, %")


def exclusive(df, prices, suppliers):
    """Бренды, которые есть только у одного поставщика — сравнивать не с кем."""
    counts = df.groupby(["Бренд", "Поставщик"])["Штрихкод"].nunique().unstack(fill_value=0)
    counts = counts.reindex(columns=suppliers, fill_value=0)
    single = counts[(counts > 0).sum(axis=1) == 1]
    rows = [{"Бренд": brand,
             "Есть только у": row[row > 0].index[0],
             "SKU": int(row.max())} for brand, row in single.iterrows()]
    return pd.DataFrame(rows).sort_values("SKU", ascending=False)


def positions_sheet(table, prices, suppliers):
    comparable = prices.notna().sum(axis=1) > 1
    rows = table[comparable.reindex(table.index, fill_value=False)].copy()
    columns = [s for s in suppliers if s in rows.columns]
    rows["Мин, KRW"] = rows[columns].min(axis=1)
    rows["Макс, KRW"] = rows[columns].max(axis=1)
    rows["Дешевле у"] = rows[columns].idxmin(axis=1)
    rows["Разброс, %"] = ((rows["Макс, KRW"] / rows["Мин, KRW"] - 1) * 100).round(1)
    rows["Экономия, руб"] = ((rows["Макс, KRW"] - rows["Мин, KRW"]) * RATE * IMPORT).round(0)
    rows["Проверить у менеджера"] = rows["Разброс, %"].ge(30).map({True: "да", False: ""})
    return rows.sort_values(["Бренд", "Разброс, %"], ascending=[True, False])


def format_sheet(worksheet, wide=("Бренд", "Название EN")):
    worksheet.freeze_panes = "B2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    titles = [str(c.value or "") for c in worksheet[1]]
    for index, title in enumerate(titles, start=1):
        letter = get_column_letter(index)
        if title == "Название EN":
            width = 50
        elif title in wide or title in ("Дешевле всех", "Дешевле у", "Второй по цене",
                                        "Есть только у", "Поставщик", "Поставщики"):
            width = 22
        elif title == "Названия":
            width = 70
        else:
            width = 13
        worksheet.column_dimensions[letter].width = width
        fmt = "0.0" if "%" in title else ("# ##0" if ("руб" in title or "KRW" in title) else None)
        if fmt:
            for row in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
                row[0].number_format = fmt


def paint(worksheet, columns, last_row, top=25):
    """Зеленый — самая низкая цена, красный — переплата."""
    for column in columns:
        letter = get_column_letter(column)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{last_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                           mid_type="num", mid_value=top / 3, mid_color="FFEB84",
                           end_type="num", end_value=top, end_color="F8696B"))


def main(country):
    if not os.path.exists(SRC):
        raise SystemExit(f"Нет файла {SRC} — сначала запустите parse_price_lists.py")

    df = load(country)
    suppliers = sorted(df["Поставщик"].unique())

    df, bad = split_conflicts(df)
    if not bad.empty:
        print(f"Один штрихкод на разные товары: {len(bad)} — спорные строки убраны")

    df, conflicts = unify_packs(df)
    if conflicts:
        print(f"Разная фасовка по одному штрихкоду: {len(conflicts)} позиций")

    column = "Цена за штуку (сводно)"
    table, prices = price_table(df, column)
    info = table[["Бренд", "Название EN", "Объем"]]

    brands, percent, money = brand_stats(prices, info, suppliers)
    summary = supplier_stats(df, prices, info, brands, suppliers)
    only = exclusive(df, prices, suppliers)
    positions = positions_sheet(table, prices, suppliers)

    sku = (df.groupby(["Бренд", "Поставщик"])["Штрихкод"].nunique().unstack(fill_value=0)
             .reindex(columns=suppliers, fill_value=0))
    # Всего SKU — это объединение по штрихкодам, а не сумма прайсов.
    sku["Всего SKU"] = df.groupby("Бренд")["Штрихкод"].nunique()
    sku = sku.sort_values("Всего SKU", ascending=False).reset_index()

    out_path = os.path.join(OUT_DIR, f"supplier_brand_matrix_{country.lower()}.xlsx")
    with pd.ExcelWriter(out_path) as writer:
        summary.to_excel(writer, sheet_name="ИТОГО", index=False)
        brands.to_excel(writer, sheet_name="ГДЕ ЗАКУПАТЬ", index=False)
        percent.to_excel(writer, sheet_name="МАТРИЦА %", index=False)
        money.to_excel(writer, sheet_name="ПЕРЕПЛАТА РУБ", index=False)
        sku.to_excel(writer, sheet_name="SKU ПО ПОСТАВЩИКАМ", index=False)
        positions.to_excel(writer, sheet_name="ПОЗИЦИИ")
        only.to_excel(writer, sheet_name="ТОЛЬКО У ОДНОГО", index=False)
        if not bad.empty:
            bad.to_excel(writer, sheet_name="СПОРНЫЕ ШТРИХКОДЫ", index=False)

        book = writer.book
        for name in book.sheetnames:
            format_sheet(book[name])

        sheet = book["МАТРИЦА %"]
        paint(sheet, range(2, len(percent.columns) + 1), len(percent) + 1)
        # Лидера подсвечиваем зеленым прямо в списке брендов.
        leaders = book["ГДЕ ЗАКУПАТЬ"]
        titles = [str(c.value or "") for c in leaders[1]]
        for row in leaders.iter_rows(min_row=2):
            row[titles.index("Дешевле всех")].fill = LEADER_FILL
        paint(leaders, [titles.index("Второй дороже, % (медиана)") + 1,
                        titles.index("Второй дороже, % (корзина)") + 1,
                        titles.index("Средняя переплата у остальных, %") + 1], len(brands) + 1)
        paint(book["ПОЗИЦИИ"], [[str(c.value or "") for c in book["ПОЗИЦИИ"][1]].index("Разброс, %") + 1],
              len(positions) + 1, top=50)
        if not bad.empty:
            for row in book["СПОРНЫЕ ШТРИХКОДЫ"].iter_rows(min_row=2, max_col=1):
                row[0].fill = ALERT_FILL

    print(f"\nСтрана {country}, поставщики: {', '.join(suppliers)}")
    print(summary.to_string(index=False))
    print("\nГде закупать, топ брендов по числу сравнимых позиций:")
    print(brands.head(25).to_string(index=False))
    print(f"\nФайл: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Матрица поставщики х бренды")
    parser.add_argument("--country", default="KR")
    main(parser.parse_args().country)
