"""Монитор вложений в товар.

Считает сколько денег заморожено в товаре на всех стадиях и сравнивает с лимитом,
чтобы не допускать перевложения:

    Всего в товаре = Заказано за границей (оплачено) + Склад + Товар на маркетах

Основной источник — ручной срез в CONFIG["snapshot"] (владелец обновляет цифры
по банкам / складу / ВБ). Выгрузки остатков и отчёт ВБ по API используются как
ПЕРЕПРОВЕРКА (команда --crosscheck): считают склад и стоимость товара на ВБ из
файлов и показывают расхождение с ручным срезом.

Запуск:
    python tools/monitor_investment.py                 # дашборд по ручному срезу
    python tools/monitor_investment.py --report         # + запись в outputs/
    python tools/monitor_investment.py --crosscheck     # сверка с файлами остатков

Зависимости: openpyxl (xlsx), xlrd (для .xls отчёта ВБ, только для --crosscheck).
"""
from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
STOCK_DIR = ROOT / "data/stock_costs"
OUT_DIR = ROOT / "outputs"

# --- Актуальный срез. Обновляй эти цифры руками. ----------------------------

CONFIG = {
    "snapshot": {
        "abroad_paid": 26_000_000,       # заказано за границей и ОПЛАЧЕНО, руб
        "sklad": 61_000_000,             # товар на складе (себестоимость), руб
        "marketplace_goods": 13_600_000,  # товар на маркетах (себестоимость), руб
        "wb_balance": 14_000_000,        # деньги на маркетах (баланс ВБ), руб
        "wb_withdraw_pct": 0.30,         # вывод с ВБ в понедельник, доля
        "cash_bank": 12_000_000,         # остатки на счетах юр лиц, руб
        "limit": 100_000_000,           # лимит вложений в товар, руб
        "warn_ratio": 0.80,              # жёлтая зона: >=80% лимита
    },

    # --- Источники для перепроверки (--crosscheck) --------------------------

    # Заказано за границей — для пересчёта валют в рубли (справочно).
    "abroad_orders": [
        {"label": "Корея (вон)", "amount": 790_000_000, "currency": "KRW",
         "krw_per_usd": 1500, "rub_per_usd": 82},
        {"label": "Китай (USD)", "amount": 67_000, "currency": "USD",
         "rub_per_usd": 82},
    ],

    # Ручные выгрузки остатков. Колонки ищем по ЗАГОЛОВКУ (файлы еженедельно
    # меняют раскладку), берём готовые суммы «СКЛАД» и «Маркеты» из блока
    # «Сумма остатков». Реальные SKU-строки = где заполнена себестоимость.
    "stock_files": [
        {
            "label": "Корея (косметика)",
            "path": STOCK_DIR / "ostatki_korea_full.xlsx",
            "sheet": None,
            "use_marketplace": False,   # маркеты Кореи берём из отчёта ВБ по API
        },
        {
            "label": "Китай (товары)",
            "path": STOCK_DIR / "ostatki_china_full.xlsx",
            "sheet": None,
            "use_marketplace": True,    # у Китая живого отчёта нет
        },
    ],

    # Живой отчёт ВБ по API (косметика). col27 = "Сумма остатка в себестоимости
    # API". Пропуски себестоимости (кабинет КРОНА) заполняем по Артикулу ВБ из
    # ручной таблицы. Кабинет Скляров не работает — не учитываем.
    "wb_api": {
        "label": "ВБ косметика (API)",
        "path": STOCK_DIR / "wb_api_report.xls",
        "art_col": 0, "stock_col": 7, "unit_cost_col": 18, "cost_sum_col": 27,
        "cabinet_names": ("Коротич", "КРОНА", "Скляров"),
        "cost_source": STOCK_DIR / "ostatki_korea_full.xlsx",
        # артикулы ВБ в таблице-источнике: Коротич/Скляров/Крона
        "cost_source_art_cols": (2, 3, 4),
    },
}


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", "").replace(" ", "").replace("%", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def is_num(v) -> bool:
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


def rub(x: float) -> str:
    return f"{x:,.0f}".replace(",", " ") + " ₽"


# --- Дашборд по ручному срезу ------------------------------------------------

def build(snap: dict) -> dict:
    abroad = snap["abroad_paid"]
    sklad = snap["sklad"]
    mkt = snap["marketplace_goods"]
    total = abroad + sklad + mkt

    wb_liquid = snap["wb_balance"] * snap["wb_withdraw_pct"]
    cash_live = snap["cash_bank"] + wb_liquid

    limit = snap["limit"]
    free = limit - total
    if total > limit:
        flag = "🔴 СТОП — перевложение"
    elif total >= limit * snap["warn_ratio"]:
        flag = "🟡 Осторожно — запас < 20%"
    else:
        flag = "🟢 ОК"

    return {
        "abroad": abroad, "sklad": sklad, "mkt": mkt, "total": total,
        "limit": limit, "free": free, "flag": flag,
        "cash_bank": snap["cash_bank"], "wb_balance": snap["wb_balance"],
        "wb_liquid": wb_liquid, "cash_live": cash_live,
        "coverage": cash_live / total if total else 0.0,
    }


def render(r: dict) -> str:
    L = []
    a = L.append
    a(f"# Монитор вложений в товар — {dt.date.today().isoformat()}\n")
    a(f"## {r['flag']}\n")
    a("## Всего в товаре\n")
    a("| Стадия | Сумма |")
    a("|---|---|")
    a(f"| 🌍 Заказано за границей (оплачено) | {rub(r['abroad'])} |")
    a(f"| 📦 На складе | {rub(r['sklad'])} |")
    a(f"| 🛒 На маркетах (товар) | {rub(r['mkt'])} |")
    a(f"| **💰 ВСЕГО В ТОВАРЕ** | **{rub(r['total'])}** |")
    a(f"| 🎯 Лимит | {rub(r['limit'])} |")
    label = "Свободный лимит" if r["free"] >= 0 else "Превышение лимита"
    a(f"| **{label}** | **{rub(r['free'])}** |\n")

    a("## Свободные деньги\n")
    a("| Источник | Сумма |")
    a("|---|---|")
    a(f"| Счета юр лиц | {rub(r['cash_bank'])} |")
    a(f"| Деньги на ВБ (баланс) | {rub(r['wb_balance'])} |")
    a(f"| — из них вывод в понедельник (30%) | {rub(r['wb_liquid'])} |")
    a(f"| **Живой кэш (счета + вывод с ВБ)** | **{rub(r['cash_live'])}** |")
    a(f"| Покрытие товара живым кэшом | {r['coverage']*100:.0f}% |\n")
    return "\n".join(L)


# --- Перепроверка из файлов --------------------------------------------------

def _pick_sheet(wb, requested):
    if requested:
        return wb[requested]
    for name in reversed(wb.sheetnames):
        if "копия" not in name.lower():
            return wb[name]
    return wb[wb.sheetnames[-1]]


def _find_col(rows, match, maxrow=6):
    """Индекс колонки, чей заголовок в первых строках проходит проверку match()."""
    for r in rows[:maxrow]:
        for i, v in enumerate(r):
            if v is not None and match(str(v).strip()):
                return i
    return None


def _cost_col(rows):
    # «Себестоимость по последнему приходу» (Корея) или «Себестоимость» (Китай),
    # но не «Себестоимость средняя карго».
    return _find_col(rows, lambda s: s.lower().startswith("себестоимость")
                     and "карго" not in s.lower() and "средн" not in s.lower())


def _analyze_stock(cfg):
    ws = _pick_sheet(openpyxl.load_workbook(cfg["path"], data_only=True), cfg["sheet"])
    rows = list(ws.iter_rows(values_only=True))
    cc = _cost_col(rows)
    sc = _find_col(rows, lambda s: s.upper() == "СКЛАД")
    mc = _find_col(rows, lambda s: s.upper() == "МАРКЕТЫ")
    sklad = mkt = 0.0
    for r in rows:
        # реальная SKU-строка = заполнена себестоимость (подытоги/заголовки без неё)
        if cc is None or num(r[cc] if cc < len(r) else None) <= 0:
            continue
        if sc is not None and sc < len(r):
            sklad += num(r[sc])
        if mc is not None and mc < len(r):
            mkt += num(r[mc])
    return {"label": cfg["label"], "sheet": ws.title, "sklad": sklad,
            "mkt": mkt, "use_marketplace": cfg["use_marketplace"]}


def _art_cost_map(path, art_cols=(2, 3, 4)):
    ws = _pick_sheet(openpyxl.load_workbook(path, data_only=True), None)
    rows = list(ws.iter_rows(values_only=True))
    cc = _cost_col(rows)
    m = {}
    if cc is None:
        return m
    for r in rows:
        cost = num(r[cc]) if cc < len(r) else 0.0
        if cost <= 0:
            continue
        for ac in art_cols:
            a = r[ac] if ac < len(r) else None
            if a is not None and is_num(a):
                m[str(int(float(a)))] = cost
    return m


def _analyze_wb_api(cfg):
    import xlrd
    sh = xlrd.open_workbook(cfg["path"]).sheet_by_index(0)
    cost_map = _art_cost_map(cfg["cost_source"], cfg.get("cost_source_art_cols", (2, 3, 4)))
    val = units_missing = 0.0
    for r in range(sh.nrows):
        a = sh.cell_value(r, cfg["art_col"])
        if isinstance(a, str) and a.strip() in cfg["cabinet_names"]:
            continue
        if not (is_num(a) and str(a).strip()):
            continue
        stock = num(sh.cell_value(r, cfg["stock_col"]))
        unit_cost = num(sh.cell_value(r, cfg["unit_cost_col"]))
        if unit_cost > 0:
            val += num(sh.cell_value(r, cfg["cost_sum_col"]))
        else:
            c = cost_map.get(str(int(float(a))))
            if c:
                val += c * stock
            else:
                units_missing += stock
    return {"mkt": val, "units_missing": units_missing}


def _order_to_rub(o):
    if o["currency"] == "USD":
        return o["amount"] * o["rub_per_usd"]
    if o["currency"] == "KRW":
        return o["amount"] / o["krw_per_usd"] * o["rub_per_usd"]
    raise ValueError(o["currency"])


def crosscheck(cfg) -> str:
    snap = cfg["snapshot"]
    stocks = [_analyze_stock(sf) for sf in cfg["stock_files"]]
    wb_api = _analyze_wb_api(cfg["wb_api"])
    f_sklad = sum(s["sklad"] for s in stocks)
    f_mkt = wb_api["mkt"] + sum(s["mkt"] for s in stocks if s["use_marketplace"])
    f_abroad = sum(_order_to_rub(o) for o in cfg["abroad_orders"])
    L = ["## Перепроверка из файлов (справочно)\n",
         "| Стадия | Ручной срез | Из файлов | Δ |", "|---|---|---|---|"]
    def line(name, manual, fromfile):
        L.append(f"| {name} | {rub(manual)} | {rub(fromfile)} | {rub(fromfile-manual)} |")
    line("Заказано за границей (полный заказ)", snap["abroad_paid"], f_abroad)
    line("Склад", snap["sklad"], f_sklad)
    line("Маркеты (товар)", snap["marketplace_goods"], f_mkt)
    L.append("")
    L.append("> Заказ за границей из файлов — это ПОЛНАЯ сумма заказа (790 млн вон "
             "+ $67k), а в срезе — только оплаченная часть; расхождение ожидаемо.")
    if wb_api["units_missing"]:
        L.append(f"> В отчёте ВБ {wb_api['units_missing']:,.0f} ед. без себестоимости "
                 f"не учтены.".replace(",", " "))
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser(description="Монитор вложений в товар")
    ap.add_argument("--report", action="store_true", help="записать отчёт в outputs/")
    ap.add_argument("--crosscheck", action="store_true", help="сверить срез с файлами")
    args = ap.parse_args()

    r = build(CONFIG["snapshot"])
    text = render(r)
    if args.crosscheck:
        text += "\n" + crosscheck(CONFIG) + "\n"
    print(text)

    if args.report:
        OUT_DIR.mkdir(exist_ok=True)
        path = OUT_DIR / f"investment_monitor_{dt.date.today().isoformat()}.md"
        path.write_text(text, encoding="utf-8")
        print(f"\n[отчёт записан] {path}")


if __name__ == "__main__":
    main()
