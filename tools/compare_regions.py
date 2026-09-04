"""На сколько Бишкек дороже Кореи и сколько на этом берет посредник.

Обычно цены сравнивают только внутри одной страны: у поставщиков разный
маршрут и разные расходы после отгрузки. Здесь задача другая — не выбрать
поставщика, а измерить разрыв. Считаем в рублях от трех баз:

- корейская цена EXW, как она стоит на складе поставщика;
- она же с доставкой и растаможкой до Бишкека (для Киргизии это 5%) —
  разница с ценой прайса и есть заработок посредника;
- она же с нашими расходами на импорт в Москву (по контейнеру 30%) —
  отсюда видно, выгоднее ли везти самим или брать готовое в Бишкеке.

Запуск:
    python3 tools/compare_regions.py --to-base 5 --to-us 30
"""

import argparse
import os

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import brand_names
from rates import rub_per_unit
from supplier_brand_matrix import split_conflicts

SRC = "outputs/prices_normalized.xlsx"
OUT_DIR = "outputs"
# Между странами сравниваем цену как она напечатана в прайсе: приведение
# к штуке делит цену на число пэдов в банке у одной стороны и не делит у
# другой, и позиция уезжает в сто раз.
COLUMN = "Закупка, KRW"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")


def load(country):
    df = pd.read_excel(SRC, dtype={"Штрихкод": str})
    df = df[df["Страна"] == country].copy()
    df["Штрихкод"] = df["Штрихкод"].fillna("").astype(str).str.replace(r"\D", "", regex=True)
    df = df[df["Штрихкод"].str.len() >= 8]
    df = df[df["Закупка, KRW"].notna() & (df["Закупка, KRW"] > 0)]
    df["Бренд в прайсе"] = df["Бренд"]
    df["Бренд"] = brand_names.resolve(df).fillna("БЕЗ БРЕНДА")
    df, _ = split_conflicts(df)
    return df


def cheapest(df, imported):
    """По каждому штрихкоду — самое дешевое предложение страны, в рублях."""
    money = sorted(df["Валюта"].dropna().unique())[0]
    rate = rub_per_unit(money, imported=imported)
    df = df.copy()
    df["Цена, руб"] = df[COLUMN] * rate
    best = df.sort_values("Цена, руб").drop_duplicates("Штрихкод")
    return best.set_index("Штрихкод"), money


def main(base_country, against_country, to_base, to_us):
    base = load(base_country)
    against = load(against_country)

    base_raw, base_money = cheapest(base, imported=False)
    local = f"Цена {base_country} + доставка в {against_country}, руб"
    ours = f"Цена {base_country} + наш импорт, руб"

    rows = []
    for supplier, group in against.groupby("Поставщик"):
        money = sorted(group["Валюта"].dropna().unique())[0]
        rate = rub_per_unit(money, imported=False)
        group = group.copy()
        group["Цена, руб"] = group[COLUMN] * rate
        group = group.sort_values("Цена, руб").drop_duplicates("Штрихкод").set_index("Штрихкод")

        shared = group.index.intersection(base_raw.index)
        if shared.empty:
            continue
        korea = base_raw.loc[shared, "Цена, руб"]
        price = group.loc[shared, "Цена, руб"]
        table = pd.DataFrame({
            "Поставщик": supplier,
            "Бренд": group.loc[shared, "Бренд"],
            "Товар": group.loc[shared, "Название EN"].fillna(group.loc[shared, "Название RU"]),
            f"Цена {against_country}, руб": price.round(2),
            f"Цена {base_country} EXW, руб": korea.round(2),
            local: (korea * (1 + to_base / 100)).round(2),
            ours: (korea * (1 + to_us / 100)).round(2),
            f"Кто дешевле в {base_country}": base_raw.loc[shared, "Поставщик"],
        })
        table["Дороже EXW, %"] = ((price / korea - 1) * 100).round(1)
        # Это и есть заработок бишкекского посредника: он привез товар за
        # 5% сверх корейской цены, а продает нам дороже.
        table["Наценка посредника, %"] = (
            (price / (korea * (1 + to_base / 100)) - 1) * 100).round(1)
        table["Против нашего импорта, %"] = (
            (price / (korea * (1 + to_us / 100)) - 1) * 100).round(1)
        rows.append(table.reset_index())

    if not rows:
        raise SystemExit("Общих штрихкодов между странами нет")
    items = pd.concat(rows, ignore_index=True).sort_values("Наценка посредника, %")
    # Разница в разы — это почти всегда разная фасовка, а не цена.
    items["Пометка"] = ""
    items.loc[items["Дороже EXW, %"].abs() > 70, "Пометка"] = "разная фасовка, сверить"
    checked = items[items["Пометка"] == ""]

    by_supplier = (checked.groupby("Поставщик")
                   .agg(**{"Общих позиций": ("Товар", "size"),
                           "Дороже корейской EXW, %": ("Дороже EXW, %", "median"),
                           "Наценка посредника, %": ("Наценка посредника, %", "median"),
                           "Против нашего импорта, %": ("Против нашего импорта, %", "median"),
                           "Дешевле нашего импорта, позиций":
                               ("Против нашего импорта, %", lambda values: int((values < 0).sum()))})
                   .round(1).reset_index()
                   .sort_values("Наценка посредника, %"))

    by_brand = (checked.groupby("Бренд")
                .agg(**{"Позиций": ("Товар", "size"),
                        "Дороже корейской EXW, %": ("Дороже EXW, %", "median"),
                        "Наценка посредника, %": ("Наценка посредника, %", "median"),
                        "Против нашего импорта, %": ("Против нашего импорта, %", "median")})
                .round(1).reset_index()
                .sort_values("Позиций", ascending=False))

    # У кого из местных дешевле по каждому бренду — с наценкой каждого.
    pivot = (checked.pivot_table(index="Бренд", columns="Поставщик",
                                 values="Наценка посредника, %", aggfunc="median")
             .round(1))
    pivot["Позиций"] = checked.groupby("Бренд").size()
    pivot["Дешевле всех"] = pivot.drop(columns="Позиций").idxmin(axis=1)
    pivot = pivot.sort_values("Позиций", ascending=False).reset_index()

    total = pd.DataFrame([
        {"Показатель": f"Позиций {against_country}, совпавших с {base_country}", "Значение": len(items)},
        {"Показатель": "Из них годятся для сравнения", "Значение": len(checked)},
        {"Показатель": "Отложено: разная фасовка",
         "Значение": int((items["Пометка"] != "").sum())},
        {"Показатель": f"Доставка и растаможка {base_country} -> {against_country}, %",
         "Значение": to_base},
        {"Показатель": "Наши расходы на импорт в Москву, %", "Значение": to_us},
        {"Показатель": "Медиана: дороже корейской цены EXW, %",
         "Значение": round(checked["Дороже EXW, %"].median(), 1)},
        {"Показатель": "Медиана: наценка бишкекского посредника, %",
         "Значение": round(checked["Наценка посредника, %"].median(), 1)},
        {"Показатель": "Медиана: против нашей себестоимости с импортом, %",
         "Значение": round(checked["Против нашего импорта, %"].median(), 1)},
        {"Показатель": "Позиций, где местная цена ниже нашего импорта",
         "Значение": int((checked["Против нашего импорта, %"] < 0).sum())},
    ])

    out_path = os.path.join(OUT_DIR, f"regions_{against_country.lower()}_vs_{base_country.lower()}.xlsx")
    with pd.ExcelWriter(out_path) as writer:
        total.to_excel(writer, sheet_name="ИТОГО", index=False)
        by_supplier.to_excel(writer, sheet_name="ПО ПОСТАВЩИКАМ", index=False)
        by_brand.to_excel(writer, sheet_name="ПО БРЕНДАМ", index=False)
        pivot.to_excel(writer, sheet_name="НАЦЕНКА ПО ПОСТАВЩИКАМ", index=False)
        items.to_excel(writer, sheet_name="ПОЗИЦИИ", index=False)

        for name in writer.book.sheetnames:
            sheet = writer.book[name]
            sheet.freeze_panes = "B2"
            titles = [str(c.value or "") for c in sheet[1]]
            for index, title in enumerate(titles, start=1):
                cell = sheet.cell(row=1, column=index)
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                sheet.column_dimensions[get_column_letter(index)].width = (
                    52 if title == "Товар" else 44 if title == "Показатель" else
                    20 if "Поставщик" in title or title == "Бренд" else 14)
                if "%" in title or "руб" in title:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = "0.0" if "%" in title else "# ##0.00"
                if "%" in title and name != "ИТОГО":
                    letter = get_column_letter(index)
                    sheet.conditional_formatting.add(
                        f"{letter}2:{letter}{sheet.max_row}",
                        ColorScaleRule(start_type="num", start_value=-20, start_color="63BE7B",
                                       mid_type="num", mid_value=20, mid_color="FFEB84",
                                       end_type="num", end_value=100, end_color="F8696B"))

    print(total.to_string(index=False))
    print("\nПо поставщикам:")
    print(by_supplier.to_string(index=False))
    print(f"\nФайл: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Разрыв цен между странами")
    parser.add_argument("--base", default="KR")
    parser.add_argument("--against", default="KG")
    parser.add_argument("--to-base", type=float, default=5.0,
                        help="доставка и растаможка из базовой страны в целевую, %%")
    parser.add_argument("--to-us", type=float, default=30.0,
                        help="наши расходы на импорт в Москву, %%")
    args = parser.parse_args()
    main(args.base, args.against, args.to_base, args.to_us)
