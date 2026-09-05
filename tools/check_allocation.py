"""Проверка плана распределения: склад, приход и кому что предлагаем.

План сотрудника отвечает на вопрос «куда что продавать»: сколько оставить
под Wildberries, сколько под Озон, сколько предложить оптовому покупателю.
Проверяем три вещи: сходится ли распределение с наличием, совпадает ли
склад с файлом остатков и совпадает ли приход с инвойсом на контейнер и
списком по машине.

Запуск:
    python3 tools/check_allocation.py
"""

import argparse
import os

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import name_match
from landed_cost import INVOICE, MANIFEST, clean_barcode, read_invoice

PLAN = "data/shipments/2026-09_plan_raspredeleniya.xlsx"
STOCK = "data/stock_costs/Остатки_31.08.2026.xlsx"
OUT = "outputs/allocation_check.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")

SHARES = ["Резерв ВБ, шт", "Резерв Озон, шт", "Предложить Егору, шт", "Свободный буфер, шт"]
NUMBERS = ["Склад, шт", "Машина", "Контейнер", "Егор уже брал, шт"] + SHARES


def read_plan(path):
    """Строки плана: с артикулом WB — то, что уже продается, без него —
    новинки из прихода, на которые карточки еще не заведены."""
    plan = pd.read_excel(path, header=0)
    plan = plan[plan["Наименование"].notna() & (plan["Наименование"] != "ИТОГО")]
    for column in NUMBERS:
        plan[column] = pd.to_numeric(plan[column], errors="coerce").fillna(0)
    plan["Новинка"] = plan["Артикул WB"].isna()
    return plan.reset_index(drop=True)


def read_stock(path):
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = []
    for line in range(3, sheet.max_row + 1):
        name = sheet.cell(line, 1).value
        if not name:
            continue
        rows.append({
            "Товар в остатках": str(name).strip(),
            "Остаток по файлу, шт": sheet.cell(line, 3).value or 0,
            "Заказ покупателя, шт": sheet.cell(line, 6).value or 0,
            "Себестоимость, руб": sheet.cell(line, 4).value,
        })
    book.close()
    return pd.DataFrame(rows)


def join(left, left_column, right, right_column):
    pairs = name_match.match(left[left_column], right[right_column])
    left = left.copy()
    left["_pair"] = left.index.map(pairs.get)
    joined = left.merge(right, left_on="_pair", right_index=True, how="left")
    return joined.drop(columns="_pair")


def incoming(invoice, manifest):
    """Приход по документам: контейнер плюс машина, одной таблицей."""
    container = invoice[["Товар", "Загружено, шт"]].rename(
        columns={"Товар": "Товар в документе", "Загружено, шт": "Контейнер по инвойсу, шт"})
    container["Машина по списку, шт"] = 0
    truck = manifest[["Товар", "Количество"]].rename(
        columns={"Товар": "Товар в документе", "Количество": "Машина по списку, шт"})
    truck["Контейнер по инвойсу, шт"] = 0
    both = pd.concat([container, truck], ignore_index=True)
    # Тонер CELIMAX едет и контейнером, и машиной. Если оставить две
    # строки, сопоставление станет неоднозначным и обе выпадут.
    both["_key"] = both["Товар в документе"].map(name_match.normal)
    grouped = both.groupby("_key", as_index=False).agg({
        "Товар в документе": lambda names: max(names, key=len),
        "Контейнер по инвойсу, шт": "sum",
        "Машина по списку, шт": "sum"})
    return grouped.drop(columns="_key")


def format_sheet(sheet):
    sheet.freeze_panes = "B2"
    titles = [str(c.value or "") for c in sheet[1]]
    for index, title in enumerate(titles, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = (
            58 if "Товар" in title or "Наименование" in title or title == "Показатель"
            else 22 if title in ("Значение", "Что не так") else 13)
        if "шт" in title or "руб" in title:
            for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                row[0].number_format = "# ##0"
    if "Что не так" in titles:
        column = titles.index("Что не так")
        for row in sheet.iter_rows(min_row=2):
            if row[column].value:
                row[column].fill = ALERT_FILL


def main():
    plan = read_plan(PLAN)
    stock = read_stock(STOCK)
    invoice, _ = read_invoice(INVOICE)
    manifest = pd.read_csv(MANIFEST, sep=";", dtype={"Штрихкод": str})
    arrivals = incoming(invoice, manifest)

    plan["Распределено, шт"] = plan[SHARES].sum(axis=1)
    plan["Всего товара, шт"] = plan["Склад, шт"] + plan["Машина"] + plan["Контейнер"]
    plan["Не распределено, шт"] = plan["Всего товара, шт"] - plan["Распределено, шт"]

    # Сверка склада и заказа покупателя.
    with_stock = join(plan[~plan["Новинка"]], "Наименование", stock, "Товар в остатках")
    with_stock["Разница по складу, шт"] = (
        with_stock["Склад, шт"] - with_stock["Остаток по файлу, шт"].fillna(0))
    with_stock["Разница с заказом, шт"] = (
        with_stock["Предложить Егору, шт"] - with_stock["Заказ покупателя, шт"].fillna(0))

    stock_check = with_stock[with_stock["Товар в остатках"].notna()
                             & (with_stock["Разница по складу, шт"] != 0)].copy()
    stock_check["Что не так"] = "в плане склада больше, чем в файле остатков"

    order_check = with_stock[with_stock["Товар в остатках"].notna()
                             & (with_stock["Разница с заказом, шт"] != 0)].copy()
    order_check["Что не так"] = order_check["Разница с заказом, шт"].map(
        lambda value: "предлагаем больше, чем просил покупатель" if value > 0
        else "предлагаем меньше, чем просил покупатель")

    # Сверка прихода с документами.
    arrival_check = join(plan[(plan["Машина"] + plan["Контейнер"]) > 0],
                         "Наименование", arrivals, "Товар в документе")
    arrival_check["Разница по контейнеру, шт"] = (
        arrival_check["Контейнер"] - arrival_check["Контейнер по инвойсу, шт"].fillna(0))
    arrival_check["Разница по машине, шт"] = (
        arrival_check["Машина"] - arrival_check["Машина по списку, шт"].fillna(0))
    arrival_check["Что не так"] = ""
    arrival_check.loc[arrival_check["Товар в документе"].isna(), "Что не так"] = \
        "в документах прихода такой позиции нет"
    mismatch = (arrival_check["Товар в документе"].notna() &
                ((arrival_check["Разница по контейнеру, шт"] != 0) |
                 (arrival_check["Разница по машине, шт"] != 0)))
    arrival_check.loc[mismatch, "Что не так"] = "количество не сходится с документом"

    not_planned = plan[(plan["Не распределено, шт"] > 0)].copy()
    unmatched = with_stock[with_stock["Товар в остатках"].isna()][["Наименование", "Склад, шт"]]

    total = pd.DataFrame([
        {"Показатель": "Позиций в плане", "Значение": len(plan)},
        {"Показатель": "Из них новинки без артикула WB", "Значение": int(plan["Новинка"].sum())},
        {"Показатель": "Склад по плану, шт", "Значение": int(plan["Склад, шт"].sum())},
        {"Показатель": "Машина по плану, шт", "Значение": int(plan["Машина"].sum())},
        {"Показатель": "Контейнер по плану, шт", "Значение": int(plan["Контейнер"].sum())},
        {"Показатель": "Контейнер по инвойсу, шт",
         "Значение": int(invoice["Загружено, шт"].sum())},
        {"Показатель": "Машина по списку, шт", "Значение": int(manifest["Количество"].sum())},
        {"Показатель": "Распределено всего, шт", "Значение": int(plan["Распределено, шт"].sum())},
        {"Показатель": "Не распределено, шт", "Значение": int(plan["Не распределено, шт"].sum())},
        {"Показатель": "Резерв ВБ, шт", "Значение": int(plan["Резерв ВБ, шт"].sum())},
        {"Показатель": "Резерв Озон, шт", "Значение": int(plan["Резерв Озон, шт"].sum())},
        {"Показатель": "Предложить покупателю, шт",
         "Значение": int(plan["Предложить Егору, шт"].sum())},
        {"Показатель": "Свободный буфер, шт", "Значение": int(plan["Свободный буфер, шт"].sum())},
        {"Показатель": "Позиций, где склад расходится с файлом остатков", "Значение": len(stock_check)},
        {"Показатель": "Позиций, где приход не сходится с документами",
         "Значение": int((arrival_check["Что не так"] != "").sum())},
    ])

    plan_columns = ["Артикул WB", "Наименование", "Склад, шт", "Машина", "Контейнер",
                    "Всего товара, шт", "Резерв ВБ, шт", "Резерв Озон, шт",
                    "Предложить Егору, шт", "Свободный буфер, шт",
                    "Распределено, шт", "Не распределено, шт"]
    with pd.ExcelWriter(OUT) as writer:
        total.to_excel(writer, sheet_name="ИТОГО", index=False)
        not_planned[plan_columns].to_excel(writer, sheet_name="НЕ РАСПРЕДЕЛЕНО", index=False)
        arrival_check[["Наименование", "Машина", "Контейнер", "Товар в документе",
                       "Контейнер по инвойсу, шт", "Машина по списку, шт",
                       "Разница по контейнеру, шт", "Разница по машине, шт",
                       "Что не так"]].to_excel(writer, sheet_name="СВЕРКА ПРИХОДА", index=False)
        stock_check[["Наименование", "Склад, шт", "Товар в остатках",
                     "Остаток по файлу, шт", "Разница по складу, шт",
                     "Что не так"]].to_excel(writer, sheet_name="СВЕРКА СКЛАДА", index=False)
        order_check[["Наименование", "Склад, шт", "Заказ покупателя, шт",
                     "Предложить Егору, шт", "Разница с заказом, шт",
                     "Что не так"]].to_excel(writer, sheet_name="ПОКУПАТЕЛЬ", index=False)
        unmatched.to_excel(writer, sheet_name="НЕ НАШЛОСЬ В ОСТАТКАХ", index=False)
        plan[plan_columns].to_excel(writer, sheet_name="ПЛАН ЦЕЛИКОМ", index=False)
        for sheet in writer.book.sheetnames:
            format_sheet(writer.book[sheet])

    print(total.to_string(index=False))
    print(f"\nФайл: {OUT}")


if __name__ == "__main__":
    argparse.ArgumentParser(description="Проверка плана распределения").parse_args()
    main()
