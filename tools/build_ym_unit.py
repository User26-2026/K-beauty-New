# -*- coding: utf-8 -*-
"""Пересчёт юнит-экономики под правила Яндекс Маркета: листы FBO (FBY) и FBS.

Исходник — рабочая модель продавца с листами `Юнитка FBO` и `Юнитка FBS`.
Скрипт читает из неё товары, цены, себестоимость, объём, выкуп и ДРР, затем
строит новую книгу, где все статьи расходов считаются по правилам Маркета,
а тариф за услуги берётся по подкатегории товара из официального файла
тарифов с 01.07.2026.

Использование:
    python3 tools/build_ym_unit.py data/unit_economics/unit_source_110826.xlsx \
        outputs/yandex_market/unit_yandex_market_FBO_FBS.xlsx
"""
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.comments import Comment

SRC_PATH = sys.argv[1] if len(sys.argv) > 1 else "data/unit_economics/unit_source_110826.xlsx"
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else "outputs/yandex_market/unit_yandex_market_FBO_FBS.xlsx"

# ставки Маркета с 01.07.2026 по подкатегориям: (FBY/Express, FBS/DBS)
RATES = {
    "Средства для ухода за лицом": (0.40, 0.47),
    "Средства для умывания":       (0.41, 0.48),
    "Средства для снятия макияжа": (0.41, 0.48),
    "Скрабы для лица":             (0.41, 0.48),
    "Маски для лица":              (0.41, 0.48),
    "Уход за кожей вокруг глаз":   (0.41, 0.48),
    "Уход за кожей губ":           (0.40, 0.47),
    "Загар и защита от солнца":    (0.41, 0.48),
    "Шампуни":                     (0.40, 0.47),
    "Маски и сыворотки (волосы)":  (0.40, 0.47),
    "Тональные средства":          (0.40, 0.47),
    "Основы под макияж":           (0.40, 0.47),
}


def classify(name):
    """Подкатегория Маркета по названию товара."""
    n = name.lower()
    if "патч" in n or "eye patch" in n or "крем для глаз" in n or "eye cream" in n or "eye serum" in n or "для век" in n:
        return "Уход за кожей вокруг глаз"
    if "lip sleeping" in n or "для губ" in n:
        return "Уход за кожей губ"
    if "шампун" in n or "shampoo" in n:
        return "Шампуни"
    if "для волос" in n or "treatment protein" in n:
        return "Маски и сыворотки (волосы)"
    if "тональный" in n or "foundation" in n or "bb крем" in n or "bb cream" in n:
        return "Тональные средства"
    if "база под макияж" in n or "glow base" in n:
        return "Основы под макияж"
    if "скраб" in n or "scrub" in n:
        return "Скрабы для лица"
    if "гидрофильное масло" in n or "cleansing oil" in n or "снятия макияжа" in n or "мицелляр" in n:
        return "Средства для снятия макияжа"
    if "пенка" in n or "cleanser" in n or "cleansing foam" in n or "для умывания" in n or "bubble foam" in n:
        return "Средства для умывания"
    if "маска для лица" in n or "mask pack" in n or "face mask" in n or "маска" in n:
        return "Маски для лица"
    return "Средства для ухода за лицом"


def read_source(path):
    """Считывает исходную модель продавца в {имя листа: [строки]}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        fbo = ws.title.endswith("FBO")
        hdr = 12 if fbo else 13
        cols = (dict(name=1, stock=4, cprice=5, cdrr=7, price=8, spp=9, cost=11, vol=24, buyout=21, cpc=27, top=29)
                if fbo else
                dict(name=2, stock=6, cprice=7, cdrr=9, price=10, spp=11, cost=13, vol=26, buyout=23, cpc=29, top=31))
        rows = []
        for r in range(hdr + 1, ws.max_row + 1):
            if ws.cell(r, cols["name"]).value is None:
                continue
            rows.append({k: ws.cell(r, c).value for k, c in cols.items()})
        out[ws.title] = rows
    return out


SRC = read_source(SRC_PATH)

FONT = "Arial"
YELLOW = PatternFill("solid", fgColor="FFFF00")
GREY = PatternFill("solid", fgColor="F2F2F2")
HDRFILL = PatternFill("solid", fgColor="D9E1F2")
GROUPFILL = PatternFill("solid", fgColor="EDEDED")
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RUB = '#,##0.00;(#,##0.00);-'
RUB0 = '#,##0;(#,##0);-'
PCT = '0.0%;(0.0%);-'

# ---- параметры листов -------------------------------------------------------
# (метка, значение, формат, вход/расчёт, примечание)
def params(fbs):
    p = [
        ("Налог с оборота", 0.07, PCT, True, "Из исходной модели продавца (УСН 7%)"),
        ("Перевод платежа, %", 0.023, PCT, True,
         "Тариф Маркета за перевод платежа при выплатах раз в неделю с отсрочкой 2 недели. "
         "Другие графики: 1,6% (отсрочка 4 недели), 2,8% (отсрочка 1 неделя)."),
        ("Приём платежа, ₽ за SKU", 0.12, RUB, True, "Тариф Маркета: 0,12 ₽ за каждый отличающийся SKU в заказе"),
        ("Обработка заказа, % от цены", 0.04, PCT, True, "Сборка и упаковка на складе Маркета: 4% от цены товара"),
        ("Обработка заказа, минимум ₽", 25.0, RUB, True, "Нижняя граница платы за сборку и упаковку"),
        ("Обработка заказа, максимум ₽", 100.0, RUB, True, "Верхняя граница платы за сборку и упаковку"),
        ("Хранение, ₽/л в день", 3.0, RUB, True,
         "Тариф Маркета для косметики. Хранение бесплатное при оборачиваемости до 365 дней, "
         "поэтому по умолчанию дней хранения = 0."),
        ("Дней платного хранения", 0.0, RUB0, True, "Ставится больше нуля только для зависших остатков старше года"),
        ("Фулфилмент наш, ₽/шт", 60.0, RUB, True, "Из исходной модели продавца: подготовка товара на нашей стороне"),
        ("Доля скидки за счёт продавца", 1.0, PCT, True,
         "1 = скидку финансируем мы (тариф и выручка считаются от цены покупателя). "
         "0 = скидку финансирует Маркет, как СПП на WB. Значение по умолчанию консервативное."),
        ("Режим тарифа: 1 стандартный / 2 льготный", 1.0, RUB0, True,
         "1 = ставка из официального файла тарифов Маркета с 01.07.2026, доставка покупателю, "
         "средняя миля и возвраты считаются включёнными в неё. "
         "2 = льготный тариф для красоты (5% FBY / 12% FBS), логистика и возвраты платятся отдельно."),
        ("Льготный тариф размещения, %", 0.12 if fbs else 0.05, PCT, True,
         "Льгота для категории Товары для красоты: 5% на FBY и Express, 12% на FBS с отгрузкой "
         "дольше 36 часов и на DBS. Действует до конца 2026 года. Проверить в личном кабинете."),
        ("Доставка покупателю, % (режим 2)", 0.05, PCT, True, "5% от цены товара, не более 1000 ₽"),
        ("Доставка покупателю, максимум ₽", 1000.0, RUB, True, "Верхняя граница платы за доставку покупателю"),
        ("Средняя миля, ₽/л (режим 2)", 5.0, RUB, True,
         "ОЦЕНКА. Маркет считает среднюю милю по объёму в литрах, не более 5500 ₽ за единицу, "
         "но саму ставку за литр подтвердить не удалось. Подставьте значение из личного кабинета."),
        ("Обработка возврата, ₽ (режим 2)", 15.0, RUB, True, "Плата за обработку возврата или невыкупа"),
        ("Обратная доставка, ₽ (режим 2)", 67.0, RUB, True,
         "Доставка от покупателя обратно идёт по междугородному тарифу. Взято значение обратной "
         "логистики из исходной модели продавца, уточнить в личном кабинете."),
        ("Баллы Плюса за наш счёт, %", 0.0, PCT, True,
         "Софинансирование программы лояльности. Ставится, если участвуем в кешбэке баллами Плюса."),
    ]
    if fbs:
        p += [
            ("Обработка отправления, ₽ (FBS)", 50.0, RUB, True, "Обработка заказа на стороне Маркета по модели FBS"),
            ("Отгрузка в СЦ или ПВЗ, ₽ (FBS)", 10.0, RUB, True,
             "Самопривоз: сортировочный центр около 10 ₽ за заказ, пункт приёма около 25 ₽. "
             "При выезде курьера ставится стоимость выезда, делённая на число заказов."),
        ]
    return p

COLS = [
    ("Товар", 52), ("Категория Маркета", 26), ("Остаток", 9),
    ("Цена конк.", 11), ("ДРР конк.", 10),
    ("Цена без скидки", 14), ("Скидка, %", 10), ("Цена покупателя", 14), ("Выручка продавца", 15),
    ("Себест.", 11), ("Фулфилмент", 11), ("База затрат", 12),
    ("Тариф Маркета %", 13), ("Тариф Маркета ₽", 14),
    ("Доставка покупателю", 15), ("Объём, л", 9), ("Средняя миля", 12),
    ("Обработка заказа", 14), ("Приём и перевод платежа", 16),
    ("Хранение", 10), ("Выкуп %", 9), ("Возвраты и невыкуп", 15),
    ("Баллы Плюса", 11), ("Налог", 10),
    ("Буст продаж %", 12), ("Буст продаж ₽", 13),
    ("Маржа до рекламы", 15), ("Маржа/шт", 12), ("Маржа % цены", 12), ("ROI %", 10),
    ("Цена безубытка", 13),
]
# индексы колонок
C_ = {name: i + 1 for i, (name, _) in enumerate(COLS)}


def build_sheet(wb, title, sheet_key, fbs):
    ws = wb.create_sheet(title)
    P = params(fbs)
    ws["A1"] = "ЮНИТ-ЭКОНОМИКА ЯНДЕКС МАРКЕТ — " + ("FBS, свой склад" if fbs else "FBO (FBY), склад Маркета")
    ws["A1"].font = Font(name=FONT, size=12, bold=True)
    ws["A2"] = "Жёлтые ячейки — вход, всё остальное считается формулами. Наведите курсор на подпись параметра: в примечании источник цифры."
    ws["A2"].font = Font(name=FONT, size=9, italic=True)

    prow = {}
    r = 4
    for label, val, fmt, _inp, note in P:
        ws.cell(r, 1, label).font = BOLD
        c = ws.cell(r, 3, val)
        c.font = BLUE; c.fill = YELLOW; c.number_format = fmt; c.border = BOX
        if note:
            ws.cell(r, 1).comment = Comment(note, "Источник")
        prow[label] = r
        r += 1

    def P_(label):
        return f"$C${prow[label]}"

    p_tax = P_("Налог с оборота"); p_pay = P_("Перевод платежа, %"); p_acq = P_("Приём платежа, ₽ за SKU")
    p_hp = P_("Обработка заказа, % от цены"); p_hmin = P_("Обработка заказа, минимум ₽"); p_hmax = P_("Обработка заказа, максимум ₽")
    p_st = P_("Хранение, ₽/л в день"); p_sd = P_("Дней платного хранения"); p_ff = P_("Фулфилмент наш, ₽/шт")
    p_sh = P_("Доля скидки за счёт продавца"); p_mode = P_("Режим тарифа: 1 стандартный / 2 льготный")
    p_soft = P_("Льготный тариф размещения, %"); p_dl = P_("Доставка покупателю, % (режим 2)")
    p_dlmax = P_("Доставка покупателю, максимум ₽"); p_mm = P_("Средняя миля, ₽/л (режим 2)")
    p_ret = P_("Обработка возврата, ₽ (режим 2)"); p_retdel = P_("Обратная доставка, ₽ (режим 2)")
    p_loy = P_("Баллы Плюса за наш счёт, %")
    p_fbsh = P_("Обработка отправления, ₽ (FBS)") if fbs else None
    p_fbss = P_("Отгрузка в СЦ или ПВЗ, ₽ (FBS)") if fbs else None

    hdr = r + 1
    for j, (name, width) in enumerate(COLS, 1):
        c = ws.cell(hdr, j, name)
        c.font = Font(name=FONT, size=10, bold=True)
        c.fill = HDRFILL; c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[gcl(j)].width = width
    ws.row_dimensions[hdr].height = 42

    rows = SRC[sheet_key]
    rr = hdr + 1
    data_rows = []
    for rec in rows:
        name = str(rec["name"])
        price = rec["price"]
        if not isinstance(price, (int, float)):
            if name.startswith("Косметика/"):
                c = ws.cell(rr, 1, name)
                c.font = BOLD; c.fill = GROUPFILL
                for j in range(1, len(COLS) + 1):
                    ws.cell(rr, j).fill = GROUPFILL
                rr += 1
            continue
        cat = classify(name)
        rate = RATES[cat][1 if fbs else 0]
        cost = rec["cost"] or 0
        vol = rec["vol"] or 0
        spp = rec["spp"] or 0
        buyout = rec["buyout"] or 0.9
        drr = (rec["cpc"] or 0) + (rec["top"] or 0)
        cprice = rec["cprice"] if isinstance(rec["cprice"], (int, float)) else None
        cdrr = rec["cdrr"] if isinstance(rec["cdrr"], (int, float)) else None
        stock = rec["stock"] if isinstance(rec["stock"], (int, float)) else None
        R = rr

        def put(col, value, fmt=None, inp=False, formula=False):
            c = ws.cell(R, C_[col], value)
            c.font = BLUE if inp else BLACK
            if inp:
                c.fill = YELLOW
            if fmt:
                c.number_format = fmt
            c.border = BOX
            return c

        put("Товар", name).alignment = Alignment(wrap_text=False)
        put("Категория Маркета", cat, inp=True)
        put("Остаток", stock, RUB0)
        put("Цена конк.", cprice, RUB0)
        put("ДРР конк.", cdrr, '0.0')
        put("Цена без скидки", price, RUB, inp=True)
        put("Скидка, %", spp, PCT, inp=True)
        put("Цена покупателя", f"=F{R}*(1-G{R})", RUB)
        put("Выручка продавца", f"=F{R}-F{R}*G{R}*{p_sh}", RUB)
        put("Себест.", cost, RUB, inp=True)
        put("Фулфилмент", f"={p_ff}", RUB)
        put("База затрат", f"=J{R}+K{R}", RUB)
        put("Тариф Маркета %", f"=IF({p_mode}=1,{rate},{p_soft})", PCT)
        put("Тариф Маркета ₽", f"=H{R}*M{R}", RUB)
        put("Доставка покупателю", f"=IF({p_mode}=2,MIN(H{R}*{p_dl},{p_dlmax}),0)", RUB)
        put("Объём, л", vol, '0.00', inp=True)
        put("Средняя миля", f"=IF({p_mode}=2,P{R}*{p_mm},0)", RUB)
        if fbs:
            put("Обработка заказа", f"={p_fbsh}+{p_fbss}", RUB)
        else:
            put("Обработка заказа", f"=MIN(MAX(H{R}*{p_hp},{p_hmin}),{p_hmax})", RUB)
        put("Приём и перевод платежа", f"=H{R}*{p_pay}+{p_acq}", RUB)
        put("Хранение", f"={p_st}*{p_sd}*P{R}", RUB)
        put("Выкуп %", buyout, PCT, inp=True)
        put("Возвраты и невыкуп", f"=IF({p_mode}=2,(1-U{R})*({p_ret}+{p_retdel}),0)", RUB)
        put("Баллы Плюса", f"=H{R}*{p_loy}", RUB)
        put("Налог", f"=I{R}*{p_tax}", RUB)
        put("Буст продаж %", drr, PCT, inp=True)
        put("Буст продаж ₽", f"=H{R}*Y{R}", RUB)
        put("Маржа до рекламы", f"=I{R}-N{R}-O{R}-Q{R}-R{R}-S{R}-T{R}-V{R}-W{R}-X{R}-L{R}", RUB)
        put("Маржа/шт", f"=AA{R}-Z{R}", RUB)
        put("Маржа % цены", f"=IFERROR(AB{R}/H{R},0)", PCT)
        put("ROI %", f"=IFERROR(AB{R}/L{R},0)", PCT)
        k = f"((1-G{R}*{p_sh})/(1-G{R}))"
        pct = f"(M{R}+{p_pay}+Y{R}+{p_loy}+IF({p_mode}=2,{p_dl},0))"
        put("Цена безубытка",
            f"=IFERROR((L{R}+R{R}+T{R}+Q{R}+V{R}+{p_acq})/({k}-{pct}-{p_tax}*{k}),0)", RUB)
        data_rows.append(R)
        rr += 1

    # итоговая строка
    tot = rr + 1
    ws.cell(tot, 1, "ИТОГО / средневзвешенно по SKU с ценой").font = BOLD
    first, last = data_rows[0], data_rows[-1]
    for col, formula in [
        ("Маржа/шт", f"=AVERAGE(AB{first}:AB{last})"),
        ("Маржа % цены", f"=IFERROR(SUM(AB{first}:AB{last})/SUM(H{first}:H{last}),0)"),
        ("ROI %", f"=IFERROR(SUM(AB{first}:AB{last})/SUM(L{first}:L{last}),0)"),
        ("Цена покупателя", f"=AVERAGE(H{first}:H{last})"),
        ("Тариф Маркета ₽", f"=SUM(N{first}:N{last})"),
    ]:
        c = ws.cell(tot, C_[col], formula)
        c.font = BOLD
        c.number_format = PCT if "%" in col else RUB
        c.fill = GREY
    ws.cell(tot, C_["Маржа до рекламы"], f"=AVERAGE(AA{first}:AA{last})").font = BOLD
    ws.cell(tot, C_["Маржа до рекламы"]).number_format = RUB
    ws.cell(tot, C_["Маржа до рекламы"]).fill = GREY

    note = tot + 2
    ws.cell(note, 1, "Как читать модель").font = BOLD
    notes = [
        "Тариф Маркета в режиме 1 взят из официального файла marketplace_services_rates_01072026.xlsx по подкатегории каждого товара.",
        "Категория Маркета проставлена по названию товара. Проверьте её по спорным позициям: ставка меняется на 14-15 п.п., если карточка висит на верхнем узле категории.",
        "Все проценты Маркета считаются от цены покупателя, а не от цены до скидки: тариф берётся от стоимости проданного товара.",
        "Выручка продавца зависит от параметра доли скидки: при 1 скидку финансируем мы, при 0 её финансирует Маркет.",
        "В режиме 1 доставка покупателю, средняя миля и возвраты обнулены, так как считаются включёнными в тариф. В режиме 2 они считаются отдельно.",
        "Наценки за нелокальные продажи и кроссдока по складам в модели нет: это логика Wildberries, у Маркета таких статей нет.",
        "Штрафы за отмену и просрочку отгрузки в модель не заложены, они разовые и зависят от дисциплины отгрузок.",
    ]
    for i, t in enumerate(notes):
        ws.cell(note + 1 + i, 1, "— " + t).font = Font(name=FONT, size=9)

    ws.freeze_panes = ws.cell(hdr + 1, 3)
    return ws, data_rows


wb = openpyxl.Workbook()
wb.remove(wb.active)
build_sheet(wb, "Юнитка ЯМ FBO", "Юнитка FBO", fbs=False)
build_sheet(wb, "Юнитка ЯМ FBS", "Юнитка FBS", fbs=True)
out = OUT_PATH
wb.save(out)
print("saved", out)
