"""Сравнение цен на один бренд у всех поставщиков.

Товары сопоставляем по штрихкоду: названия у поставщиков расходятся
(MEDI-PEEL против MEDIPEEL, разные хвосты и приписки), а штрихкод один.

Запуск:
    python3 tools/compare_brand.py MEDIPEEL
    python3 tools/compare_brand.py "ROUND LAB" --min-diff 5
"""

import argparse
import os
import re

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from price_unit import unify_packs

SRC = "outputs/prices_normalized.xlsx"
OUT_DIR = "outputs"
RATE = 0.058          # рублей за вону, тот же курс, что в разборе прайсов
IMPORT = 1.4          # логистика, пошлина и приемка


def load_brand(brand):
    df = pd.read_excel(SRC)
    mask = (df["Бренд"].fillna("").astype(str).str.upper().str.contains(brand.upper())
            | df["Название EN"].fillna("").astype(str).str.upper().str.contains(brand.upper()))
    rows = df[mask].copy()
    rows["Штрихкод"] = rows["Штрихкод"].astype(str).str.replace(r"\D", "", regex=True)
    return rows[rows["Штрихкод"].str.len() >= 8]


HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BEST_FILL = PatternFill("solid", fgColor="E2EFDA")


def format_sheet(worksheet, price_columns):
    """Ширины, закрепление шапки и денежный формат — чтобы таблицу читали."""
    worksheet.freeze_panes = "B2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    widths = {"Название EN": 52, "Объем": 18, "Штрихкод": 15, "Дешевле у": 18}
    for index, cell in enumerate(worksheet[1], start=1):
        title = str(cell.value or "")
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = widths.get(title, 15)
        if title in price_columns or "KRW" in title or "руб" in title:
            for row in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
                row[0].number_format = "# ##0"

    # Подсветим цену того поставщика, у кого дешевле.
    titles = [str(c.value or "") for c in worksheet[1]]
    if "Дешевле у" not in titles:
        return
    winner_col = titles.index("Дешевле у") + 1
    for row in worksheet.iter_rows(min_row=2):
        winner = row[winner_col - 1].value
        if winner in titles:
            row[titles.index(winner)].fill = BEST_FILL


def overpay_table(prices, info, bases):
    """На сколько процентов каждый поставщик дороже самого дешевого."""
    cheapest = prices.min(axis=1)
    overpay = prices.div(cheapest, axis=0).sub(1).mul(100).round(1)
    table = info.join(overpay)
    table["Базисы"] = bases.apply(lambda row: ", ".join(sorted(row.dropna().unique())), axis=1)
    table["Мин. цена, KRW"] = cheapest.round(0)
    table["Дешевле у"] = prices.idxmin(axis=1)
    return table


def summary_table(prices, bases):
    """Свод по поставщикам: охват, победы, переплата, стоимость корзины."""
    cheapest = prices.min(axis=1)
    common = prices.dropna()
    rows = []
    for supplier in prices.columns:
        column = prices[supplier]
        over = (column.dropna() / cheapest[column.notna()] - 1) * 100
        basket = common[supplier].sum() if not common.empty else None
        rows.append({
            "Поставщик": supplier,
            "Базис": ", ".join(sorted(bases[supplier].dropna().unique())),
            "Позиций": int(column.notna().sum()),
            "Дешевле всех": int((column == cheapest).sum()),
            "Переплата медиана, %": round(over.median(), 1),
            "Переплата средняя, %": round(over.mean(), 1),
            "Максимум переплаты, %": round(over.max(), 1),
            f"Корзина из {len(common)} позиций, KRW": basket,
        })
    result = pd.DataFrame(rows).sort_values("Переплата медиана, %")
    basket_col = f"Корзина из {len(common)} позиций, KRW"
    if result[basket_col].notna().any():
        base = result[basket_col].min()
        result["Корзина дороже минимума, %"] = ((result[basket_col] / base - 1) * 100).round(1)
    return result


def paint_overpay(worksheet, first_col, last_col, first_row, last_row):
    """Зеленый — самая низкая цена, красный — самая высокая переплата."""
    span = (f"{get_column_letter(first_col)}{first_row}"
            f":{get_column_letter(last_col)}{last_row}")
    worksheet.conditional_formatting.add(span, ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="num", mid_value=10, mid_color="FFEB84",
        end_type="num", end_value=30, end_color="F8696B"))
    for row in worksheet.iter_rows(min_row=first_row, max_row=last_row,
                                   min_col=first_col, max_col=last_col):
        for cell in row:
            cell.number_format = "0.0\%"


def main(brand, min_diff):
    if not os.path.exists(SRC):
        raise SystemExit(f"Нет файла {SRC} — сначала запустите parse_price_lists.py")

    rows = load_brand(brand)
    if rows.empty:
        raise SystemExit(f"Бренд {brand} не найден")

    # Фасовку сводим по штрихкоду: часть поставщиков объем не пишет вовсе.
    rows, pack_conflicts = unify_packs(rows)
    if pack_conflicts:
        print(f"Разное число штук в упаковке у поставщиков: {len(pack_conflicts)} "
              f"позиций, напр. {', '.join(pack_conflicts[:3])}")

    # У одного поставщика товар может встретиться дважды — берем дешевле.
    best = (rows.sort_values("Цена за штуку (сводно)")
                .drop_duplicates(["Штрихкод", "Поставщик"]))
    suppliers = sorted(best["Поставщик"].unique())
    print(f"{brand}: SKU со штрихкодом по поставщикам")
    print(best.groupby("Поставщик").size().to_string())

    # Сравниваем цену за штуку: у одного поставщика это может быть набор.
    prices = best.pivot(index="Штрихкод", columns="Поставщик", values="Цена за штуку (сводно)")
    units = best.pivot(index="Штрихкод", columns="Поставщик", values="Единица цены")
    units.columns = [f"{c}: единица" for c in units.columns]
    bases = best.pivot(index="Штрихкод", columns="Поставщик", values="Базис")
    info = (best.sort_values("Название EN", key=lambda s: s.str.len(), ascending=False)
                .drop_duplicates("Штрихкод")
                .set_index("Штрихкод")[["Название EN", "Объем", "MSRP, KRW"]])

    table = info.join(prices).join(units)
    # EXW и FOB напрямую сравнивать нельзя: при EXW доставку до порта платим мы.
    table["Базисы"] = bases.apply(lambda row: ", ".join(sorted(row.dropna().unique())), axis=1)
    shared = table[(prices.notna().sum(axis=1) > 1).reindex(table.index, fill_value=False)].copy()
    print(f"\nЕсть более чем у одного поставщика: {len(shared)} из {len(table)}")
    if shared.empty:
        return

    price_cols = [c for c in suppliers if c in shared.columns]
    shared["Дешевле у"] = shared[price_cols].idxmin(axis=1)
    shared["Мин, KRW"] = shared[price_cols].min(axis=1)
    shared["Макс, KRW"] = shared[price_cols].max(axis=1)
    shared["Разница, %"] = ((1 - shared["Мин, KRW"] / shared["Макс, KRW"]) * 100).round(1)
    shared["Экономия, руб"] = ((shared["Макс, KRW"] - shared["Мин, KRW"]) * RATE * IMPORT).round(0)
    shared["Себестоимость мин, руб"] = (shared["Мин, KRW"] * RATE * IMPORT).round(0)
    # Фасовку считаем сопоставимой, только если единица цены совпала у всех.
    unit_cols = [c for c in shared.columns if c.endswith(": единица")]
    shared["Фасовка совпала"] = shared[unit_cols].apply(
        lambda row: row.dropna().nunique() <= 1, axis=1)
    shared["Базис совпал"] = ~shared["Базисы"].str.contains(",")

    print("\nКто дешевле, по числу позиций:")
    print(shared["Дешевле у"].value_counts().to_string())
    print(f"\nМедиана расхождения: {shared['Разница, %'].median():.1f}%")

    shared = shared.sort_values("Разница, %", ascending=False)
    only_one = table[(prices.notna().sum(axis=1) == 1).reindex(table.index, fill_value=False)].copy()
    only_one["Есть только у"] = only_one[price_cols].apply(
        lambda row: row.dropna().index[0] if row.notna().any() else "", axis=1)

    out_path = os.path.join(OUT_DIR, f"brand_{re.sub(r'[^a-z0-9]+', '_', brand.lower())}_by_supplier.xlsx")
    overpay = overpay_table(prices.loc[shared.index], info.loc[shared.index],
                            bases.loc[shared.index])
    overpay = overpay.sort_values("Мин. цена, KRW", ascending=False)
    summary = summary_table(prices.loc[shared.index], bases.loc[shared.index])

    with pd.ExcelWriter(out_path) as writer:
        summary.to_excel(writer, sheet_name="ИТОГО", index=False)
        overpay.to_excel(writer, sheet_name="ПЕРЕПЛАТА %")
        shared.to_excel(writer, sheet_name="СРАВНЕНИЕ")
        only_one.to_excel(writer, sheet_name="ТОЛЬКО У ОДНОГО")
        table.to_excel(writer, sheet_name="ВСЕ ПОЗИЦИИ")
        for sheet in writer.book.worksheets:
            format_sheet(sheet, price_cols)
        sheet = writer.sheets["ПЕРЕПЛАТА %"]
        titles = [str(c.value) for c in sheet[1]]
        columns = [titles.index(s) + 1 for s in price_cols if s in titles]
        if columns:
            paint_overpay(sheet, min(columns), max(columns), 2, len(overpay) + 1)

    print("\n=== Насколько дороже самого дешевого ===")
    print(summary.to_string(index=False))
    print(f"\nПозиций только у одного поставщика: {len(only_one)}")
    print(only_one["Есть только у"].value_counts().to_string())
    mismatch = int((~shared["Фасовка совпала"]).sum())
    if mismatch:
        print(f"\nФасовка не совпала у {mismatch} позиций — их надо уточнить у менеджера.")
    mixed = int((~shared["Базис совпал"]).sum())
    if mixed:
        print(f"Разный базис поставки у {mixed} позиций: сравнивать EXW с FOB напрямую "
              f"нельзя, при EXW доставку до порта оплачиваем мы.")
        print(shared.loc[~shared["Базис совпал"], "Базисы"].value_counts().to_string())

    top = shared[shared["Разница, %"] >= min_diff]
    print(f"\nТоп расхождений (от {min_diff}%):")
    print(top.head(15).to_string())
    print(f"\nСохранено: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("brand", help="название бренда, часть тоже подойдет")
    parser.add_argument("--min-diff", type=float, default=0.0, help="порог расхождения, %%")
    ns = parser.parse_args()
    main(ns.brand, ns.min_diff)
