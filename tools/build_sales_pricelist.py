"""Прайс для покупателя: закупочная цена плюс наценка, в вонах.

Одна строка — один товар. Если позиция есть у нескольких поставщиков,
берем более дорогую цену: так наценка не съедается, если закупать
придется не у самого дешевого.

Цена и описание фасовки берутся из одной и той же строки прайса — иначе
можно поставить цену за набор рядом с фасовкой за штуку.

В файл не попадают ни поставщики, ни наша закупочная цена: он уходит
покупателю.

Запуск:
    python3 tools/build_sales_pricelist.py
    python3 tools/build_sales_pricelist.py --markup 15 --brand MEDIPEEL
"""

import argparse
import math
import os

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = "outputs/prices_normalized.xlsx"
OUT_DIR = "outputs"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="B4C6E7")


def round_up(price, step=10):
    """Округляем вверх до десятка вон, чтобы наценка не терялась."""
    return int(math.ceil(price / step) * step)


def build(brand, markup):
    df = pd.read_excel(SRC, dtype={"Штрихкод": str})
    if brand:
        mask = (df["Бренд"].fillna("").str.upper().str.contains(brand.upper())
                | df["Название EN"].fillna("").str.upper().str.contains(brand.upper()))
        df = df[mask]
        if df.empty:
            raise SystemExit(f"Бренд {brand} не найден")

    df = df[df["Закупка, KRW"].notna() & (df["Закупка, KRW"] > 0)]
    # Ключ товара: штрихкод, а без него — название с объемом.
    key = df["Штрихкод"].fillna("").str.replace(r"\D", "", regex=True)
    fallback = df["Название EN"].fillna("") + "|" + df["Объем"].fillna("").astype(str)
    df["Ключ"] = key.where(key.str.len() >= 8, fallback)

    # Более дорогая цена и ее собственное описание фасовки — из одной строки.
    dearest = df.sort_values("Закупка, KRW", ascending=False).drop_duplicates("Ключ")

    out = pd.DataFrame({
        "Barcode": dearest["Штрихкод"],
        "Brand": dearest["Бренд"].fillna("").str.upper(),
        "Product name": dearest["Название EN"].fillna(dearest["Название KR"]),
        "Volume": dearest["Объем"],
        "Unit": dearest["Единица цены"].map({"за шт": "piece", "за набор": "set"}),
        "Pcs in set": dearest["Штук в упаковке"],
        "Qty per box": dearest["Шт/короб"],
        "Price, KRW": [round_up(p * (1 + markup / 100)) for p in dearest["Закупка, KRW"]],
    })
    return out.sort_values(["Brand", "Product name"]).reset_index(drop=True)


def write_excel(table, path, markup, brand):
    title = f"PRICE LIST — {brand.upper()}" if brand else "PRICE LIST — KOREAN COSMETICS"
    with pd.ExcelWriter(path) as writer:
        table.to_excel(writer, sheet_name="PRICE LIST", index=False, startrow=2)
        book, sheet = writer.book, writer.sheets["PRICE LIST"]

        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A2"] = "Currency: KRW. Prices are per unit as stated in the Unit column."
        sheet["A2"].font = Font(italic=True, size=10)

        header_row = 3
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        widths = {"Barcode": 16, "Brand": 20, "Product name": 58, "Volume": 20,
                  "Unit": 10, "Pcs in set": 11, "Qty per box": 12, "Price, KRW": 14,
                  "Order qty": 12, "Amount, KRW": 16}
        titles = [str(c.value) for c in sheet[header_row]]

        # Колонки для заказа: покупатель проставляет количество, сумма считается.
        order_col = len(titles) + 1
        amount_col = order_col + 1
        sheet.cell(header_row, order_col, "Order qty")
        sheet.cell(header_row, amount_col, "Amount, KRW")
        for col in (order_col, amount_col):
            cell = sheet.cell(header_row, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        price_col = titles.index("Price, KRW") + 1
        first, last = header_row + 1, header_row + len(table)
        for row in range(first, last + 1):
            sheet.cell(row, price_col).number_format = "# ##0"
            amount = sheet.cell(row, amount_col)
            amount.value = (f"={get_column_letter(price_col)}{row}"
                            f"*{get_column_letter(order_col)}{row}")
            amount.number_format = "# ##0"
            for col in range(1, amount_col + 1):
                sheet.cell(row, col).border = Border(bottom=THIN)

        total = last + 1
        sheet.cell(total, price_col - 1, "TOTAL").font = Font(bold=True)
        for col, formula in ((order_col, "SUM"), (amount_col, "SUM")):
            cell = sheet.cell(total, col)
            letter = get_column_letter(col)
            cell.value = f"={formula}({letter}{first}:{letter}{last})"
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
            cell.number_format = "# ##0"

        for index in range(1, amount_col + 1):
            name = sheet.cell(header_row, index).value
            sheet.column_dimensions[get_column_letter(index)].width = widths.get(str(name), 14)
        sheet.freeze_panes = f"A{header_row + 1}"


def main(brand, markup):
    if not os.path.exists(SRC):
        raise SystemExit(f"Нет файла {SRC} — сначала запустите parse_price_lists.py")
    table = build(brand, markup)
    suffix = f"_{brand.lower().replace(' ', '_')}" if brand else ""
    path = os.path.join(OUT_DIR, f"sales_pricelist{suffix}.xlsx")
    write_excel(table, path, markup, brand)

    print(f"Позиций в прайсе: {len(table)}   Брендов: {table['Brand'].nunique()}")
    print(f"Наценка: {markup}%   Валюта: KRW")
    print(f"Цена: минимум {table['Price, KRW'].min():,.0f}, "
          f"медиана {table['Price, KRW'].median():,.0f}, "
          f"максимум {table['Price, KRW'].max():,.0f}")
    print(f"Позиций с ценой за набор: {(table['Unit'] == 'set').sum()}")
    print(f"\nСохранено: {path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--markup", type=float, default=10.0, help="наценка в процентах")
    parser.add_argument("--brand", help="только один бренд")
    ns = parser.parse_args()
    main(ns.brand, ns.markup)
