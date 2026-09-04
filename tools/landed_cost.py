"""Себестоимость товара в пути: инвойс плюс растаможка и доставка.

Инвойсная цена — это цена на складе в Корее. Чтобы понять, во что товар
обойдется в Москве, сверху накидываем процент на транспорт, пошлину и
приемку. Для поставки сентября 2026 договорились считать 30%.

Товар, который идет машиной, приходит без цен — только штрихкоды и
количество. Цену для него берем сначала из инвойса того же контейнера,
потом из прайсов поставщиков. Источник цены пишем в отдельной колонке,
чтобы было видно, где цена подтверждена документом, а где взята из прайса.

Запуск:
    python3 tools/landed_cost.py --markup 30
"""

import argparse
import os
import re

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

INVOICE = "data/shipments/2026-09_container_invoice.xlsx"
MANIFEST = "data/shipments/2026-09_truck_manifest.csv"
PRICES = "outputs/prices_normalized.xlsx"
OUT = "outputs/landed_cost_2026-09.xlsx"
KRW_RUB = 0.058

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
ALERT_FILL = PatternFill("solid", fgColor="FFF2CC")


def clean_barcode(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[:13] if len(digits) >= 8 else None


def read_invoice(path):
    """Строки инвойса: товар, штрихкод, количество и цена за штуку."""
    raw = pd.read_excel(path, header=None)
    rows, service = [], []
    for _, line in raw.iloc[3:].iterrows():
        name = line[2]
        if not isinstance(name, str) or not name.strip():
            continue
        quantity = pd.to_numeric(line[4], errors="coerce")
        price = pd.to_numeric(line[5], errors="coerce")
        if pd.isna(quantity) or pd.isna(price):
            continue
        # Справа в инвойсе стоит то, что реально загрузили в контейнер.
        # По Manyo заказ 1200, а погрузили 600 — сумму инвойс при этом
        # посчитал по заказу, поэтому обе цифры держим рядом.
        loaded = pd.to_numeric(line[20], errors="coerce")
        loaded = int(loaded) if pd.notna(loaded) else int(quantity)
        record = {
            "HS-код": line[0],
            "Бренд": str(line[1]).strip() if isinstance(line[1], str) else None,
            "Товар": name.strip(),
            "Штрихкод": clean_barcode(line[3]),
            "Количество": int(quantity),
            "Загружено, шт": loaded,
            "Цена, KRW": float(price),
            "Сумма, KRW": float(pd.to_numeric(line[6], errors="coerce")),
            "Сумма по погрузке, KRW": float(price) * loaded,
        }
        # Погрузка контейнера — услуга, а не товар: наценку на нее не ставим.
        if str(line[0]).strip().lower().startswith("container"):
            service.append(record)
        else:
            rows.append(record)
    return pd.DataFrame(rows), pd.DataFrame(service)


def price_book(invoice):
    """Где брать цену для машины: сначала инвойс, потом прайсы поставщиков."""
    book = {}
    for _, row in invoice.iterrows():
        if row["Штрихкод"]:
            book[row["Штрихкод"]] = (row["Цена, KRW"], "инвойс контейнера")

    prices = pd.read_excel(PRICES, dtype={"Штрихкод": str})
    prices = prices[(prices["Страна"] == "KR") & prices["Закупка, KRW"].notna()]
    prices["Штрихкод"] = prices["Штрихкод"].fillna("").astype(str)
    cheapest = prices.sort_values("Закупка, KRW").drop_duplicates("Штрихкод")
    for _, row in cheapest.iterrows():
        code = clean_barcode(row["Штрихкод"])
        if code and code not in book:
            book[code] = (float(row["Закупка, KRW"]), f"прайс {row['Поставщик']}")
    return book


def read_manifest(path, book):
    rows = pd.read_csv(path, sep=";", dtype={"Штрихкод": str})
    rows["Штрихкод"] = rows["Штрихкод"].map(clean_barcode)
    found = rows["Штрихкод"].map(lambda code: book.get(code, (None, "цены нет")))
    rows["Цена, KRW"] = [item[0] for item in found]
    rows["Источник цены"] = [item[1] for item in found]
    rows["Сумма, KRW"] = rows["Цена, KRW"] * rows["Количество"]
    return rows.rename(columns={"Количество": "Количество", "Товар": "Товар"})


def add_landed(table, markup, rate):
    """Цена с растаможкой и доставкой — та, по которой считаем юнитку."""
    factor = 1 + markup / 100
    table = table.copy()
    table["Цена до Москвы, KRW"] = (table["Цена, KRW"] * factor).round(0)
    table["Себестоимость, руб/шт"] = (table["Цена, KRW"] * factor * rate).round(2)
    table["Сумма до Москвы, KRW"] = (table["Сумма, KRW"] * factor).round(0)
    table["Сумма, руб"] = (table["Сумма, KRW"] * factor * rate).round(0)
    return table


def format_sheet(sheet):
    sheet.freeze_panes = "B2"
    titles = [str(c.value or "") for c in sheet[1]]
    for index, title in enumerate(titles, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = (
            56 if title in ("Товар", "Показатель") else
            20 if title in ("Источник цены", "Бренд", "Значение") else 15)
        if "KRW" in title or "руб" in title or title == "Количество":
            for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                row[0].number_format = "# ##0"
    if "Источник цены" in titles:
        source = titles.index("Источник цены")
        for row in sheet.iter_rows(min_row=2):
            if str(row[source].value or "").startswith("прайс"):
                row[source].fill = ALERT_FILL


def main(markup, rate):
    invoice, service = read_invoice(INVOICE)
    truck = read_manifest(MANIFEST, price_book(invoice))

    container = add_landed(invoice, markup, rate)
    gap = container["Количество"] - container["Загружено, шт"]
    container["Расхождение заказ/погрузка"] = gap.map(
        lambda value: f"недогруз {value} шт" if value else "")
    truck = add_landed(truck, markup, rate)

    by_brand = (container.groupby("Бренд")
                .agg(**{"Позиций": ("Товар", "size"), "Штук": ("Загружено, шт", "sum"),
                        "Сумма инвойса, KRW": ("Сумма, KRW", "sum"),
                        "Сумма до Москвы, KRW": ("Сумма до Москвы, KRW", "sum"),
                        "Сумма, руб": ("Сумма, руб", "sum")})
                .sort_values("Сумма, руб", ascending=False).reset_index())

    loading = float(service["Сумма, KRW"].sum()) if not service.empty else 0.0
    goods = float(container["Сумма, KRW"].sum())
    truck_sum = float(truck["Сумма, KRW"].sum())
    total = [
        {"Показатель": "Наценка на растаможку и доставку, %", "Значение": markup},
        {"Показатель": "Курс, руб за вону", "Значение": rate},
        {"Показатель": "КОНТЕЙНЕР: позиций", "Значение": len(container)},
        {"Показатель": "КОНТЕЙНЕР: заказано, шт", "Значение": int(container["Количество"].sum())},
        {"Показатель": "КОНТЕЙНЕР: погружено, шт", "Значение": int(container["Загружено, шт"].sum())},
        {"Показатель": "КОНТЕЙНЕР: инвойс, KRW", "Значение": round(goods)},
        {"Показатель": "КОНТЕЙНЕР: погрузка контейнера, KRW", "Значение": round(loading)},
        {"Показатель": "КОНТЕЙНЕР: с наценкой, KRW", "Значение": round(goods * (1 + markup / 100))},
        {"Показатель": "КОНТЕЙНЕР: с наценкой, руб",
         "Значение": round(goods * (1 + markup / 100) * rate)},
        {"Показатель": "МАШИНА: позиций", "Значение": len(truck)},
        {"Показатель": "МАШИНА: штук", "Значение": int(truck["Количество"].sum())},
        {"Показатель": "МАШИНА: инвойс, KRW", "Значение": round(truck_sum)},
        {"Показатель": "МАШИНА: с наценкой, KRW", "Значение": round(truck_sum * (1 + markup / 100))},
        {"Показатель": "МАШИНА: с наценкой, руб",
         "Значение": round(truck_sum * (1 + markup / 100) * rate)},
        {"Показатель": "ВСЕГО в пути, руб",
         "Значение": round((goods + truck_sum) * (1 + markup / 100) * rate)},
        {"Показатель": "Недогружено против заказа, шт",
         "Значение": int((container["Количество"] - container["Загружено, шт"]).sum())},
        {"Показатель": "Инвойс выставлен на недогруз, KRW",
         "Значение": round(float(container["Сумма, KRW"].sum()
                                 - container["Сумма по погрузке, KRW"].sum()))},
        {"Показатель": "Позиций без цены",
         "Значение": int(truck["Цена, KRW"].isna().sum())},
    ]

    columns = ["Бренд", "Товар", "Штрихкод", "Количество", "Загружено, шт", "Цена, KRW",
               "Цена до Москвы, KRW", "Себестоимость, руб/шт",
               "Сумма, KRW", "Сумма по погрузке, KRW", "Сумма до Москвы, KRW",
               "Сумма, руб", "Расхождение заказ/погрузка"]
    truck_columns = ["Товар", "Штрихкод", "Количество", "Срок годности", "Источник цены",
                     "Цена, KRW", "Цена до Москвы, KRW", "Себестоимость, руб/шт",
                     "Сумма, KRW", "Сумма до Москвы, KRW", "Сумма, руб"]

    with pd.ExcelWriter(OUT) as writer:
        pd.DataFrame(total).to_excel(writer, sheet_name="ИТОГО", index=False)
        container[columns].to_excel(writer, sheet_name="КОНТЕЙНЕР", index=False)
        truck[truck_columns].to_excel(writer, sheet_name="МАШИНА", index=False)
        by_brand.to_excel(writer, sheet_name="ПО БРЕНДАМ", index=False)
        if not service.empty:
            service.to_excel(writer, sheet_name="УСЛУГИ В ИНВОЙСЕ", index=False)
        for name in writer.book.sheetnames:
            format_sheet(writer.book[name])

    print(pd.DataFrame(total).to_string(index=False))
    print(f"\nПо брендам:\n{by_brand.to_string(index=False)}")
    missing = truck[truck["Цена, KRW"].isna()]
    if not missing.empty:
        print("\nНет цены:")
        print(missing[["Товар", "Штрихкод", "Количество"]].to_string(index=False))
    print(f"\nФайл: {OUT}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Себестоимость товара в пути")
    parser.add_argument("--markup", type=float, default=30.0)
    parser.add_argument("--rate", type=float, default=KRW_RUB)
    args = parser.parse_args()
    main(args.markup, args.rate)
