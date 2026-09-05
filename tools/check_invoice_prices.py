"""Сверка инвойса с прайсами: не переплатили ли мы за закупку.

По каждой позиции инвойса ищем тот же штрихкод в прайсах корейских
поставщиков и считаем, во сколько обошлась бы та же партия у них.
Переплату считаем на фактическое количество — процент без штук ничего не
говорит о деньгах.

Сравниваем цену как она напечатана: приведение к штуке делит цену на
число пэдов в банке у одной стороны и не делит у другой.

Запуск:
    python3 tools/check_invoice_prices.py
"""

import argparse
import os

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from landed_cost import INVOICE, read_invoice
from rates import KRW_RUB
from supplier_brand_matrix import split_conflicts, tokens

PRICES = "outputs/prices_normalized.xlsx"
OUT = "outputs/Сравнение цен купленного товара в Корее.xlsx"
COUNTRY = "KR"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")


def supplier_prices(country):
    df = pd.read_excel(PRICES, dtype={"Штрихкод": str})
    df = df[(df["Страна"] == country) & df["Закупка, KRW"].notna()]
    df = df[df["Закупка, KRW"] > 0].copy()
    df["Штрихкод"] = df["Штрихкод"].fillna("").astype(str).str.replace(r"\D", "", regex=True)
    df = df[df["Штрихкод"].str.len() >= 8]
    # Сначала выкидываем строки, где под одним штрихкодом лежит другой
    # товар: в прайсе SKIN1004 у Papa Cosmetic стоят чужие штрихкоды, и
    # без этого дешевая чужая позиция притворяется выгодным предложением.
    df["Бренд в прайсе"] = df["Бренд"]
    df, _ = split_conflicts(df)
    # У одного поставщика товар может встретиться в двух прайсах — берем дешевле.
    return df.sort_values("Закупка, KRW").drop_duplicates(["Штрихкод", "Поставщик"])


def same_goods(our_name, their_name):
    """Один штрихкод на разные товары бывает: сверяем название.

    Одного общего слова мало: «CICA» есть и в PINE CALMING CICA BODY WASH,
    и в SKIN1004 Probio-Cica Cream. Просим два совпадения там, где слов
    хватает.
    """
    ours, theirs = tokens(our_name), tokens(their_name)
    if not ours or not theirs:
        return True
    common = ours & theirs
    if min(len(ours), len(theirs)) <= 2:
        return bool(common)
    return len(common) >= 2


def with_total(table, label_column, sums, label="ИТОГО"):
    """Дописывает строку ИТОГО: без нее таблицу приходится считать руками."""
    footer = {column: None for column in table.columns}
    footer[label_column] = label
    for column in sums:
        if column in table.columns:
            footer[column] = table[column].sum()
    return pd.concat([table, pd.DataFrame([footer])], ignore_index=True)


def main(country):
    invoice, _ = read_invoice(INVOICE)
    invoice = invoice[invoice["Штрихкод"].notna()].copy()
    market = supplier_prices(country)
    suppliers = sorted(market["Поставщик"].unique())

    prices = market.pivot(index="Штрихкод", columns="Поставщик", values="Закупка, KRW")
    names = market.pivot(index="Штрихкод", columns="Поставщик", values="Название EN")

    rows = []
    for _, item in invoice.iterrows():
        code = item["Штрихкод"]
        offers = {}
        if code in prices.index:
            for supplier in suppliers:
                price = prices.loc[code, supplier]
                if pd.isna(price):
                    continue
                if not same_goods(item["Товар"], names.loc[code, supplier]):
                    continue
                offers[supplier] = float(price)

        record = {
            "Бренд": item["Бренд"],
            "Товар": item["Товар"],
            "Штрихкод": code,
            "Куплено, шт": item["Загружено, шт"],
            "Наша цена, KRW": item["Цена, KRW"],
            "Сумма закупки, KRW": item["Цена, KRW"] * item["Загружено, шт"],
        }
        record.update({supplier: offers.get(supplier) for supplier in suppliers})
        if offers:
            best = min(offers, key=offers.get)
            record["Дешевле всех"] = best
            record["Лучшая цена, KRW"] = offers[best]
            record["Разница на штуке, KRW"] = round(item["Цена, KRW"] - offers[best])
            record["Мы дороже на, %"] = round(
                (item["Цена, KRW"] / offers[best] - 1) * 100, 1)
            record["Переплата на партии, KRW"] = round(
                record["Разница на штуке, KRW"] * item["Загружено, шт"])
            record["Переплата на партии, руб"] = round(
                record["Переплата на партии, KRW"] * KRW_RUB)
        rows.append(record)

    table = pd.DataFrame(rows)
    found = table[table["Дешевле всех"].notna()].copy()
    missing = table[table["Дешевле всех"].isna()][
        ["Бренд", "Товар", "Штрихкод", "Куплено, шт", "Наша цена, KRW"]]

    # Переплата — только там, где чужая цена ниже нашей.
    overpaid = found[found["Переплата на партии, KRW"] > 0].sort_values(
        "Переплата на партии, руб", ascending=False)

    by_supplier = (overpaid.groupby("Дешевле всех")
                   .agg(**{"Позиций": ("Товар", "size"),
                           "Переплата, руб": ("Переплата на партии, руб", "sum")})
                   .reset_index().sort_values("Переплата, руб", ascending=False))
    by_supplier = with_total(by_supplier, "Дешевле всех", ["Позиций", "Переплата, руб"])

    # Свод по брендам: где деньги, а где просто проценты.
    by_brand = (found.groupby("Бренд")
                .agg(**{"Позиций": ("Товар", "size"),
                        "Куплено, шт": ("Куплено, шт", "sum"),
                        "Сумма закупки, руб": ("Сумма закупки, KRW",
                                               lambda values: round(values.sum() * KRW_RUB)),
                        "Переплата, руб": ("Переплата на партии, руб",
                                           lambda values: int(values[values > 0].sum()))})
                .reset_index())
    by_brand["Переплата к закупке, %"] = (by_brand["Переплата, руб"] /
                                          by_brand["Сумма закупки, руб"] * 100).round(1)
    by_brand = by_brand.sort_values("Переплата, руб", ascending=False)
    by_brand = with_total(by_brand, "Бренд",
                          ["Позиций", "Куплено, шт", "Сумма закупки, руб", "Переплата, руб"])
    by_brand.iloc[-1, by_brand.columns.get_loc("Переплата к закупке, %")] = round(
        by_brand["Переплата, руб"].iloc[-1] / by_brand["Сумма закупки, руб"].iloc[-1] * 100, 1)

    # Весь инвойс одной таблицей, снизу строка ИТОГО.
    whole = found.sort_values("Переплата на партии, руб", ascending=False).copy()
    whole["Сумма закупки, руб"] = (whole["Сумма закупки, KRW"] * KRW_RUB).round(0)
    # Там, где дешевле нашей цены никто не дает, переплаты нет — не пишем
    # в этой колонке отрицательные числа, они читаются как скидка.
    whole.loc[whole["Переплата на партии, руб"] <= 0, "Переплата на партии, руб"] = 0
    footer = {column: None for column in whole.columns}
    footer.update({
        "Бренд": "ИТОГО",
        "Куплено, шт": whole["Куплено, шт"].sum(),
        "Сумма закупки, KRW": whole["Сумма закупки, KRW"].sum(),
        "Сумма закупки, руб": whole["Сумма закупки, руб"].sum(),
        "Переплата на партии, руб": int(overpaid["Переплата на партии, руб"].sum()),
        "Дешевле всех": f"переплата {int(overpaid['Переплата на партии, руб'].sum())} руб",
    })
    whole = pd.concat([whole, pd.DataFrame([footer])], ignore_index=True)

    paid = with_total(overpaid.copy(), "Бренд",
                      ["Куплено, шт", "Переплата на партии, KRW", "Переплата на партии, руб"])
    absent = with_total(missing.copy(), "Бренд", ["Куплено, шт"])

    total = pd.DataFrame([
        {"Показатель": "Позиций в инвойсе", "Значение": len(table)},
        {"Показатель": "Нашлось в прайсах поставщиков", "Значение": len(found)},
        {"Показатель": "Не с чем сравнить", "Значение": len(missing)},
        {"Показатель": "Позиций, где мы купили дороже", "Значение": len(overpaid)},
        {"Показатель": "Позиций, где дешевле нашей цены нет",
         "Значение": len(found) - len(overpaid)},
        {"Показатель": "Переплата всего, KRW",
         "Значение": int(overpaid["Переплата на партии, KRW"].sum())},
        {"Показатель": "Переплата всего, руб",
         "Значение": int(overpaid["Переплата на партии, руб"].sum())},
        {"Показатель": "Сумма закупки по сравнимым позициям, KRW",
         "Значение": int(found["Сумма закупки, KRW"].sum())},
        {"Показатель": "Переплата к сравнимой закупке, %",
         "Значение": round(overpaid["Переплата на партии, KRW"].sum()
                           / found["Сумма закупки, KRW"].sum() * 100, 1)},
    ])

    with pd.ExcelWriter(OUT) as writer:
        total.to_excel(writer, sheet_name="ИТОГО", index=False)
        whole.to_excel(writer, sheet_name="ВЕСЬ ИНВОЙС", index=False)
        by_brand.to_excel(writer, sheet_name="ПО БРЕНДАМ", index=False)
        by_supplier.to_excel(writer, sheet_name="У КОГО ДЕШЕВЛЕ", index=False)
        paid.to_excel(writer, sheet_name="ПЕРЕПЛАТИЛИ", index=False)
        absent.to_excel(writer, sheet_name="НЕ С ЧЕМ СРАВНИТЬ", index=False)

        for name in writer.book.sheetnames:
            sheet = writer.book[name]
            sheet.freeze_panes = "B2"
            titles = [str(c.value or "") for c in sheet[1]]
            for index, title in enumerate(titles, start=1):
                cell = sheet.cell(row=1, column=index)
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                sheet.column_dimensions[get_column_letter(index)].width = (
                    52 if title == "Товар" else 40 if title == "Показатель" else
                    18 if title in ("Дешевле всех", "Бренд") else 14)
                if "KRW" in title or "руб" in title or "шт" in title:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = "# ##0"
                if "%" in title:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = "0.0"
            last = sheet.cell(row=sheet.max_row, column=1).value
            if isinstance(last, str) and last.startswith("ИТОГО"):
                for cell in sheet[sheet.max_row]:
                    cell.font = Font(bold=True)
                    cell.fill = HEADER_FILL
            if "Переплата на партии, руб" in titles:
                column = titles.index("Переплата на партии, руб")
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row - 1):
                    value = row[column].value
                    if isinstance(value, (int, float)) and value > 0:
                        row[column].fill = ALERT_FILL
                    elif isinstance(value, (int, float)) and value < 0:
                        row[column].fill = GOOD_FILL
            if "Мы дороже на, %" in titles and sheet.max_row > 1:
                letter = get_column_letter(titles.index("Мы дороже на, %") + 1)
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{sheet.max_row}",
                    ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                                   mid_type="num", mid_value=10, mid_color="FFEB84",
                                   end_type="num", end_value=30, end_color="F8696B"))

    print(total.to_string(index=False))
    print("\nУ кого дешевле:")
    print(by_supplier.to_string(index=False))
    print(f"\nФайл: {OUT}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Сверка инвойса с прайсами")
    parser.add_argument("--country", default=COUNTRY)
    main(parser.parse_args().country)
