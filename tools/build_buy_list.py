"""Список закупа: какой бренд и какой товар у какого поставщика брать.

Берем цены за штуку в вонах внутри одной страны, сводим по штрихкоду и по
каждой позиции показываем, у кого дешевле и на сколько процентов дороже
следующий. Отдельно даем план по брендам: заказ обычно собирают целиком у
одного поставщика, а не по одной банке у каждого.

Запуск:
    python3 tools/build_buy_list.py
"""

import argparse
import os

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from price_unit import unify_packs
from supplier_brand_matrix import (IMPORT, MIN_COVERAGE, RATE, brand_stats, load,
                                   price_table, split_conflicts)

OUT_DIR = "outputs"
COLUMN = "Цена за штуку (сводно)"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
ONLY_FILL = PatternFill("solid", fgColor="FFF2CC")
ALERT_FILL = PatternFill("solid", fgColor="FFC7CE")


def positions(table, prices, suppliers):
    """По каждому товару: у кого дешевле, кто второй и на сколько дороже."""
    columns = [s for s in suppliers if s in prices.columns]
    rows = table.copy()
    values = prices[columns]

    rows["Брать у"] = values.idxmin(axis=1)
    rows["Цена, KRW/шт"] = values.min(axis=1).round(0)
    rows["Себестоимость, руб/шт"] = (rows["Цена, KRW/шт"] * RATE * IMPORT).round(0)
    rows["Поставщиков"] = values.notna().sum(axis=1)

    # Второй по цене нужен как запасной вариант и как мера риска по цене.
    # Там, где поставщик один, второго нет — idxmin по пустой строке падает.
    ranked = values.rank(axis=1, method="first")
    second = values.where(ranked == 2)
    has_second = second.notna().any(axis=1)
    rows["Второй поставщик"] = pd.NA
    rows.loc[has_second, "Второй поставщик"] = second[has_second].idxmin(axis=1)
    rows["Цена второго, KRW"] = second.min(axis=1).round(0)
    rows["Второй дороже на, %"] = (
        (rows["Цена второго, KRW"] / rows["Цена, KRW/шт"] - 1) * 100).round(1)
    rows["Пометка"] = ""
    rows.loc[rows["Поставщиков"] == 1, "Пометка"] = "единственный поставщик"
    rows.loc[rows["Второй дороже на, %"] >= 30, "Пометка"] = "разница 30%+, сверить"
    return rows


def brand_plan(brands, df, suppliers):
    """План по брендам: где заказывать бренд целиком."""
    catalog = df.groupby(["Бренд", "Поставщик"])["Штрихкод"].nunique().unstack(fill_value=0)
    catalog = catalog.reindex(columns=suppliers, fill_value=0)
    leaders = brands.set_index("Бренд")

    rows = []
    for brand, counts in catalog.iterrows():
        present = counts[counts > 0]
        if brand in leaders.index:
            record = leaders.loc[brand]
            supplier = record["Дешевле всех"]
            rows.append({
                "Бренд": brand,
                "Заказывать у": supplier,
                "SKU у него": int(counts[supplier]),
                "SKU всего у всех": int(present.sum()),
                "Поставщиков с брендом": int(len(present)),
                "Сравнимых позиций": int(record["Позиций в сравнении"]),
                "Покрытие лидера, %": record["Покрытие лидера, %"],
                "Второй по цене": record["Второй по цене"],
                "Второй дороже, %": record["Второй дороже, % (корзина)"],
                "Переплата, если брать не у лидера, руб":
                    record["Переплата, если брать не у лидера, руб"],
                "Основание": "дешевле по сравнению цен",
            })
        else:
            supplier = present.index[0]
            rows.append({
                "Бренд": brand,
                "Заказывать у": supplier,
                "SKU у него": int(present.iloc[0]),
                "SKU всего у всех": int(present.sum()),
                "Поставщиков с брендом": int(len(present)),
                "Сравнимых позиций": 0,
                "Покрытие лидера, %": 100,
                "Второй по цене": None,
                "Второй дороже, %": None,
                "Переплата, если брать не у лидера, руб": 0,
                "Основание": ("бренд есть только у него" if len(present) == 1
                              else "нет общих штрихкодов, цены не сравнить"),
            })
    plan = pd.DataFrame(rows)
    return plan.sort_values(["Сравнимых позиций", "SKU всего у всех"], ascending=False)


def supplier_plan(plan, buy, suppliers):
    """Свод: что и на сколько позиций заказываем у каждого поставщика.

    Поставщика без единого бренда в плане не прячем: ноль в строке — это
    и есть ответ, брать у него нечего.
    """
    rows = []
    for supplier in suppliers:
        group = plan[plan["Заказывать у"] == supplier]
        items = buy[buy["Брать у"] == supplier]
        rows.append({
            "Поставщик": supplier,
            "Брендов забираем": len(group),
            "SKU по плану брендов": int(group["SKU у него"].sum()),
            "Позиций, где он дешевле всех": len(items),
            "Из них единственный поставщик":
                int((items["Пометка"] == "единственный поставщик").sum()),
            "Бренды": ", ".join(sorted(group["Бренд"])[:12]) +
                      (" ..." if len(group) > 12 else ""),
        })
    return pd.DataFrame(rows).sort_values("SKU по плану брендов", ascending=False)


def format_sheet(worksheet):
    worksheet.freeze_panes = "B2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    titles = [str(c.value or "") for c in worksheet[1]]
    for index, title in enumerate(titles, start=1):
        letter = get_column_letter(index)
        if title == "Название EN":
            width = 50
        elif title == "Бренды":
            width = 60
        elif title in ("Бренд", "Брать у", "Заказывать у", "Второй поставщик",
                       "Поставщик", "Пометка", "Основание"):
            width = 24
        else:
            width = 13
        worksheet.column_dimensions[letter].width = width
        fmt = "0.0" if "%" in title else ("# ##0" if ("руб" in title or "KRW" in title) else None)
        if fmt:
            for row in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
                row[0].number_format = fmt


def main(country):
    df = load(country)
    suppliers = sorted(df["Поставщик"].unique())
    df, bad = split_conflicts(df)
    df, _ = unify_packs(df)

    table, prices = price_table(df, COLUMN)
    info = table[["Бренд", "Название EN", "Объем"]]
    brands, _, _ = brand_stats(prices, info, suppliers)

    buy = positions(table, prices, suppliers)
    plan = brand_plan(brands, df, suppliers)
    by_supplier = supplier_plan(plan, buy, suppliers)

    # В таблицу товаров подставляем рекомендацию по бренду: заказ обычно
    # собирают целиком у одного поставщика.
    recommended = plan.set_index("Бренд")["Заказывать у"]
    buy.insert(buy.columns.get_loc("Брать у"), "Бренд заказываем у",
               buy["Бренд"].map(recommended))
    buy["Совпадает с планом бренда"] = (
        buy["Брать у"] == buy["Бренд заказываем у"]).map({True: "да", False: "нет"})
    columns = ["Бренд", "Название EN", "Объем", "Бренд заказываем у", "Брать у",
               "Совпадает с планом бренда", "Цена, KRW/шт", "Себестоимость, руб/шт",
               "Второй поставщик", "Цена второго, KRW", "Второй дороже на, %",
               "Поставщиков", "Пометка"]
    buy = buy[columns].sort_values(["Бренд", "Название EN"])

    check = buy[buy["Пометка"] == "разница 30%+, сверить"]

    out_path = os.path.join(OUT_DIR, f"buy_list_{country.lower()}.xlsx")
    with pd.ExcelWriter(out_path) as writer:
        plan.to_excel(writer, sheet_name="ПЛАН ПО БРЕНДАМ", index=False)
        by_supplier.to_excel(writer, sheet_name="ПО ПОСТАВЩИКАМ", index=False)
        buy.to_excel(writer, sheet_name="ЧТО ПОКУПАТЬ")
        check.to_excel(writer, sheet_name="НА СВЕРКУ")
        if not bad.empty:
            bad.to_excel(writer, sheet_name="СПОРНЫЕ ШТРИХКОДЫ", index=False)

        book = writer.book
        for name in book.sheetnames:
            format_sheet(book[name])

        titles = [str(c.value or "") for c in book["ПЛАН ПО БРЕНДАМ"][1]]
        letter = get_column_letter(titles.index("Второй дороже, %") + 1)
        book["ПЛАН ПО БРЕНДАМ"].conditional_formatting.add(
            f"{letter}2:{letter}{len(plan) + 1}",
            ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                           mid_type="num", mid_value=8, mid_color="FFEB84",
                           end_type="num", end_value=25, end_color="F8696B"))

        sheet = book["ЧТО ПОКУПАТЬ"]
        titles = [str(c.value or "") for c in sheet[1]]
        mark = titles.index("Пометка")
        for row in sheet.iter_rows(min_row=2):
            note = row[mark].value
            if note == "единственный поставщик":
                row[mark].fill = ONLY_FILL
            elif note:
                row[mark].fill = ALERT_FILL

    print(f"Страна {country}: {len(buy)} товаров, {len(plan)} брендов, "
          f"{len(suppliers)} поставщиков")
    print(by_supplier.drop(columns="Бренды").to_string(index=False))
    print(f"\nНа сверку: {len(check)} позиций. Файл: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Список закупа по брендам и товарам")
    parser.add_argument("--country", default="KR")
    main(parser.parse_args().country)
