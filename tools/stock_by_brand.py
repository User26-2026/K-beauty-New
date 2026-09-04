"""Остатки по брендам и позициям: что ходовое, а что лежит.

Количество берем из плана распределения — там свежий склад и артикул WB.
Себестоимость — из файла остатков, а где не нашлось, из отчета WB.
Скорость продаж — из отчета по выкупам: сколько штук ушло за период.

Запас в месяцах = склад / продажи в месяц. Это главная цифра: она сразу
показывает, где деньги стоят мертво, а где товар вот-вот кончится.

Продажи только по Wildberries и только за период отчета — Озон и опт в
эту скорость не входят.

Запуск:
    python3 tools/stock_by_brand.py
"""

import argparse
import os
import re

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import name_match

PLAN = "data/shipments/2026-09_plan_raspredeleniya.xlsx"
STOCK = "data/stock_costs/Остатки_31.08.2026.xlsx"
SALES = "workspace/wb_audit/sales_profit_with_cogs_2026-04-01_2026-05-24.csv"
SALES_DAYS = 54
OUT = "outputs/stock_by_brand.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
DEAD_FILL = PatternFill("solid", fgColor="FFC7CE")
HOT_FILL = PatternFill("solid", fgColor="C6EFCE")

# Один бренд пишут по-разному в трех документах — сводим к одному виду.
BRAND_FIX = {
    "FARM STAY": "FARMSTAY", "ROUND LAB": "ROUND LAB", "ETUDE HOUSE": "ETUDE",
    "DR.JART+": "DR. JART", "SOME BY MI": "SOME BY MI", "JMSOLUTION": "JMSOLUTION",
    "MANYO FACTORY": "MANYO", "I'M SORRY FOR MY SKIN": "IM SORRY FOR MY SKIN",
}
KNOWN = ["CELIMAX", "FARMSTAY", "ROUND LAB", "PETITFEE", "JIGOTT", "MANYO", "ETUDE",
         "ENOUGH", "JMSOLUTION", "ELIZAVECCA", "DEOPROCE", "LEBELAGE", "COSRX",
         "SKIN1004", "MEDICUBE", "DERMA FACTORY", "HEIMISH", "EKEL", "DR. JART",
         "SOME BY MI", "THE ORDINARY", "MASIL", "LANEIGE", "VT"]


def clean_brand(value):
    text = re.sub(r"\s+", " ", str(value or "").strip().upper())
    return BRAND_FIX.get(text, text)


def brand_from_name(name):
    text = str(name).upper()
    for brand in sorted(KNOWN, key=len, reverse=True):
        if text.startswith(brand) or f" {brand}" in text[:40]:
            return brand
    return text.split()[0] if text.split() else "БЕЗ БРЕНДА"


def read_plan(path):
    plan = pd.read_excel(path, header=0)
    plan = plan[plan["Артикул WB"].notna()].copy()
    plan["Артикул WB"] = pd.to_numeric(plan["Артикул WB"], errors="coerce").astype("Int64")
    for column in ("Склад, шт", "Резерв ВБ, шт", "Резерв Озон, шт",
                   "Предложить Егору, шт", "Егор уже брал, шт"):
        plan[column] = pd.to_numeric(plan[column], errors="coerce").fillna(0)
    return plan.reset_index(drop=True)


def read_stock(path):
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = []
    for line in range(3, sheet.max_row + 1):
        name = sheet.cell(line, 1).value
        if not name:
            continue
        rows.append({"Товар в остатках": str(name).strip(),
                     "Себестоимость из остатков": sheet.cell(line, 4).value,
                     "Срок годности": sheet.cell(line, 5).value})
    book.close()
    return pd.DataFrame(rows)


BAD_CHARS = re.compile(r"[\\/*?:\[\]]")


def sheet_name(brand, used):
    """Имя вкладки: Excel не берет длиннее 31 символа и часть знаков."""
    name = BAD_CHARS.sub(" ", str(brand)).strip()[:31] or "БЕЗ БРЕНДА"
    if name in used:
        name = f"{name[:28]}_{len(used)}"
    used.add(name)
    return name


def brand_sheet(rows, money_total, units_total):
    """Лист одного бренда: его товары, деньги и доля в общем складе."""
    columns = ["Наименование", "Артикул WB", "Склад, шт", "Себестоимость, руб",
               "Остаток, руб", "Доля в бренде, %", "Доля в складе, %",
               "Продажи, шт/мес", "Запас, месяцев", "Статус", "Приход в пути, шт"]
    table = rows.copy()
    brand_money = table["Остаток, руб"].sum()
    table["Доля в бренде, %"] = (table["Остаток, руб"] / brand_money * 100).round(1) \
        if brand_money else 0.0
    table["Доля в складе, %"] = (table["Остаток, руб"] / money_total * 100).round(1)
    table = table[columns].sort_values("Остаток, руб", ascending=False)

    total = {column: None for column in columns}
    total.update({
        "Наименование": "ИТОГО ПО БРЕНДУ",
        "Склад, шт": table["Склад, шт"].sum(),
        "Остаток, руб": brand_money,
        "Доля в бренде, %": 100.0,
        "Доля в складе, %": round(brand_money / money_total * 100, 1),
        "Продажи, шт/мес": round(table["Продажи, шт/мес"].sum(), 1),
        "Приход в пути, шт": table["Приход в пути, шт"].sum(),
        "Статус": f"{table['Склад, шт'].sum() / units_total * 100:.1f}% всех штук склада",
    })
    return pd.concat([table, pd.DataFrame([total])], ignore_index=True)


def status(cover, sold, in_report):
    """Ходовой товар мерим запасом в месяцах, а не выручкой.

    Позицию, которой нет в отчете продаж, в неликвид не записываем: скорее
    всего карточка появилась уже после периода отчета.
    """
    if not in_report:
        return "нет данных о продажах"
    if sold <= 0:
        return "продаж не было"
    if cover is None or pd.isna(cover):
        return "нет данных о продажах"
    if cover > 12:
        return "неликвид, запас больше года"
    if cover > 6:
        return "затоварен"
    if cover >= 2:
        return "нормальный запас"
    return "ходовой, запас кончается"


def main():
    plan = read_plan(PLAN)
    stock = read_stock(STOCK)
    sales = pd.read_csv(SALES)
    sales["Артикул WB"] = pd.to_numeric(sales["Артикул WB"], errors="coerce").astype("Int64")
    sales = sales.groupby("Артикул WB", as_index=False).agg(
        **{"Продано, шт": ("Выкупы, шт", "sum"), "Продано, руб": ("Выкупы, ₽", "sum"),
           "Себестоимость из отчета": ("cost", "max"), "Бренд из отчета": ("Бренд", "first")})

    table = plan.merge(sales, on="Артикул WB", how="left")

    pairs = name_match.match(table["Наименование"], stock["Товар в остатках"])
    table["_pair"] = table.index.map(pairs.get)
    table = table.merge(stock, left_on="_pair", right_index=True, how="left").drop(columns="_pair")

    table["Себестоимость, руб"] = pd.to_numeric(
        table["Себестоимость из остатков"], errors="coerce").fillna(
        pd.to_numeric(table["Себестоимость из отчета"], errors="coerce"))
    table["Бренд"] = table["Бренд из отчета"].map(clean_brand).replace("", pd.NA)
    table["Бренд"] = table["Бренд"].fillna(table["Наименование"].map(brand_from_name))
    table.loc[~table["Бренд"].isin(KNOWN), "Бренд"] = \
        table.loc[~table["Бренд"].isin(KNOWN), "Наименование"].map(brand_from_name)

    table["Остаток, руб"] = (table["Склад, шт"] * table["Себестоимость, руб"]).round(0)
    table["Продано, шт"] = table["Продано, шт"].fillna(0)
    table["Продажи, шт/мес"] = (table["Продано, шт"] * 30 / SALES_DAYS).round(1)
    table["Запас, месяцев"] = (table["Склад, шт"] / table["Продажи, шт/мес"]
                              ).replace([float("inf")], float("nan")).round(1)
    table["Есть в отчете продаж"] = table["Бренд из отчета"].notna()
    table["Статус"] = [status(cover, sold, seen) for cover, sold, seen
                       in zip(table["Запас, месяцев"], table["Продано, шт"],
                              table["Есть в отчете продаж"])]
    table.loc[table["Себестоимость, руб"].isna(), "Статус"] = "нет себестоимости"
    table["Приход в пути, шт"] = (pd.to_numeric(table["Машина"], errors="coerce").fillna(0)
                                  + pd.to_numeric(table["Контейнер"], errors="coerce").fillna(0))

    dead = table["Статус"].isin(["продаж не было", "неликвид, запас больше года"])
    table["Деньги в неликвиде, руб"] = table["Остаток, руб"].where(dead, 0)

    by_brand = (table.groupby("Бренд")
                .agg(**{"Позиций": ("Наименование", "size"),
                        "Штук на складе": ("Склад, шт", "sum"),
                        "Остаток, руб": ("Остаток, руб", "sum"),
                        "Продажи, шт/мес": ("Продажи, шт/мес", "sum"),
                        "Продано за период, руб": ("Продано, руб", "sum"),
                        "Деньги в неликвиде, руб": ("Деньги в неликвиде, руб", "sum"),
                        "Приход в пути, шт": ("Приход в пути, шт", "sum")})
                .reset_index())
    by_brand["Доля остатка, %"] = (by_brand["Остаток, руб"] /
                                   by_brand["Остаток, руб"].sum() * 100).round(1)
    by_brand["Доля по штукам, %"] = (by_brand["Штук на складе"] /
                                     by_brand["Штук на складе"].sum() * 100).round(1)
    by_brand["Запас, месяцев"] = (by_brand["Штук на складе"] / by_brand["Продажи, шт/мес"]
                                 ).replace([float("inf")], float("nan")).round(1)
    by_brand["Неликвид, % остатка"] = (by_brand["Деньги в неликвиде, руб"] /
                                       by_brand["Остаток, руб"] * 100).round(1)
    by_brand = by_brand[["Бренд", "Позиций", "Штук на складе", "Доля по штукам, %",
                         "Остаток, руб", "Доля остатка, %", "Продажи, шт/мес",
                         "Запас, месяцев", "Продано за период, руб",
                         "Деньги в неликвиде, руб", "Неликвид, % остатка",
                         "Приход в пути, шт"]]
    by_brand = by_brand.sort_values("Остаток, руб", ascending=False)

    columns = ["Бренд", "Наименование", "Артикул WB", "Склад, шт", "Себестоимость, руб",
               "Остаток, руб", "Продано, шт", "Продажи, шт/мес", "Продано, руб",
               "Запас, месяцев", "Статус", "Приход в пути, шт", "Резерв ВБ, шт",
               "Резерв Озон, шт", "Егор уже брал, шт", "Срок годности"]
    items = table[columns].sort_values("Остаток, руб", ascending=False)
    dead_items = items[items["Статус"].isin(["продаж не было", "неликвид, запас больше года"])]
    hot_items = items[items["Статус"] == "ходовой, запас кончается"].sort_values("Продажи, шт/мес",
                                                                                 ascending=False)

    total = pd.DataFrame([
        {"Показатель": "Позиций на складе", "Значение": len(items)},
        {"Показатель": "Штук", "Значение": int(items["Склад, шт"].sum())},
        {"Показатель": "Остаток по себестоимости, руб", "Значение": round(items["Остаток, руб"].sum())},
        {"Показатель": "Продажи WB, шт/мес", "Значение": round(items["Продажи, шт/мес"].sum())},
        {"Показатель": "Запас всего склада, месяцев",
         "Значение": round(items["Склад, шт"].sum() / items["Продажи, шт/мес"].sum(), 1)},
        {"Показатель": "Неликвид и позиции без продаж, шт позиций", "Значение": len(dead_items)},
        {"Показатель": "Позиций без данных о продажах",
         "Значение": int((items["Статус"] == "нет данных о продажах").sum())},
        {"Показатель": "Позиций без себестоимости",
         "Значение": int((items["Статус"] == "нет себестоимости").sum())},
        {"Показатель": "Денег в неликвиде, руб", "Значение": round(dead_items["Остаток, руб"].sum())},
        {"Показатель": "Ходовых позиций с запасом меньше 2 месяцев", "Значение": len(hot_items)},
        {"Показатель": "Период отчета продаж, дней", "Значение": SALES_DAYS},
    ])

    money_total = float(items["Остаток, руб"].sum())
    units_total = float(items["Склад, шт"].sum())

    with pd.ExcelWriter(OUT) as writer:
        total.to_excel(writer, sheet_name="ИТОГО", index=False)
        by_brand.to_excel(writer, sheet_name="ПО БРЕНДАМ", index=False)
        # Дальше вкладка на каждый бренд, по убыванию денег в остатке.
        used = {"ИТОГО", "ПО БРЕНДАМ"}
        for brand in by_brand["Бренд"]:
            rows = items[items["Бренд"] == brand]
            brand_sheet(rows, money_total, units_total).to_excel(
                writer, sheet_name=sheet_name(brand, used), index=False)
        items.to_excel(writer, sheet_name="ПО ПОЗИЦИЯМ", index=False)
        dead_items.to_excel(writer, sheet_name="НЕ ХОДОВОЙ", index=False)
        hot_items.to_excel(writer, sheet_name="ХОДОВОЙ", index=False)

        book = writer.book
        for name in book.sheetnames:
            sheet = book[name]
            sheet.freeze_panes = "B2"
            titles = [str(c.value or "") for c in sheet[1]]
            for index, title in enumerate(titles, start=1):
                cell = sheet.cell(row=1, column=index)
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                sheet.column_dimensions[get_column_letter(index)].width = (
                    58 if title in ("Наименование", "Показатель") else
                    26 if title in ("Статус", "Бренд") else 13)
                if "%" in title:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = "0.0"
                if "руб" in title or "шт" in title:
                    for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                        row[0].number_format = "# ##0"
            # Строку ИТОГО в листе бренда выделяем жирным.
            last = sheet.cell(row=sheet.max_row, column=1).value
            if isinstance(last, str) and last.startswith("ИТОГО"):
                for cell in sheet[sheet.max_row]:
                    cell.font = Font(bold=True)
            if "Статус" in titles:
                column = titles.index("Статус")
                for row in sheet.iter_rows(min_row=2):
                    value = row[column].value
                    if value in ("продаж не было", "неликвид, запас больше года"):
                        row[column].fill = DEAD_FILL
                    elif value == "ходовой, запас кончается":
                        row[column].fill = HOT_FILL
            if "Запас, месяцев" in titles:
                letter = get_column_letter(titles.index("Запас, месяцев") + 1)
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{sheet.max_row}",
                    ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                                   mid_type="num", mid_value=6, mid_color="FFEB84",
                                   end_type="num", end_value=24, end_color="F8696B"))

    print(total.to_string(index=False))
    print("\nПо брендам:")
    print(by_brand.to_string(index=False))
    print(f"\nФайл: {OUT}")


if __name__ == "__main__":
    argparse.ArgumentParser(description="Остатки по брендам и ходовость").parse_args()
    main()
