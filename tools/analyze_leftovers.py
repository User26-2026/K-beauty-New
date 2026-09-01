"""Разбор остатка, который не забирает покупатель.

Зеленая колонка в файле остатков — то, что клиент забирает, желтая — то,
что остается на складе и требует сбыта. По каждой оставшейся позиции
показываем, сколько за нее дают российские оптовики: это потолок цены, от
которого имеет смысл торговаться.

Штрихкода в файле остатков нет, поэтому связываем через действующий прайс
по наименованию, а уже из прайса берем штрихкод.

Запуск:
    python3 tools/analyze_leftovers.py --stock <файл> --price <прайс>
"""

import argparse
import os
import re

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PRICES = "outputs/prices_normalized.xlsx"
OUT = "outputs/leftovers.xlsx"
RU_SUPPLIERS = ("KEAUTY", "KoreaTrade", "SAFIYA")


def read_stock(path):
    """Остатки: сколько забирают, сколько остается, по какой себестоимости."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        rows.append({
            "Товар": str(name).strip(),
            "Остаётся, шт": ws.cell(r, 2).value or 0,
            "Всего, шт": ws.cell(r, 3).value or 0,
            "Себестоимость, руб": ws.cell(r, 4).value,
            "Срок годности": ws.cell(r, 5).value,
            "Заберут, шт": ws.cell(r, 6).value or 0,
        })
    wb.close()
    table = pd.DataFrame(rows)
    for column in ("Остаётся, шт", "Всего, шт", "Заберут, шт", "Себестоимость, руб"):
        table[column] = pd.to_numeric(table[column], errors="coerce")
    return table


def key(name):
    """Ключ сопоставления: слова названия плюс все числа из него.

    Без чисел в ключе разные фасовки одного товара сливаются: у CELIMAX
    есть и тонер-пэды на 10 штук, и на 60, названия у них совпадают до
    самого размера.
    """
    text = re.sub(r"[^0-9a-zа-я ]", " ", str(name).lower())
    words = text.split()
    numbers = sorted(w for w in words if any(c.isdigit() for c in w))
    return " ".join(words[:6] + numbers)


def barcodes_from_pricelist(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for _, name, barcode, _, price in ws.iter_rows(min_row=6, max_col=5, values_only=True):
        code = re.sub(r"\D", "", str(barcode or ""))[:13]
        if len(code) >= 8 and name:
            rows.append({"Ключ": key(name), "Штрихкод": code,
                         "Цена в прайсе, руб": price})
    wb.close()
    return pd.DataFrame(rows).drop_duplicates("Ключ")


def ru_ceiling():
    """Минимальная цена по штрихкоду у российских оптовиков, в рублях."""
    df = pd.read_excel(PRICES, dtype={"Штрихкод": str})
    ru = df[df["Поставщик"].isin(RU_SUPPLIERS) & df["Цена, руб"].notna()].copy()
    ru["Штрихкод"] = ru["Штрихкод"].fillna("").str.replace(r"\D", "", regex=True).str[:13]
    ru = ru[ru["Штрихкод"].str.len() >= 8]
    wide = (ru.sort_values("Цена, руб")
              .drop_duplicates(["Штрихкод", "Поставщик"])
              .pivot(index="Штрихкод", columns="Поставщик", values="Цена, руб"))
    wide.columns = [f"{c}, руб" for c in wide.columns]
    wide["Потолок РФ, руб"] = wide.min(axis=1)
    return wide


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
DEAD_FILL = PatternFill("solid", fgColor="FCE4E4")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")


def write_excel(left, table):
    """Книга по остатку: весь остаток и отдельно то, что не берут вовсе."""
    dead = left[left["Статус"] == "не забирают вовсе"]
    with pd.ExcelWriter(OUT) as writer:
        left.to_excel(writer, sheet_name="ОСТАТОК", index=False)
        dead.to_excel(writer, sheet_name="НЕ ЗАБИРАЮТ", index=False)
        for name, frame in (("ОСТАТОК", left), ("НЕ ЗАБИРАЮТ", dead)):
            format_sheet(writer.sheets[name], frame)


def format_sheet(sheet, frame):
    titles = [str(c.value) for c in sheet[1]]
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    widths = {"Товар": 70, "Статус": 20, "Срок годности": 14, "Штрихкод": 15}
    for index, title in enumerate(titles, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(title, 14)
        if "руб" in title or "шт" in title:
            for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                row[0].number_format = "# ##0"

    # Позиции, которые не берут вовсе, подсвечиваем: их сбывать в первую очередь.
    status = titles.index("Статус")
    for row in sheet.iter_rows(min_row=2, max_row=len(frame) + 1):
        if row[status].value == "не забирают вовсе":
            for cell in row:
                cell.fill = DEAD_FILL

    total = len(frame) + 2
    sheet.cell(total, 1, "ИТОГО").font = Font(bold=True)
    for title in ("Остаётся, шт", "Остаток, руб", "Выручка по потолку, руб"):
        if title not in titles:
            continue
        column = titles.index(title) + 1
        letter = get_column_letter(column)
        cell = sheet.cell(total, column)
        cell.value = f"=SUM({letter}2:{letter}{len(frame) + 1})"
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.number_format = "# ##0"
    sheet.freeze_panes = "B2"


def main(stock_path, price_path):
    stock = read_stock(stock_path)
    stock["Ключ"] = stock["Товар"].map(key)
    table = stock.merge(barcodes_from_pricelist(price_path), on="Ключ", how="left")
    table = table.merge(ru_ceiling(), on="Штрихкод", how="left")

    table["Остаток, руб"] = (table["Остаётся, шт"] * table["Себестоимость, руб"]).round(0)
    table["Наценка до потолка, %"] = (
        (table["Потолок РФ, руб"] / table["Себестоимость, руб"] - 1) * 100).round(0)
    table["Выручка по потолку, руб"] = (
        table["Остаётся, шт"] * table["Потолок РФ, руб"] * 0.88).round(0)

    left = table[table["Остаётся, шт"] > 0].sort_values("Остаток, руб", ascending=False)
    dead = table[(table["Заберут, шт"] == 0) & (table["Остаётся, шт"] > 0)]

    columns = ["Товар", "Срок годности", "Всего, шт", "Заберут, шт", "Остаётся, шт",
               "Себестоимость, руб", "Остаток, руб", "Потолок РФ, руб",
               "Наценка до потолка, %", "Выручка по потолку, руб", "Штрихкод"]
    left = left.reindex(columns=columns)
    left.insert(1, "Статус", ["не забирают вовсе" if q == 0 else "забирают частично"
                              for q in table.loc[left.index, "Заберут, шт"]])
    write_excel(left, table)

    print(f"Позиций всего: {len(table)}   остается на складе: {len(left)}   "
          f"из них не забирают вовсе: {len(dead)}")
    print(f"\n{'Забирают':<26} {table['Заберут, шт'].sum():>9,.0f} шт   "
          f"{(table['Заберут, шт'] * table['Себестоимость, руб']).sum():>14,.0f} руб по себестоимости")
    print(f"{'Остается':<26} {left['Остаётся, шт'].sum():>9,.0f} шт   "
          f"{left['Остаток, руб'].sum():>14,.0f} руб по себестоимости")
    matched = left[left["Потолок РФ, руб"].notna()]
    print(f"\nИз остатка есть у российских оптовиков: {len(matched)} позиций "
          f"на {matched['Остаток, руб'].sum():,.0f} руб по себестоимости")
    print(f"Их можно продать примерно за {matched['Выручка по потолку, руб'].sum():,.0f} руб "
          f"(потолок минус 12%)")
    print(f"Медианная наценка до потолка: {matched['Наценка до потолка, %'].median():.0f}%")

    print("\nТоп-15 остатка по деньгам:")
    cols = ["Товар", "Остаётся, шт", "Себестоимость, руб", "Остаток, руб",
            "Потолок РФ, руб", "Наценка до потолка, %"]
    print(left[cols].head(15).to_string(index=False))
    print(f"\nСохранено: {OUT}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--stock", required=True)
    parser.add_argument("--price", required=True)
    ns = parser.parse_args()
    main(ns.stock, ns.price)
