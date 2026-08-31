"""Разбор прайсов поставщиков в единую таблицу.

Каждый поставщик лежит в своей папке внутри data/price_lists. Прайсы внутри
приходят от разных брендов с разной версткой: заголовок может быть на любой
из первых строк, часть шапок двухэтажные, названия колонок пишутся
по-разному. Скрипт сам находит строку заголовка и раскладывает данные в
общий набор полей.

Себестоимость считаем по правилу проекта:
    себестоимость_руб = закупка_KRW * курс * 1.4

Запуск:
    python3 tools/parse_price_lists.py                    # все поставщики
    python3 tools/parse_price_lists.py --supplier classic # один поставщик
    python3 tools/parse_price_lists.py --rate 0.0598      # свой курс воны
    python3 tools/parse_price_lists.py путь.xlsx ...      # конкретные файлы
"""

import argparse
import glob
import os
import re

import openpyxl
import pandas as pd
import xlrd

from price_unit import detect as detect_price_unit

PRICE_ROOT = "data/price_lists"
OUT_DIR = "outputs"

# Имя папки поставщика -> как называем его в отчетах. Файлы, лежащие прямо
# в price_lists, — это прайсы, полученные от брендов напрямую.
# Страна и валюта нужны, чтобы не сравнивать поставщиков из разных стран:
# у них разный маршрут и разные расходы после отгрузки.
SUPPLIERS = {
    "annecy": ("Аннеси", "KR", "KRW"),
    "classic": ("Классик", "KR", "KRW"),
    "ge_global": ("G&E Global", "KR", "KRW"),
    "glowbeauty": ("GlowBeauty", "KR", "KRW"),
    "papacosmetic": ("Papa Cosmetic", "KR", "KRW"),
    "price_lists": ("Прямой прайс бренда", "KR", "KRW"),
    "koreatrade": ("KoreaTrade", "RU", "RUB"),
    "korshop": ("Korshop", "KG", "USD"),
    "keauty": ("KEAUTY", "RU", "RUB"),
    # Казахстанская компания со складом в Екатеринбурге: товар получают
    # внутри РФ, поэтому сравниваем с российскими. Прайс в долларах.
    "safiya": ("SAFIYA", "RU", "USD"),
    "aibeauty": ("Aibeauty", "KG", "USD"),
}


def supplier_dirs():
    """Папки поставщиков: сам price_lists и любая подпапка с прайсами внутри."""
    found = []
    if glob.glob(os.path.join(PRICE_ROOT, "*.xls*")):
        found.append(PRICE_ROOT)
    for path in sorted(glob.glob(os.path.join(PRICE_ROOT, "*"))):
        if os.path.isdir(path) and glob.glob(os.path.join(path, "*.xls*")):
            found.append(path)
    return found


def supplier_info(path):
    """Имя, страна и валюта поставщика по имени его папки."""
    key = os.path.basename(path.rstrip("/"))
    return SUPPLIERS.get(key, (key.replace("_", " ").title(), "KR", "KRW"))

# Базис поставки. По умолчанию EXW — это подтверждено по всем поставщикам,
# но отдельные прайсы написаны на FOB, и тогда берем то, что в прайсе.
DEFAULT_BASIS = "EXW"
BASIS_WORDS = r"\b(FOB|EXW|CIF|CIP|DAP|DDP|FCA)\b"

# Курсы ЦБ, рублей за единицу валюты. Обновляем на дату расчета.
KRW_RUB = 0.058
USD_RUB = 87.0
# Логистика, пошлина и приемка сверх закупочной цены.
IMPORT_MULTIPLIER = 1.4

# Как узнаем колонку: поле -> список шаблонов по тексту шапки.
COLUMN_PATTERNS = {
    # 구분/Division в прайсах — порядковый номер строки, не бренд.
    "brand": [r"^brand$", r"^бренд$", r"^бренд\b"],
    # SAFIYA подписывает штрихкод как GTIN.
    "code": [r"sku\s*no", r"^code$", r"product\s*code", r"sap\s*code", r"^артикул$"],
    "barcode": [r"bar\s*code", r"barcode", r"바코드", r"штрихкод", r"^gtin$"],
    # Названия: сначала колонки с явной пометкой языка (STRONG_PATTERNS),
    # потом общие подписи. Где обе колонки подписаны одинаково, корейская
    # идет первой, поэтому name_kr проверяется раньше name_en.
    "name_kr": [r"^pro\w*t\s*name$", r"^product$", r"^name$"],
    "name_en": [
        r"^pro\w*t\s*name$", r"^product$", r"^name$",
        # GlowBeauty подписывает колонку названия как Product list.
        r"^product\s*list$", r"item\s*description", r"\bname\b",
        r"номенклатура", r"наименование",
    ],
    # LEBELAGE ведет отдельную колонку с русскими названиями — забираем.
    "name_ru": [r"наименование", r"название"],
    "type": [r"^type$", r"^category$", r"product\s*line"],
    "volume": [r"^vol", r"volume", r"^size$", r"capacity", r"용량", r"규격",
               r"вес\s*\(объем\)", r"^вес$", r"^объем$"],
    "msrp_krw": [
        r"msrp", r"^srp\b", r"^retail\b", r"retail\s*price", r"list\s*price",
        r"consumer\s*price", r"regular\s*price\s*\(krw",
        # Корейские подписи розничной цены: 소비자가 / 정상소가 / 기준가.
        r"소비자", r"정상소가", r"기준가",
    ],
    # Закупку пишут по-разному: supply / unit / просто price (-VAT).
    "supply_krw": [
        r"sup\w*ly\w*\s*price", r"wholesale\s*price", r"fob\s*price", r"공급가",
        r"distributor\s*price",
        r"unit\s*price", r"^price\s*\(\s*-?\s*vat",
        # Оптовики РФ и КГ: берем цену самого крупного опта, это их нижний предел.
        r"опт\s*от\s*300", r"от\s*300\s*т\.?\s*р", r"^цена$", r"^цена,\s*\$$",
    ],
    "qty_per_box": [
        r"q'?ty\s*/?\s*box", r"qty\s*per\s*outbox", r"1\s*box\s*qty", r"ea\s*/\s*box",
        r"^master$", r"кол-?во\s*в\s*упаковке", r"информация\s*об\s*упаковке",
    ],
    "moq": [r"^moq", r"moq\s*qty"],
    "shelf_life": [r"shelf\s*life", r"유통기한"],
    "status": [r"^status$", r"^remark$", r"^비고$"],
}

# Колонки, которые нельзя отдавать полю, даже если шапка подошла.
EXCLUDE_PATTERNS = {
    "msrp_krw": [r"\busd\b", r"\beur\b", r"\$"],
    "supply_krw": [r"\busd\b", r"\beur\b", r"\$"],
}

# Явные языковые подписи — их разбираем до общих шаблонов названия.
STRONG_PATTERNS = {
    "name_kr": [
        r"\bname\b.*\b(ko|kr|kor|korean)\b", r"^korean$",
        r"국문", r"한글명", r"제품명", r"품명",
    ],
    "name_en": [
        r"\bname\b.*\b(en|eng|english)\b", r"^eng$", r"^english$", r"영문",
        r"description.*\b(en|eng|english)\b",
    ],
}

FALLBACK_PATTERNS = {
    "supply_krw": [r"supply\s*rate"],
}

# По этим словам ищем саму строку заголовка.
HEADER_HINTS = [
    "barcode", "bar code", "product name", "supply price", "msrp", "vol",
    "brand", "moq", "retail price", "code",
    # Аннеси подписывает шапку целиком по-корейски.
    "바코드", "품명", "제품명", "공급가", "소비자", "용량", "규격", "브랜드", "입수",
    # Российские и киргизские оптовики — по-русски.
    "штрихкод", "номенклатура", "бренд", "цена", "остаток", "артикул", "упаковк",
]


def norm(value):
    """Текст ячейки в нижнем регистре, переносы строк схлопнуты в пробел."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("_", " ")).strip().lower()


def find_header_row(rows, limit=12):
    """Строка заголовка — та, где больше всего опорных слов."""
    best_idx, best_score = None, 0
    for idx, row in enumerate(rows[:limit]):
        cells = [norm(c) for c in row]
        score = sum(1 for c in cells if c and any(h in c for h in HEADER_HINTS))
        if score > best_score:
            best_idx, best_score = idx, score
    return best_idx if best_score >= 2 else None


def looks_like_header(cells):
    """Строка похожа на часть шапки, если в ней есть опорные слова."""
    extra = ["price", "weight", "size", "cbm", "국문", "영문", "korean", "english"]
    return sum(1 for c in cells if c and any(h in c for h in HEADER_HINTS + extra)) >= 2


def merge_header(rows, header_idx):
    """Склеиваем шапку из нескольких строк.

    Подписи бывают и снизу (двухэтажная шапка с уточнениями), и сверху:
    korshop пишет «Цена» строкой выше остальных названий колонок.
    """
    header = [norm(c) for c in rows[header_idx]]

    if header_idx + 1 < len(rows):
        below = [norm(c) for c in rows[header_idx + 1]]
        if looks_like_header(below):
            header = [(t + " " + b).strip() for t, b in zip(header, below)]

    if header_idx > 0:
        above = [norm(c) for c in rows[header_idx - 1]]
        if looks_like_header(above):
            # Сверху берем только то, для чего в самой шапке подписи нет.
            header = [t or a for t, a in zip(header, above + [""] * len(header))]
    return header


def map_columns(header):
    """Шапка -> {поле: номер колонки}.

    Два прохода. Сначала колонки с явной пометкой языка: JIGOTT называет
    английское название "Description (ENG)", CP1 подписывает языки во втором
    этаже шапки. Потом общие подписи — там, где AMORE и FLOR DE MAN зовут обе
    колонки одинаково ("Product Name" / "Product"), первая колонка корейская.
    """
    mapping = map_once(header, exclude=True)
    # Отбрасывать цены в долларах имеет смысл, только когда рядом есть цена в
    # местной валюте: у DEAR KLAIRS так, а у SAFIYA весь прайс в долларах.
    if "supply_krw" not in mapping:
        mapping = map_once(header, exclude=False)
    return mapping


def map_once(header, exclude):
    mapping, used = {}, set()
    for patterns_set in (STRONG_PATTERNS, COLUMN_PATTERNS, FALLBACK_PATTERNS):
        for col_idx, title in enumerate(header):
            if not title or col_idx in used:
                continue
            for field, patterns in patterns_set.items():
                if field in mapping:
                    continue
                if not any(re.search(p, title) for p in patterns):
                    continue
                if exclude and any(re.search(p, title) for p in EXCLUDE_PATTERNS.get(field, [])):
                    continue
                mapping[field] = col_idx
                used.add(col_idx)
                break
    return mapping


def to_number(value):
    """Число из ячейки: убираем валюту, пробелы и разделители разрядов.

    Запятая бывает и разделителем тысяч (40,000), и десятичной точкой:
    korshop пишет цены как 7,73. Различаем по числу цифр после запятой.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^\d.,\-]", "", str(value))
    if not re.search(r"\d", text):
        return None
    decimal = re.fullmatch(r"-?\d+,\d{1,2}", text)
    text = text.replace(",", "." if decimal else "")
    try:
        return float(text)
    except ValueError:
        return None


KOREAN = r"\uac00-\ud7a3"

# Один бренд поставщики пишут по-разному, и сравнение по нему дробится.
BRAND_ALIASES = {
    "ROUNDLAB": "ROUND LAB",
    "MAANYO": "MANYO",
    "DR ALTHEA": "DR. ALTHEA",
    "VT COSMETICS": "VT",
    "AMORE PACIFIC": "AMORE",
    "3W CLINIC LAB": "3W CLINIC",
    "3W CLINIC PREMIUM": "3W CLINIC",
}


def normalize_brand(brand):
    """Единое написание бренда: без корейской расшифровки в скобках."""
    if not brand:
        return brand
    name = re.sub(r"\s*\(.*?\)\s*$", "", str(brand)).strip()
    return BRAND_ALIASES.get(name.upper(), name)


def split_languages(text):
    """Разносит название на английское и корейское.

    Часть поставщиков (FLORODORA, BERGAMO, AMORE, CP1) кладет оба названия
    в одну ячейку — через перенос строки или просто подряд.
    """
    if not text or not re.search(f"[{KOREAN}]", text):
        return text, ""
    korean = " ".join(re.findall(f"[{KOREAN}][{KOREAN}\\s]*", text)).strip()
    latin = " ".join(
        part.strip() for part in re.split(f"[{KOREAN}]+", text)
        if re.search(r"[A-Za-z]{2}", part)
    ).strip()
    return re.sub(r"\s+", " ", latin), re.sub(r"\s+", " ", korean)


def clean_barcode(value):
    """Штрихкод к 13 цифрам.

    Российские оптовики дописывают к штрихкоду свой суффикс через дефис:
    8809647394136-11. Без обрезки получается пятнадцатизначное число, и
    товар перестает находиться у других поставщиков.
    """
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) > 13:
        digits = digits[:13]
    return digits if len(digits) >= 8 else ""


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def brand_from_filename(path):
    """Бренд из имени файла.

    Убираем префикс папки поставщика, слова PRICE LIST (встречаются опечатки
    вида PRICE_LEST) и дату в конце — у G&E Global она идет как _2608.
    """
    folder = os.path.basename(os.path.dirname(path))
    name = os.path.splitext(os.path.basename(path))[0]
    # Дата в начале имени: 260810_ у Аннеси, 26.07.01_ у Papa Cosmetic.
    name = re.sub(r"^\d{6}_|^\d{2}\.\d{2}\.\d{2}_", "", name, flags=re.I)
    # Имя поставщика встречается и в префиксе папки, и внутри самого имени.
    name = re.sub(rf"{re.escape(folder)}_?", "", name, flags=re.I)
    name = re.sub(r"^\d{6}_|^\d{2}\.\d{2}\.\d{2}_", "", name)
    name = re.split(r"_?PRICE[_ ]?L[EI]S?T|_LIST\d*\b|_\d{2}\.\d{2}", name, flags=re.I)[0]
    name = re.sub(r"_\d{4}$", "", name)
    return name.replace("_", " ").strip(" ._").upper()


def parse_sheet(ws, source, sheet_name, fallback_brand, supplier, country, currency):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], "заголовок не найден"

    # Базис ищем по шапке прайса: он пишется в заголовке или в подписи цены.
    top_text = " ".join(str(c) for row in rows[:header_idx + 2] for c in row if c is not None)
    found_basis = re.search(BASIS_WORDS, top_text, re.I)
    basis = found_basis.group(1).upper() if found_basis else DEFAULT_BASIS

    header = merge_header(rows, header_idx)
    mapping = map_columns(header)
    if "supply_krw" not in mapping:
        return [], "нет колонки закупочной цены"

    records = []
    last_brand = fallback_brand
    for row in rows[header_idx + 1:]:
        supply = to_number(row[mapping["supply_krw"]]) if mapping["supply_krw"] < len(row) else None
        if not supply:
            # Строка-разделитель с одним словом — это название бренда:
            # korshop и SKIN APPLE так разбивают прайс на секции.
            filled = [clean_text(c) for c in row if clean_text(c)]
            if len(filled) == 1 and 2 <= len(filled[0]) <= 40 and not filled[0].isdigit():
                last_brand = filled[0]
            continue

        def cell(field):
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        name_en = clean_text(cell("name_en"))
        name_kr = clean_text(cell("name_kr"))
        if not name_en and not name_kr:
            continue
        # Колонку без языка в шапке могли занять под латиницу — это название EN.
        if not name_en and not re.search(f"[{KOREAN}]", name_kr):
            name_en, name_kr = name_kr, ""
        # Разносим языки, только если английского названия еще нет или в нем
        # оказался корейский: иначе затрем уже правильное название объемом.
        needs_split = not name_en or bool(re.search(f"[{KOREAN}]", name_en))
        for candidate in (name_en, name_kr) if needs_split else ():
            latin, korean = split_languages(candidate)
            if not (latin and korean):
                continue
            name_en = latin
            if not re.search(f"[{KOREAN}]", name_kr) or name_kr == candidate:
                name_kr = korean
            break

        brand = clean_text(cell("brand")) or last_brand
        last_brand = brand  # в прайсах бренд ставят только в первой строке блока
        brand = normalize_brand(brand)

        unit, per_pack = detect_price_unit(clean_text(cell("volume")), name_en or name_kr)
        records.append({
            "Поставщик": supplier,
            "Страна": country,
            "Валюта": currency,
            "Файл": os.path.basename(source),
            "Лист": sheet_name,
            "Бренд": brand,
            "Артикул": clean_text(cell("code")),
            "Штрихкод": clean_barcode(cell("barcode")),
            "Название EN": name_en,
            "Название KR": name_kr,
            "Название RU": clean_text(cell("name_ru")),
            "Тип": clean_text(cell("type")),
            "Объем": clean_text(cell("volume")),
            "Базис": basis,
            "Единица цены": unit,
            "Штук в упаковке": per_pack,
            "MSRP, KRW": to_number(cell("msrp_krw")),
            "Закупка, KRW": supply,
            "Шт/короб": to_number(cell("qty_per_box")),
            "MOQ": to_number(cell("moq")),
            "Срок годности": clean_text(cell("shelf_life")),
            "Примечание": clean_text(cell("status")),
        })
    return records, None


def dedupe(records, notes):
    """Часть поставщиков дублирует каталог на листах 국문 и English.

    Совпадение ищем по артикулу, а если его нет — по паре штрихкод + название.
    Оставляем первую строку, но если у дубля другая закупочная цена, это не
    техническая копия, а второй уровень цен — говорим об этом отдельно.
    """
    unique, seen, conflicts = [], {}, []
    for rec in records:
        key = rec["Артикул"] or (rec["Штрихкод"], rec["Название EN"])
        if not (rec["Артикул"] or rec["Штрихкод"]):
            unique.append(rec)
            continue
        if key in seen:
            kept = seen[key]
            if kept["Закупка, KRW"] != rec["Закупка, KRW"]:
                conflicts.append((kept, rec))
            continue
        seen[key] = rec
        unique.append(rec)

    dropped = len(records) - len(unique)
    if dropped:
        notes.append(f"убрано дублей между листами: {dropped}")
    if conflicts:
        kept, other = conflicts[0]
        notes.append(
            f"РАЗНЫЕ ЦЕНЫ на листах ({len(conflicts)} SKU), взят лист "
            f"'{kept['Лист']}': напр. {kept['Артикул'] or kept['Штрихкод']} — "
            f"{kept['Закупка, KRW']:.0f} против {other['Закупка, KRW']:.0f} "
            f"{kept.get('Валюта', '')}"
        )
    return unique, notes


def pick_sheets(sheet_names):
    """Отбираем листы, по которым считать закупку.

    LAGOM дает один каталог дважды: VAT포함 (с НДС) и VAT별도 (без НДС), плюс
    листы 검산/검수 — внутренняя сверка. Нам нужна цена без НДС.
    """
    sheets = [s for s in sheet_names if not re.search(r"검산|검수", s)]
    if any("별도" in s for s in sheets):
        sheets = [s for s in sheets if "포함" not in s]
    return sheets


class LegacySheet:
    """Лист старого .xls в том же виде, что дает openpyxl."""

    def __init__(self, sheet):
        self._sheet = sheet
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        last = min(max_row or self.max_row, self.max_row)
        for index in range(min_row - 1, last):
            yield tuple(cell.value if cell.value != "" else None
                        for cell in self._sheet.row(index))


class LegacyWorkbook:
    """Книга старого .xls: G&E Global присылает часть прайсов в этом формате."""

    def __init__(self, path):
        self._book = xlrd.open_workbook(path)
        self.sheetnames = self._book.sheet_names()

    def __getitem__(self, name):
        return LegacySheet(self._book.sheet_by_name(name))

    def close(self):
        self._book.release_resources()


def open_workbook(path):
    if path.lower().endswith(".xls"):
        return LegacyWorkbook(path)
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def parse_file(path, supplier, country, currency):
    wb = open_workbook(path)
    fallback_brand = brand_from_filename(path)
    records, notes = [], []
    for sheet_name in pick_sheets(wb.sheetnames):
        found, problem = parse_sheet(wb[sheet_name], path, sheet_name, fallback_brand,
                                     supplier, country, currency)
        records.extend(found)
        if problem:
            notes.append(f"{sheet_name}: {problem}")
    wb.close()

    unique, notes = dedupe(records, notes)
    return unique, notes


def main(paths, rate, usd, only_supplier):
    """Разбираем прайсы и складываем в одну таблицу по всем поставщикам."""
    if paths:
        jobs = [(p,) + supplier_info(os.path.dirname(p)) for p in paths]
    else:
        dirs = supplier_dirs()
        if only_supplier:
            dirs = [d for d in dirs if os.path.basename(d) == only_supplier]
            if not dirs:
                print(f"Поставщик {only_supplier} не найден в {PRICE_ROOT}")
                return
        jobs = [
            (path,) + supplier_info(d)
            for d in dirs
            for path in sorted(glob.glob(os.path.join(d, "*.xls*")))
        ]

    all_records = []
    current = None
    for path, supplier, country, currency in jobs:
        if supplier != current:
            current = supplier
            print(f"\n=== {supplier} ({country}, {currency}) ===")
            print(f"{'Файл':<52} {'SKU':>5}  Замечания")
            print("-" * 92)
        records, notes = parse_file(path, supplier, country, currency)
        all_records.extend(records)
        print(f"{os.path.basename(path):<52} {len(records):>5}  {'; '.join(notes)}")

    if not all_records:
        print("\nНичего не разобрано.")
        return

    df = pd.DataFrame(all_records)
    # Наценка поставщика к рекомендованной рознице — грубый ориентир по марже.
    df["MSRP/Закупка"] = (df["MSRP, KRW"] / df["Закупка, KRW"]).round(2)
    # Цену набора делим на число штук: сравнивать можно только штуку со штукой.
    per_pack = df["Штук в упаковке"].where(df["Единица цены"] == "за набор").fillna(1)
    df["Цена за штуку, KRW"] = (df["Закупка, KRW"] / per_pack).round(2)

    # Приводим все валюты к рублям, чтобы цены разных стран были сопоставимы.
    to_rub = df["Валюта"].map({"KRW": rate, "USD": usd, "RUB": 1.0}).fillna(1.0)
    df["Цена, руб"] = (df["Закупка, KRW"] * to_rub).round(2)
    df["Цена за штуку, руб"] = (df["Цена за штуку, KRW"] * to_rub).round(2)

    # Множитель импорта применяем только к корейской закупке: у российских
    # оптовиков товар уже ввезен, у них это цена продажи, а не наша закупка.
    imported = (df["Страна"] == "KR").map({True: IMPORT_MULTIPLIER, False: 1.0})
    df["Себестоимость, руб"] = (df["Цена, руб"] * imported).round(2)
    df["Себестоимость штуки, руб"] = (df["Цена за штуку, руб"] * imported).round(2)
    df["Курс KRW"] = rate

    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, "prices_normalized.xlsx")
    df.to_excel(out_path, index=False)
    print(f"\nВсего SKU: {len(df)}   Брендов: {df['Бренд'].nunique()}"
          f"   Поставщиков: {df['Поставщик'].nunique()}")
    print(f"Курсы: {rate} руб/вона, {usd} руб/доллар. "
          f"Множитель импорта {IMPORT_MULTIPLIER} — только для корейской закупки")
    print(f"Сохранено: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("paths", nargs="*", help="конкретные файлы прайсов")
    parser.add_argument("--rate", type=float, default=KRW_RUB, help="курс рублей за 1 вону")
    parser.add_argument("--usd", type=float, default=USD_RUB, help="курс рублей за 1 доллар")
    parser.add_argument("--supplier", help="имя папки поставщика в data/price_lists")
    ns = parser.parse_args()
    main(ns.paths, ns.rate, ns.usd, ns.supplier)
