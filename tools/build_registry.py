"""Сборка РЕЕСТРА ЦЕН по структуре инструкции по закупкам.

Одна строка = один товар одного поставщика на одну дату. Колонки A-Q идут
в том порядке, что задан в инструкции, чтобы таблицу можно было вставить в
Google Sheets без перестановки.

Чего в прайсах нет и что приходится доставать отдельно:

- Дата прайса берется из имени файла. Внутри прайсов она стоит не везде,
  а число почти нигде, поэтому точность до месяца — день помечаем как 01
  и пишем об этом в примечании.
- Базис поставки и минимальную партию ищем в шапке прайса по словам FOB,
  EXW, CIF, DAP и MOA/MOQ. Где не нашли — ставим «уточнить».

Запуск:
    python3 tools/build_registry.py
"""

import glob
import os
import re

import openpyxl
import pandas as pd

SRC = "outputs/prices_normalized.xlsx"
OUT_XLSX = "outputs/registry_prices.xlsx"
OUT_CSV = "outputs/registry_prices.csv"
PRICE_ROOT = "data/price_lists"

COUNTRY = "KR"
# Поставщик из сводной таблицы -> папка с его прайсами.
FOLDERS = {"Классик": "classic", "G&E Global": "ge_global", "GlowBeauty": "glowbeauty"}

BASIS_WORDS = r"\b(FOB|EXW|CIF|CIP|DAP|DDP|FCA)\b"
COLUMNS = [
    "Штрихкод", "Бренд", "Наименование", "Объём", "Код поставщика", "Страна",
    "Цена", "Валюта", "Единица цены", "Кол-во в коробе", "Базис поставки",
    "Условия оплаты", "Мин. партия", "Срок годности", "Дата прайса",
    "Файл-источник", "Примечание",
]


def price_date(filename):
    """Дата прайса из имени файла. Возвращает пару (дата, точность)."""
    name = os.path.splitext(filename)[0]
    full = re.search(r"(20\d{2})\.(\d{2})", name)            # 2026.07
    if full and 1 <= int(full.group(2)) <= 12:
        return f"{full.group(1)}-{full.group(2)}-01", "месяц"
    short = re.search(r"(?<!\d)(\d{2})\.(\d{2})", name)       # 26.04
    if short and 1 <= int(short.group(2)) <= 12:
        return f"20{short.group(1)}-{short.group(2)}-01", "месяц"
    packed = re.search(r"_(\d{2})(\d{2})(?:\D|$)", name)      # 2608
    if packed and 24 <= int(packed.group(1)) <= 30 and 1 <= int(packed.group(2)) <= 12:
        return f"20{packed.group(1)}-{packed.group(2)}-01", "месяц"
    year = re.search(r"_(20\d{2})(?:\D|$)", name)             # 2026
    if year:
        return f"{year.group(1)}-01-01", "только год"
    return "", "не найдена"


def file_header_text(path, rows=12):
    """Первые строки прайса одной строкой — там пишут базис и минималку."""
    try:
        if path.lower().endswith(".xls"):
            import xlrd
            book = xlrd.open_workbook(path)
            sheet = book.sheet_by_index(0)
            cells = [str(sheet.cell_value(r, c))
                     for r in range(min(rows, sheet.nrows)) for c in range(sheet.ncols)]
            book.release_resources()
        else:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            cells = [str(c) for row in ws.iter_rows(max_row=rows, values_only=True)
                     for c in row if c is not None]
            wb.close()
        return " ".join(cells)
    except Exception:
        return ""


def file_terms(path):
    """Базис поставки и минимальная партия из шапки прайса."""
    text = file_header_text(path)
    basis = re.search(BASIS_WORDS, text, re.I)
    # MOA обычно сумма заказа (в 만원, это 10 000 вон), MOQ бывает и в штуках.
    moa = re.search(r"(?:MOA|MOQ|최소)\s*[:：]?\s*([\d,]+)\s*\(?\s*(만원|만|원|ea|개|pcs|USD|\$)?",
                    text, re.I)
    minimum = ""
    if moa:
        amount = int(moa.group(1).replace(",", ""))
        unit = (moa.group(2) or "").lower()
        if unit in ("만원", "만"):
            minimum = f"{amount * 10000} KRW"
        elif unit in ("ea", "개", "pcs"):
            minimum = f"{amount} шт"
        else:
            minimum = f"{amount} {moa.group(2)}".strip() if moa.group(2) else f"{amount} (единица не указана)"
    return (basis.group(1).upper() if basis else "уточнить"), (minimum or "уточнить")


def supplier_code(supplier, brand, filename):
    """Код поставщика в формате СТРАНА_названиелатиницей."""
    if supplier in FOLDERS:
        return f"{COUNTRY}_{FOLDERS[supplier]}"
    # Прямые прайсы: у каждого бренда свой производитель, код по бренду.
    slug = re.split(r"_PRICE|_\d{2}\.", filename)[0]
    slug = re.sub(r"[^A-Za-z0-9]+", "", slug) or re.sub(r"[^A-Za-z0-9]+", "", str(brand))
    return f"{COUNTRY}_{slug.lower()}"


def main():
    if not os.path.exists(SRC):
        raise SystemExit(f"Нет файла {SRC} — сначала запустите parse_price_lists.py")
    df = pd.read_excel(SRC)

    paths = {os.path.basename(p): p
             for p in glob.glob(os.path.join(PRICE_ROOT, "*.xls*"))
             + glob.glob(os.path.join(PRICE_ROOT, "*", "*.xls*"))}
    terms = {name: file_terms(path) for name, path in paths.items()}
    dates = {name: price_date(name) for name in paths}

    out = pd.DataFrame({
        "Штрихкод": df["Штрихкод"].astype(str).str.replace(r"\D", "", regex=True),
        "Бренд": df["Бренд"].fillna("").astype(str).str.upper(),
        "Наименование": df["Название EN"].fillna(df["Название KR"]),
        "Объём": df["Объем"],
        "Код поставщика": [supplier_code(s, b, f)
                           for s, b, f in zip(df["Поставщик"], df["Бренд"], df["Файл"])],
        "Страна": COUNTRY,
        "Цена": df["Закупка, KRW"],
        "Валюта": "KRW",
        "Единица цены": df["Единица цены"],
        "Кол-во в коробе": df["Шт/короб"],
        "Базис поставки": [terms.get(f, ("уточнить", ""))[0] for f in df["Файл"]],
        "Условия оплаты": "уточнить",
        "Мин. партия": [terms.get(f, ("", "уточнить"))[1] for f in df["Файл"]],
        "Срок годности": df["Срок годности"],
        "Дата прайса": [dates.get(f, ("", ""))[0] for f in df["Файл"]],
        "Файл-источник": df["Файл"],
    })

    notes = []
    for accuracy, unit, per_pack in zip(
            (dates.get(f, ("", ""))[1] for f in df["Файл"]),
            df["Единица цены"], df["Штук в упаковке"]):
        parts = []
        if accuracy != "месяц":
            parts.append(f"дата прайса: {accuracy}")
        else:
            parts.append("день в дате не указан")
        if unit == "за набор":
            parts.append(f"набор {int(per_pack)} шт" if pd.notna(per_pack) else "набор, число штук не указано")
        notes.append("; ".join(parts))
    out["Примечание"] = notes
    out = out[COLUMNS]

    os.makedirs("outputs", exist_ok=True)
    out.to_excel(OUT_XLSX, index=False)
    out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

    print(f"Строк в реестре: {len(out)}   Поставщиков: {out['Код поставщика'].nunique()}")
    print("\nКоды поставщиков (топ-12 по числу позиций):")
    print(out["Код поставщика"].value_counts().head(12).to_string())
    print("\nДата прайса определена:", int((out["Дата прайса"] != "").sum()), "из", len(out))
    print("Диапазон дат:", out.loc[out["Дата прайса"] != "", "Дата прайса"].min(),
          "—", out.loc[out["Дата прайса"] != "", "Дата прайса"].max())
    print("\nБазис поставки:")
    print(out["Базис поставки"].value_counts().to_string())
    print("\nМин. партия:")
    print(out["Мин. партия"].value_counts().head(6).to_string())
    print(f"\nСохранено: {OUT_XLSX}\n           {OUT_CSV}")


if __name__ == "__main__":
    main()
