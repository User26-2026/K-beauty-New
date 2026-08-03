"""Живой Excel-шаблон юнит-экономики Ozon (формулы, а не значения).

Меняешь жёлтые ячейки (цена, параметры, при желании себестоимость/комиссию) —
книга пересчитывается сама средствами Excel. Данные (комиссии, логистика,
объём) берутся из выгрузки prices.json, себестоимость — из COSTS.

Вход:  data/ozon_api/prices.json
Выход: outputs/ozon/ozon_unit_model_live_2026-08-03.xlsx
"""
import json
from pathlib import Path

from openpyxl.utils import get_column_letter

from calc_unit_ozon import (COSTS, DRR, TAX, ACQUIRING, STORAGE_L_DAY,
                            STORAGE_DAYS, FULFILLMENT, buyout, num)

ROOT = Path(__file__).parent.parent.parent
PRICES = ROOT / 'data' / 'ozon_api' / 'prices.json'
INFO = ROOT / 'data' / 'ozon_api' / 'product_info.json'
OUT = ROOT / 'outputs' / 'ozon'


def load_liters():
    """Объём в литрах по product_id (из габаритов, поле liters в product_info)."""
    if not INFO.exists():
        return {}
    data = json.loads(INFO.read_text())
    out = {}
    for pid, meta in data.items():
        if isinstance(meta, dict) and meta.get('liters'):
            out[str(pid)] = float(meta['liters'])
    return out


def scheme_inputs(item, scheme):
    c = item.get('commissions', {}) or {}
    if scheme == 'fbo':
        return (num(c.get('sales_percent_fbo')) / 100,
                num(c.get('fbo_direct_flow_trans_min_amount')),
                num(c.get('fbo_deliv_to_customer_amount')),
                num(c.get('fbo_return_flow_amount')))
    return (num(c.get('sales_percent_fbs')) / 100,
            num(c.get('fbs_direct_flow_trans_min_amount')),
            num(c.get('fbs_deliv_to_customer_amount')),
            num(c.get('fbs_return_flow_amount')))


HEAD = ['Товар', 'Цена', 'СПП, %', 'Цена с СПП', 'Себест.', 'Фулфилмент',
        'База затрат', 'Комиссия %', 'Комиссия ₽', 'Логистика',
        'Наценка нелок., ₽', 'Доставка до ПВЗ',
        'Эквайринг', 'Налог', 'Выкуп %', 'Обр.лог база',
        'Обр.лог', 'Объём, л', 'Хранение',
        'Оплата за клик, %', 'Оплата за клик ₽', 'Вывод в топ, %', 'Вывод в топ ₽',
        'Реклама всего ₽', 'ДРР %', 'Соинвест, %', 'Соинвест ₽', 'Маржа до рекл.',
        'Маржа/шт', 'Маржа % цены', 'ROI %', 'Цена безубыт.']
# буквы колонок по порядку HEAD (поддержка >26 колонок → AA, AB, ...)
COL = {name: get_column_letter(i + 1) for i, name in enumerate(HEAD)}


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = json.loads(PRICES.read_text())
    liters = load_liters()  # объём в литрах из габаритов (если выгружены)
    rows = []
    for it in items:
        if it.get('offer_id') not in COSTS:
            continue
        # только FBO
        comm, logi, lastm, retb = scheme_inputs(it, 'fbo')
        pid = str(it.get('product_id'))
        # объём: реальные литры из габаритов, иначе fallback на volume_weight
        vol = liters.get(pid) or num(it.get('volume_weight'))
        rows.append({
            'offer': it.get('offer_id'),
            'price': num(it.get('price', {}).get('price')),
            'cost': COSTS[it.get('offer_id')],
            'comm': comm, 'logi': logi, 'lastm': lastm, 'retb': retb,
            'buyout': buyout(it.get('offer_id')),
            'vol': vol,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = 'Юнитка Ozon (live)'

    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='305496')
    in_fill = PatternFill('solid', fgColor='FFF2CC')     # редактируемые (жёлтые)
    res_fill = PatternFill('solid', fgColor='E2EFDA')    # результат
    bold = Font(bold=True)

    # --- Блок параметров (редактируемый) ---
    ws['A1'] = 'ПАРАМЕТРЫ — меняй жёлтые ячейки, всё пересчитается'
    ws['A1'].font = Font(bold=True, size=12)
    params = [
        ('Налог с оборота', TAX, '0%'),
        ('Эквайринг', ACQUIRING, '0.0%'),
        ('Фулфилмент, ₽/шт', FULFILLMENT, '#,##0'),
        ('Хранение, ₽/л/день', STORAGE_L_DAY, '0.00'),
        ('Дней хранения', STORAGE_DAYS, '#,##0'),
        ('Наценка за нелокальные продажи, %', 0.0, '0.0%'),
    ]
    # значения в B3..B8
    for i, (label, val, fmt) in enumerate(params):
        r = 3 + i
        ws[f'A{r}'] = label
        cell = ws[f'B{r}']
        cell.value = val
        cell.number_format = fmt
        cell.fill = in_fill
        cell.border = border
        cell.font = bold
    # ДРР задаётся по строкам; наценка за нелокальность — параметр B8 (% к логистике).
    P = {'tax': '$B$3', 'acq': '$B$4', 'ff': '$B$5', 'stl': '$B$6',
         'std': '$B$7', 'nonlocal': '$B$8'}

    # --- Заголовок таблицы ---
    HROW = 11
    for j, name in enumerate(HEAD):
        c = ws.cell(HROW, j + 1, name)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = head_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[HROW].height = 30

    money = '#,##0'
    money2 = '#,##0.00'
    pct = '0.0%'

    input_cols = {'Цена', 'СПП, %', 'Себест.', 'Комиссия %', 'Логистика',
                  'Доставка до ПВЗ',
                  'Выкуп %', 'Обр.лог база', 'Объём, л',
                  'Оплата за клик, %', 'Вывод в топ, %', 'Соинвест, %'}

    for k, d in enumerate(rows):
        rr = HROW + 1 + k
        # ссылки на ячейки текущей строки
        def cr(name):
            return f'{COL[name]}{rr}'
        # входные значения
        vals = {
            'Товар': d['offer'], 'Цена': d['price'], 'СПП, %': 0,
            'Себест.': round(d['cost'], 2), 'Комиссия %': d['comm'],
            'Логистика': d['logi'],
            'Доставка до ПВЗ': d['lastm'],
            'Выкуп %': 0.90, 'Обр.лог база': d['retb'], 'Объём, л': d['vol'],
            'Оплата за клик, %': DRR,   # CPC-продвижение товаров, ДРР (доля от цены)
            'Вывод в топ, %': 0,        # CPO — ставка % от цены проданного заказа
            'Соинвест, %': 0,           # % из кабинета (Цены → Баллы за скидки)
        }
        # формулы
        formulas = {
            'Цена с СПП': f'={cr("Цена")}*(1-{cr("СПП, %")})',
            'Фулфилмент': f'={P["ff"]}',
            'База затрат': f'={cr("Себест.")}+{cr("Фулфилмент")}',
            'Комиссия ₽': f'={cr("Цена")}*{cr("Комиссия %")}',
            'Эквайринг': f'={cr("Цена")}*{P["acq"]}',
            'Налог': f'={cr("Цена с СПП")}*{P["tax"]}',  # налог от цены с СПП
            'Обр.лог': f'={cr("Обр.лог база")}*(1-{cr("Выкуп %")})',
            # Наценка за нелокальность = % (B8) от логистики.
            'Наценка нелок., ₽': f'={cr("Логистика")}*{P["nonlocal"]}',
            'Хранение': f'={P["stl"]}*{P["std"]}*{cr("Объём, л")}',
            # Актуальная реклама Ozon: Оплата за клик (CPC, продвижение товаров —
            # моделируем как ДРР, доля от цены) + Вывод в топ (CPO — ставка % от
            # цены проданного заказа). Сумма — общий рекламный расход.
            'Оплата за клик ₽': f'={cr("Цена")}*{cr("Оплата за клик, %")}',
            'Вывод в топ ₽': f'={cr("Цена")}*{cr("Вывод в топ, %")}',
            'Реклама всего ₽': f'={cr("Оплата за клик ₽")}+{cr("Вывод в топ ₽")}',
            # Доля рекламных расходов = весь рекламный расход / цена.
            'ДРР %': f'={cr("Реклама всего ₽")}/{cr("Цена")}',
            # Соинвест: Ozon возвращает часть баллами, они гасят комиссию/логистику,
            # поэтому в марже прибавляется (компенсация издержек).
            'Соинвест ₽': f'={cr("Цена")}*{cr("Соинвест, %")}',
            'Маржа до рекл.': (f'={cr("Цена")}-{cr("Комиссия ₽")}-{cr("Логистика")}'
                              f'-{cr("Наценка нелок., ₽")}-{cr("Доставка до ПВЗ")}'
                              f'-{cr("Эквайринг")}-{cr("Налог")}'
                              f'-{cr("Обр.лог")}-{cr("Хранение")}-{cr("База затрат")}'
                              f'+{cr("Соинвест ₽")}'),
            'Маржа/шт': f'={cr("Маржа до рекл.")}-{cr("Реклама всего ₽")}',
            'Маржа % цены': f'={cr("Маржа/шт")}/{cr("Цена")}',
            'ROI %': f'={cr("Маржа/шт")}/{cr("База затрат")}',
            # налог зависит от цены с СПП → в знаменателе (1-СПП)*налог
            'Цена безубыт.': (f'=({cr("Логистика")}+{cr("Наценка нелок., ₽")}'
                             f'+{cr("Доставка до ПВЗ")}+{cr("Обр.лог")}'
                             f'+{cr("Хранение")}+{cr("База затрат")})'
                             f'/(1-({cr("Комиссия %")}+{P["acq"]}+(1-{cr("СПП, %")})*{P["tax"]}'
                             f'+{cr("Оплата за клик, %")}+{cr("Вывод в топ, %")}'
                             f'-{cr("Соинвест, %")}))'),
        }

        for name in HEAD:
            cell = ws.cell(rr, HEAD.index(name) + 1)
            cell.value = vals.get(name, formulas.get(name))
            cell.border = border
            # форматы
            if name in ('СПП, %', 'Комиссия %', 'Выкуп %', 'Оплата за клик, %',
                        'Вывод в топ, %', 'ДРР %', 'Соинвест, %', 'Маржа % цены', 'ROI %'):
                cell.number_format = pct
            elif name in ('Себест.', 'Фулфилмент', 'База затрат'):
                cell.number_format = money2
            elif name == 'Объём, л':
                cell.number_format = '0.00'
            elif name != 'Товар':
                cell.number_format = money
            # заливка
            if name in input_cols:
                cell.fill = in_fill
            if name in ('Маржа/шт', 'ROI %'):
                cell.fill = res_fill
                cell.font = bold

    # ширины
    widths = {'Товар': 42, 'СПП, %': 8, 'Цена с СПП': 11, 'База затрат': 12,
              'Наценка нелок., ₽': 15, 'Обр.лог база': 12,
              'Оплата за клик, %': 13, 'Оплата за клик ₽': 13, 'Вывод в топ, %': 12,
              'Вывод в топ ₽': 12, 'Реклама всего ₽': 13, 'ДРР %': 8,
              'Доставка до ПВЗ': 13,
              'Соинвест, %': 11, 'Соинвест ₽': 11, 'Маржа до рекл.': 13,
              'Маржа % цены': 11, 'Цена безубыт.': 12}
    for name in HEAD:
        ws.column_dimensions[COL[name]].width = widths.get(name, 10)
    ws.freeze_panes = f'B{HROW + 1}'

    # подсказка
    note = ws.cell(HROW + len(rows) + 3, 1,
                   'Только FBO. Жёлтые ячейки — вход (цена, СПП %, себест., '
                   'комиссия, логистика, наценка за нелокальность, выкуп, объём, '
                   'реклама %, соинвест %) и параметры B3:B7. Остальное — формулы. '
                   'Цена с СПП = Цена×(1−СПП%); налог считается от цены с СПП. '
                   'Объём — в литрах из габаритов карточки (если выгружены, иначе '
                   'объёмный вес). Реклама по актуальным форматам Ozon: Оплата за '
                   'клик % (CPC) и Вывод в топ % (CPO). Соинвест % — из кабинета '
                   '(Цены → Баллы за скидки), баллы прибавляются к марже. '
                   'Меняешь цену в колонке B → маржа и ROI пересчитываются.')
    note.font = Font(italic=True, color='808080')

    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / 'ozon_unit_model_live_2026-08-03.xlsx'
    wb.save(path)
    print(f'Сохранено: {path}  (строк: {len(rows)})')


if __name__ == '__main__':
    build()
