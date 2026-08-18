"""Юнит-экономика для Ozon по реальным тарифам из Seller API.

Комиссии и плечи логистики берём из выгрузки data/ozon_api/prices.json
(метод /v5/product/info/prices), себестоимость — из желтой колонки
«Себестоимость по последнему приходу» файла остатков (задаётся тут словарём
COSTS по offer_id). Считаем сценарий FBO (основной) и FBS для сравнения.

Формула (на 1 продажу):
  комиссия      = цена * sales_percent/100
  логистика     = direct_flow_trans_min          (нижний порог тарифа)
  последняя миля= deliv_to_customer_amount
  эквайринг     = цена * ACQUIRING
  налог         = цена * TAX
  обр. логистика= return_flow_amount * (1 - выкуп)
  хранение      = STORAGE_L_DAY * STORAGE_DAYS * объём_л
  реклама       = цена * ДРР
  маржа_до_рекл = цена - комиссия - логистика - последняя_миля - эквайринг
                       - налог - обр.логистика - хранение - себестоимость
  маржа         = маржа_до_рекл - реклама
  ROI           = маржа / себестоимость * 100
"""
import json
from pathlib import Path

ROOT = Path(__file__).parent.parent.parent
PRICES = ROOT / 'data' / 'ozon_api' / 'prices.json'
OUT = ROOT / 'outputs' / 'ozon'

# --- Параметры сценария (явно, чтобы менять руками) --------------------
DRR = 0.10           # доля рекламных расходов от цены
TAX = 0.08           # налог с оборота (как в Celimax_unit_model, Юнит!B8)
ACQUIRING = 0.015    # эквайринг/приём платежей Ozon (допущение ~1.5%)
STORAGE_L_DAY = 0.16 # ₽/литр/день (Юнит!B6) — прокси для хранения
STORAGE_DAYS = 30    # дней хранения (Юнит!B7)
FULFILLMENT = 60.0   # фулфилмент/обработка, ₽/шт — складываем с себестоимостью

# Себестоимость по последнему приходу, ₽ (желтая колонка остатков).
COSTS = {
    'Celimax Dual Barrier Creamy Toner': 663.86,
    'FARMSTAY COLLAGEN&HYALURONIC ACID ALL-IN ONE': 347.76,
    'ROUND LAB 1025 DOKDO CLEANSER_150ml': 463.19,
}


def buyout(offer_id: str) -> float:
    """Доля выкупа: для пенок/гелей/мицелляр. воды 0.96, иначе 0.9219."""
    low = offer_id.lower()
    if any(w in low for w in ['cleanser', 'пенк', 'foam', 'мицелляр', 'гель для душа']):
        return 0.96
    return 0.9219


def num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def calc(item, scheme='fbo'):
    offer = item.get('offer_id', '')
    cost = COSTS.get(offer)
    if cost is None:
        return None
    price = num(item.get('price', {}).get('price'))
    vol = num(item.get('volume_weight'))
    c = item.get('commissions', {}) or {}

    if scheme == 'fbo':
        comm_pct = num(c.get('sales_percent_fbo'))
        logistics = num(c.get('fbo_direct_flow_trans_min_amount'))
        last_mile = num(c.get('fbo_deliv_to_customer_amount'))
        ret_base = num(c.get('fbo_return_flow_amount'))
    else:
        comm_pct = num(c.get('sales_percent_fbs'))
        logistics = num(c.get('fbs_direct_flow_trans_min_amount'))
        last_mile = num(c.get('fbs_deliv_to_customer_amount'))
        ret_base = num(c.get('fbs_return_flow_amount'))

    commission = price * comm_pct / 100
    acquiring = price * ACQUIRING
    tax = price * TAX
    reverse = ret_base * (1 - buyout(offer))
    storage = STORAGE_L_DAY * STORAGE_DAYS * vol
    ads = price * DRR
    cost_total = cost + FULFILLMENT  # себестоимость + фулфилмент

    margin_before = (price - commission - logistics - last_mile - acquiring
                     - tax - reverse - storage - cost_total)
    margin = margin_before - ads
    roi = margin / cost_total * 100 if cost_total else None
    # Цена безубыточности (маржа с рекламой = 0): решаем по цене.
    # Все проценты (comm_pct, ACQUIRING, TAX, DRR) — от цены, поэтому:
    #   price*(1 - k) = logistics + last_mile + reverse + storage + cost_total
    k = comm_pct / 100 + ACQUIRING + TAX + DRR
    fixed = logistics + last_mile + reverse + storage + cost_total
    breakeven = fixed / (1 - k) if k < 1 else None

    return {
        'offer': offer, 'scheme': scheme.upper(), 'price': price, 'cost': round(cost, 2),
        'fulfillment': round(FULFILLMENT, 2), 'cost_total': round(cost_total, 2),
        'comm_pct': comm_pct, 'commission': round(commission, 2),
        'logistics': round(logistics, 2), 'last_mile': round(last_mile, 2),
        'acquiring': round(acquiring, 2), 'tax': round(tax, 2),
        'reverse': round(reverse, 2), 'storage': round(storage, 2),
        'ads': round(ads, 2), 'drr_pct': round(DRR * 100, 1),
        'margin_before_ads': round(margin_before, 2),
        'margin': round(margin, 2),
        'roi_pct': round(roi, 1) if roi is not None else None,
        'margin_pct_price': round(margin / price * 100, 1) if price else None,
        'breakeven_price': round(breakeven) if breakeven else None,
    }


HEAD = ['Товар', 'Схема', 'Цена', 'Себест.', 'Фулфилмент', 'Себест.+фулфилмент',
        'Комиссия %', 'Комиссия ₽', 'Логистика', 'Посл.миля', 'Эквайринг',
        'Налог', 'Обр.лог', 'Хранение', 'Реклама (ДРР)', 'Маржа до рекл.',
        'Маржа/шт', 'Маржа % цены', 'ROI %', 'Цена безубыт.']


def row(r):
    return [r['offer'], r['scheme'], r['price'], r['cost'], r['fulfillment'],
            r['cost_total'], r['comm_pct'], r['commission'], r['logistics'],
            r['last_mile'], r['acquiring'], r['tax'], r['reverse'], r['storage'],
            r['ads'], r['margin_before_ads'], r['margin'], r['margin_pct_price'],
            r['roi_pct'], r['breakeven_price']]


def write_xlsx(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print('  openpyxl нет — пропускаю xlsx')
        return
    wb = Workbook()
    ws = wb.active
    ws.title = 'Юнитка Ozon'

    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='305496')
    cost_fill = PatternFill('solid', fgColor='FCE4D6')   # база затрат
    res_fill = PatternFill('solid', fgColor='E2EFDA')    # маржа/ROI

    ws.append(HEAD)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = head_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    money = '#,##0.00'
    int_fmt = '#,##0'
    pct = '0.0"%"'
    col = {name: i + 1 for i, name in enumerate(HEAD)}

    for r in rows:
        ws.append(row(r))
        rr = ws.max_row
        for i in range(1, len(HEAD) + 1):
            ws.cell(rr, i).border = border
        for name in ['Цена', 'Комиссия ₽', 'Логистика', 'Посл.миля', 'Эквайринг',
                     'Налог', 'Обр.лог', 'Хранение', 'Реклама (ДРР)',
                     'Маржа до рекл.', 'Маржа/шт', 'Цена безубыт.']:
            ws.cell(rr, col[name]).number_format = int_fmt
        for name in ['Себест.', 'Фулфилмент', 'Себест.+фулфилмент']:
            ws.cell(rr, col[name]).number_format = money
            ws.cell(rr, col[name]).fill = cost_fill
        for name in ['Комиссия %', 'Маржа % цены', 'ROI %']:
            ws.cell(rr, col[name]).number_format = pct
        ws.cell(rr, col['Маржа/шт']).fill = res_fill
        ws.cell(rr, col['ROI %']).fill = res_fill
        ws.cell(rr, col['Маржа/шт']).font = Font(bold=True)
        ws.cell(rr, col['ROI %']).font = Font(bold=True)

    widths = {'Товар': 42, 'Схема': 7, 'Себест.+фулфилмент': 16, 'Реклама (ДРР)': 12,
              'Маржа до рекл.': 12, 'Маржа % цены': 11, 'Цена безубыт.': 12}
    for name, i in col.items():
        letter = ws.cell(1, i).column_letter
        ws.column_dimensions[letter].width = widths.get(name, 11)
    ws.freeze_panes = 'C2'
    ws.row_dimensions[1].height = 30

    xlsx = OUT / 'ozon_unit_economics_2026-08-03.xlsx'
    wb.save(xlsx)
    print(f"Сохранено: {xlsx}")


if __name__ == '__main__':
    items = json.loads(PRICES.read_text())
    rows = []
    for it in items:
        if it.get('offer_id') not in COSTS:
            continue
        for scheme in ('fbo', 'fbs'):
            r = calc(it, scheme)
            if r:
                rows.append(r)

    # печать в консоль
    print(f"Параметры: ДРР={DRR*100:.0f}%  налог={TAX*100:.0f}%  "
          f"эквайринг={ACQUIRING*100:.1f}%  выкуп: пенка 96% / прочее 92,19%")
    for r in rows:
        print(f"\n{r['offer']}  [{r['scheme']}]")
        print(f"  цена {r['price']:.0f} | себест. {r['cost']:.2f} + фулфилмент "
              f"{r['fulfillment']:.0f} = {r['cost_total']:.2f} | "
              f"комиссия {r['comm_pct']:.0f}% = {r['commission']:.0f} | "
              f"логистика {r['logistics']:.0f} | посл.миля {r['last_mile']:.0f} | "
              f"эквайринг {r['acquiring']:.0f} | налог {r['tax']:.0f} | "
              f"обр.лог {r['reverse']:.1f} | хранение {r['storage']:.2f} | "
              f"реклама {r['ads']:.0f}")
        print(f"  МАРЖА/шт {r['margin']:.0f} ₽ ({r['margin_pct_price']:.1f}% цены) | "
              f"ROI {r['roi_pct']:.1f}% | безубыток при цене ~{r['breakeven_price']} ₽")

    OUT.mkdir(parents=True, exist_ok=True)
    tsv = OUT / 'ozon_unit_economics_2026-08-03.tsv'
    lines = ['\t'.join(HEAD)]
    for r in rows:
        lines.append('\t'.join(str(x) for x in row(r)))
    tsv.write_text('\n'.join(lines), encoding='utf-8')
    print(f"\nСохранено: {tsv}")

    write_xlsx(rows)
