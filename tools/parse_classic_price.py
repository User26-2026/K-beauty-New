"""Разбор прайсов поставщика Классик в единую таблицу.

Прайсы приходят от разных брендов с разной версткой: заголовок может быть
на любой из первых строк, часть шапок двухэтажные, названия колонок пишутся
по-разному. Скрипт сам находит строку заголовка и раскладывает данные в
общий набор полей.

Себестоимость считаем по правилу проекта:
    себестоимость_руб = закупка_KRW * курс * 1.4

Запуск:
    python3 tools/parse_classic_price.py                    # все прайсы Классика
    python3 tools/parse_classic_price.py --rate 0.0598      # свой курс воны
    python3 tools/parse_classic_price.py путь.xlsx ...      # конкретные файлы
"""

import argparse
import glob
import os
import re

import openpyxl
import pandas as pd

PRICE_DIR = "data/price_lists/classic"
OUT_DIR = "outputs"

# Курс ЦБ, рублей за 1 вону. Обновляем на дату расчета через --rate.
KRW_RUB = 0.058
# Логистика, пошлина и приемка сверх закупочной цены.
IMPORT_MULTIPLIER = 1.4

# Как узнаем колонку: поле -> список шаблонов по тексту шапки.
COLUMN_PATTERNS = {
    # 구분/Division в прайсах — порядковый номер строки, не бренд.
    "brand": [r"^brand$", r"^бренд$"],
    "code": [r"sku\s*no", r"^code$", r"product\s*code", r"sap\s*code", r"^артикул$"],
    "barcode": [r"bar\s*code", r"barcode", r"바코드"],
    # Названия: сначала колонки с явной пометкой языка (STRONG_PATTERNS),
    # потом общие подписи. Где обе колонки подписаны одинаково, корейская
    # идет первой, поэтому name_kr проверяется раньше name_en.
    "name_kr": [r"^product\s*name$", r"^product$", r"^name$"],
    "name_en": [r"^product\s*name$", r"^product$", r"^name$"],
    # LEBELAGE ведет отдельную колонку с русскими названиями — забираем.
    "name_ru": [r"наименование", r"название"],
    "type": [r"^type$", r"^category$", r"product\s*line"],
    "volume": [r"^vol", r"volume", r"^size$", r"capacity", r"용량"],
    "msrp_krw": [
        r"msrp", r"^srp\b", r"^retail\b", r"retail\s*price", r"list\s*price",
        r"consumer\s*price", r"regular\s*price\s*\(krw",
        # Корейские подписи розничной цены: 소비자가 / 정상소가 / 기준가.
        r"소비자", r"정상소가", r"기준가",
    ],
    # Закупку пишут по-разному: supply / unit / просто price (-VAT).
    "supply_krw": [
        r"supply\s*price", r"fob\s*price", r"공급가", r"distributor\s*price",
        r"unit\s*price", r"^price\s*\(\s*-?\s*vat",
    ],
    "qty_per_box": [r"q'?ty\s*/?\s*box", r"qty\s*per\s*outbox", r"1\s*box\s*qty", r"ea\s*/\s*box", r"^master$"],
    "moq": [r"^moq", r"moq\s*qty"],
    "shelf_life": [r"shelf\s*life", r"유통기한"],
    "status": [r"^status$", r"^remark$", r"^비고$"],
}

# Явные языковые подписи — их разбираем до общих шаблонов названия.
STRONG_PATTERNS = {
    "name_kr": [r"\bname\b.*\b(ko|kr|kor|korean)\b", r"국문", r"제품명", r"품명"],
    "name_en": [
        r"\bname\b.*\b(en|eng|english)\b", r"영문", r"^eng$",
        r"description.*\b(en|eng|english)\b",
    ],
}

# По этим словам ищем саму строку заголовка.
HEADER_HINTS = [
    "barcode", "bar code", "product name", "supply price", "msrp", "vol",
    "brand", "moq", "retail price", "code",
]


def norm(value):
    """Текст ячейки в нижнем регистре, переносы строк схлопнуты в пробел."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def find_header_row(rows, limit=12):
    """Строка заголовка — та, где больше всего опорных слов."""
    best_idx, best_score = None, 0
    for idx, row in enumerate(rows[:limit]):
        cells = [norm(c) for c in row]
        score = sum(1 for c in cells if c and any(h in c for h in HEADER_HINTS))
        if score > best_score:
            best_idx, best_score = idx, score
    return best_idx if best_score >= 2 else None


def merge_header(rows, header_idx):
    """Склеиваем двухэтажную шапку: подпись сверху + уточнение снизу."""
    top = [norm(c) for c in rows[header_idx]]
    if header_idx + 1 >= len(rows):
        return top
    below = [norm(c) for c in rows[header_idx + 1]]
    extra = ["price", "weight", "size", "cbm", "국문", "영문", "korean", "english"]
    hits = sum(1 for c in below if c and any(h in c for h in HEADER_HINTS + extra))
    if hits < 2:
        return top
    return [(t + " " + b).strip() for t, b in zip(top, below)]


def map_columns(header):
    """Шапка -> {поле: номер колонки}.

    Два прохода. Сначала колонки с явной пометкой языка: JIGOTT называет
    английское название "Description (ENG)", CP1 подписывает языки во втором
    этаже шапки. Потом общие подписи — там, где AMORE и FLOR DE MAN зовут обе
    колонки одинаково ("Product Name" / "Product"), первая колонка корейская.
    """
    mapping, used = {}, set()
    for patterns_set in (STRONG_PATTERNS, COLUMN_PATTERNS):
        for col_idx, title in enumerate(header):
            if not title or col_idx in used:
                continue
            for field, patterns in patterns_set.items():
                if field in mapping:
                    continue
                if any(re.search(p, title) for p in patterns):
                    mapping[field] = col_idx
                    used.add(col_idx)
                    break
    return mapping


def to_number(value):
    """Число из ячейки: убираем валюту, пробелы и разделители тысяч."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^\d.,\-]", "", str(value)).replace(",", "")
    if not re.search(r"\d", text):
        return None
    try:
        return float(text)
    except ValueError:
        return None


KOREAN = r"\uac00-\ud7a3"


def split_languages(text):
    """Разносит название на английское и корейское.

    Часть поставщиков (FLORODORA, BERGAMO, AMORE, CP1) кладет оба названия
    в одну ячейку — через перенос строки или просто подряд.
    """
    if not text or not re.search(f"[{KOREAN}]", text):
        return text, ""
    korean = " ".join(re.findall(f"[{KOREAN}][{KOREAN}\\s]*", text)).strip()
    latin = " ".join(
        part.strip() for part in re.split(f"[{KOREAN}]+", text)
        if re.search(r"[A-Za-z]{2}", part)
    ).strip()
    return re.sub(r"\s+", " ", latin), re.sub(r"\s+", " ", korean)


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def brand_from_filename(path):
    """Бренд из имени файла. В именах встречаются опечатки вида PRICE_LEST."""
    name = re.sub(r"^classic_", "", os.path.basename(path))
    name = re.split(r"_?PRICE[_ ]?L[EI]S?T|_\d{2}\.\d{2}", name, flags=re.I)[0]
    return name.replace("_", " ").strip(" ._").upper()


def parse_sheet(ws, source, sheet_name, fallback_brand):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx = find_header_row(rows)
    if header_idx is None:
        return [], "заголовок не найден"

    header = merge_header(rows, header_idx)
    mapping = map_columns(header)
    if "supply_krw" not in mapping:
        return [], "нет колонки закупочной цены"

    records = []
    last_brand = fallback_brand
    for row in rows[header_idx + 1:]:
        supply = to_number(row[mapping["supply_krw"]]) if mapping["supply_krw"] < len(row) else None
        if not supply:
            continue

        def cell(field):
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        name_en = clean_text(cell("name_en"))
        name_kr = clean_text(cell("name_kr"))
        if not name_en and not name_kr:
            continue
        # Колонку без языка в шапке могли занять под латиницу — это название EN.
        if not name_en and not re.search(f"[{KOREAN}]", name_kr):
            name_en, name_kr = name_kr, ""
        for candidate in (name_en, name_kr):
            latin, korean = split_languages(candidate)
            if not (latin and korean):
                continue
            name_en = latin
            if not re.search(f"[{KOREAN}]", name_kr) or name_kr == candidate:
                name_kr = korean
            break

        brand = clean_text(cell("brand")) or last_brand
        last_brand = brand  # в прайсах бренд ставят только в первой строке блока

        records.append({
            "Файл": os.path.basename(source),
            "Лист": sheet_name,
            "Бренд": brand,
            "Артикул": clean_text(cell("code")),
            "Штрихкод": clean_text(cell("barcode")),
            "Название EN": name_en,
            "Название KR": name_kr,
            "Название RU": clean_text(cell("name_ru")),
            "Тип": clean_text(cell("type")),
            "Объем": clean_text(cell("volume")),
            "MSRP, KRW": to_number(cell("msrp_krw")),
            "Закупка, KRW": supply,
            "Шт/короб": to_number(cell("qty_per_box")),
            "MOQ": to_number(cell("moq")),
            "Срок годности": clean_text(cell("shelf_life")),
            "Примечание": clean_text(cell("status")),
        })
    return records, None


def dedupe(records, notes):
    """Часть поставщиков дублирует каталог на листах 국문 и English.

    Совпадение ищем по артикулу, а если его нет — по паре штрихкод + название.
    Оставляем первую строку, но если у дубля другая закупочная цена, это не
    техническая копия, а второй уровень цен — говорим об этом отдельно.
    """
    unique, seen, conflicts = [], {}, []
    for rec in records:
        key = rec["Артикул"] or (rec["Штрихкод"], rec["Название EN"])
        if not (rec["Артикул"] or rec["Штрихкод"]):
            unique.append(rec)
            continue
        if key in seen:
            kept = seen[key]
            if kept["Закупка, KRW"] != rec["Закупка, KRW"]:
                conflicts.append((kept, rec))
            continue
        seen[key] = rec
        unique.append(rec)

    dropped = len(records) - len(unique)
    if dropped:
        notes.append(f"убрано дублей между листами: {dropped}")
    if conflicts:
        kept, other = conflicts[0]
        notes.append(
            f"РАЗНЫЕ ЦЕНЫ на листах ({len(conflicts)} SKU), взят лист "
            f"'{kept['Лист']}': напр. {kept['Артикул'] or kept['Штрихкод']} — "
            f"{kept['Закупка, KRW']:.0f} против {other['Закупка, KRW']:.0f} KRW"
        )
    return unique, notes


def pick_sheets(sheet_names):
    """Отбираем листы, по которым считать закупку.

    LAGOM дает один каталог дважды: VAT포함 (с НДС) и VAT별도 (без НДС), плюс
    листы 검산/검수 — внутренняя сверка. Нам нужна цена без НДС.
    """
    sheets = [s for s in sheet_names if not re.search(r"검산|검수", s)]
    if any("별도" in s for s in sheets):
        sheets = [s for s in sheets if "포함" not in s]
    return sheets


def parse_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    fallback_brand = brand_from_filename(path)
    records, notes = [], []
    for sheet_name in pick_sheets(wb.sheetnames):
        found, problem = parse_sheet(wb[sheet_name], path, sheet_name, fallback_brand)
        records.extend(found)
        if problem:
            notes.append(f"{sheet_name}: {problem}")
    wb.close()

    unique, notes = dedupe(records, notes)
    return unique, notes


def main(paths, rate):
    if not paths:
        paths = sorted(glob.glob(os.path.join(PRICE_DIR, "*.xlsx")))

    all_records = []
    print(f"{'Файл':<52} {'SKU':>5}  Замечания")
    print("-" * 92)
    for path in paths:
        records, notes = parse_file(path)
        all_records.extend(records)
        print(f"{os.path.basename(path):<52} {len(records):>5}  {'; '.join(notes)}")

    if not all_records:
        print("\nНичего не разобрано.")
        return

    df = pd.DataFrame(all_records)
    # Наценка поставщика к рекомендованной рознице — грубый ориентир по марже.
    df["MSRP/Закупка"] = (df["MSRP, KRW"] / df["Закупка, KRW"]).round(2)
    df["Себестоимость, руб"] = (df["Закупка, KRW"] * rate * IMPORT_MULTIPLIER).round(2)
    df["Курс KRW"] = rate

    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, "classic_prices_normalized.xlsx")
    df.to_excel(out_path, index=False)
    print(f"\nВсего SKU: {len(df)}   Брендов: {df['Бренд'].nunique()}")
    print(f"Курс: {rate} руб/вона, множитель импорта {IMPORT_MULTIPLIER}")
    print(f"Сохранено: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("paths", nargs="*", help="конкретные файлы прайсов")
    parser.add_argument("--rate", type=float, default=KRW_RUB, help="курс рублей за 1 вону")
    ns = parser.parse_args()
    main(ns.paths, ns.rate)
