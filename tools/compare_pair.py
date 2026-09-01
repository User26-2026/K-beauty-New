"""Сравнение двух поставщиков по всем брендам сразу.

Пример: Классик против FINESKIN. Обоих сводим по штрихкоду, цену берем
за штуку в воне, фасовку выравниваем через unify_packs. Сравниваем только
внутри одной страны — у разных стран разный маршрут и разные расходы.

Запуск:
    python3 tools/compare_pair.py Классик FINESKIN
"""

import argparse
import os
import re

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from price_unit import unify_packs

SRC = "outputs/prices_normalized.xlsx"
OUT_DIR = "outputs"
RATE = 0.058          # рублей за вону
IMPORT = 1.4          # логистика, пошлина и приемка

TRANSLIT = {"а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ж": "zh",
            "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n",
            "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f",
            "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch", "ы": "y",
            "э": "e", "ю": "yu", "я": "ya", "ь": "", "ъ": ""}


def slugify(text):
    """Имя файла латиницей: поставщики называются и по-русски."""
    lowered = "".join(TRANSLIT.get(char, char) for char in text.lower())
    return re.sub(r"[^a-z0-9]+", "_", lowered).strip("_") or "pair"


HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BEST_FILL = PatternFill("solid", fgColor="E2EFDA")


def load(first, second, country):
    df = pd.read_excel(SRC, dtype={"Штрихкод": str})
    df = df[(df["Страна"] == country) & df["Поставщик"].isin([first, second])]
    df["Штрихкод"] = df["Штрихкод"].fillna("").astype(str).str.replace(r"\D", "", regex=True)
    df = df[df["Штрихкод"].str.len() >= 8]
    return df[df["Закупка, KRW"].notna() & (df["Закупка, KRW"] > 0)].copy()


def brand_of(rows):
    """Бренд у поставщиков пишется по-разному, а в части прайсов в колонке
    бренда стоит категория (SERUM, CUSHION). По штрихкоду берем то значение,
    которое чаще встречается как бренд во всей выборке."""
    brands = rows["Бренд"].fillna("").astype(str).str.upper().str.strip()
    brands = brands.replace("", pd.NA)
    counts = brands.value_counts()
    frame = pd.DataFrame({"ean": rows["Штрихкод"], "brand": brands}).dropna()
    frame["weight"] = frame["brand"].map(counts)
    best = (frame.sort_values("weight", ascending=False)
                 .drop_duplicates("ean").set_index("ean")["brand"])
    return rows["Штрихкод"].map(best).fillna(brands)


def positions(df, first, second, column):
    """Одна строка на штрихкод: цена у первого, цена у второго, разница."""
    best = df.sort_values(column).drop_duplicates(["Штрихкод", "Поставщик"])
    prices = best.pivot(index="Штрихкод", columns="Поставщик", values=column)
    units = best.pivot(index="Штрихкод", columns="Поставщик", values="Единица цены")
    packs = best.pivot(index="Штрихкод", columns="Поставщик", values="Штук в упаковке (сводно)")
    info = (best.sort_values("Название EN", key=lambda s: s.astype(str).str.len(), ascending=False)
                .drop_duplicates("Штрихкод")
                .set_index("Штрихкод")[["Бренд", "Название EN", "Объем", "Базис", "MSRP, KRW"]])

    table = info.join(prices, how="right")
    for supplier in (first, second):
        if supplier not in table.columns:
            table[supplier] = pd.NA
        table[f"{supplier}: единица"] = units.get(supplier)
        table[f"{supplier}: шт в уп."] = packs.get(supplier)
    table["Фасовка совпала"] = (packs.get(first).fillna(1) == packs.get(second).fillna(1))
    return table


def diff_columns(table, first, second):
    pair = table[[first, second]]
    table["Дешевле у"] = pair.idxmin(axis=1)
    table["Мин, KRW"] = pair.min(axis=1)
    table["Макс, KRW"] = pair.max(axis=1)
    # Насколько дешевле победитель по отношению к проигравшему.
    table["Дешевле на, %"] = ((1 - table["Мин, KRW"] / table["Макс, KRW"]) * 100).round(1)
    # Насколько дороже проигравший по отношению к победителю — это переплата.
    table["Переплата, %"] = ((table["Макс, KRW"] / table["Мин, KRW"] - 1) * 100).round(1)
    table["Разница, KRW"] = (table["Макс, KRW"] - table["Мин, KRW"]).round(0)
    table["Разница, руб"] = (table["Разница, KRW"] * RATE * IMPORT).round(0)
    # Расхождение больше 30% бывает настоящим, но чаще это разная единица цены
    # или опечатка в прайсе. Из рейтинга не выкидываем, помечаем на сверку.
    table["Проверить у менеджера"] = table["Переплата, %"].ge(30).map({True: "да", False: ""})
    return table


def by_brand(shared, first, second):
    """Свод по брендам: где дешевле, на сколько и какая корзина."""
    rows = []
    for brand, group in shared.groupby(shared["Бренд"].fillna("БЕЗ БРЕНДА")):
        wins_first = int((group["Дешевле у"] == first).sum())
        wins_second = int((group["Дешевле у"] == second).sum())
        basket_first = group[first].sum()
        basket_second = group[second].sum()
        cheaper = first if basket_first < basket_second else second
        gap = abs(basket_first - basket_second) / max(basket_first, basket_second) * 100
        rows.append({
            "Бренд": brand,
            "Общих позиций": len(group),
            f"Дешевле у {first}": wins_first,
            f"Дешевле у {second}": wins_second,
            "Медиана разницы, %": round(group["Дешевле на, %"].median(), 1),
            f"Корзина {first}, KRW": round(basket_first),
            f"Корзина {second}, KRW": round(basket_second),
            "Корзина дешевле у": cheaper,
            "Корзина дешевле на, %": round(gap, 1),
            "Разница корзины, руб": round(abs(basket_first - basket_second) * RATE * IMPORT),
        })
    return pd.DataFrame(rows).sort_values("Общих позиций", ascending=False)


def brand_coverage(df, first, second):
    """Какие бренды есть у каждого и где они пересекаются."""
    counts = (df.groupby(["Бренд", "Поставщик"])["Штрихкод"].nunique()
                .unstack(fill_value=0))
    for supplier in (first, second):
        if supplier not in counts.columns:
            counts[supplier] = 0
    counts = counts[[first, second]].reset_index()
    counts.columns = ["Бренд", f"SKU у {first}", f"SKU у {second}"]
    counts["Есть у обоих"] = ((counts[f"SKU у {first}"] > 0)
                              & (counts[f"SKU у {second}"] > 0)).map({True: "да", False: ""})
    return counts.sort_values([f"SKU у {first}", f"SKU у {second}"], ascending=False)


def format_sheet(worksheet, percent_titles=(), money_titles=()):
    worksheet.freeze_panes = "B2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    widths = {"Название EN": 52, "Объем": 16, "Штрихкод": 15, "Бренд": 20,
              "Дешевле у": 16, "Корзина дешевле у": 18, "Есть только у": 16}
    titles = [str(c.value or "") for c in worksheet[1]]
    for index, title in enumerate(titles, start=1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = widths.get(title, 14)
        if title in percent_titles:
            fmt = "0.0"
        elif title in money_titles or "KRW" in title or "руб" in title:
            fmt = "# ##0"
        else:
            continue
        for row in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
            row[0].number_format = fmt
    if "Дешевле у" in titles:
        winner_col = titles.index("Дешевле у")
        for row in worksheet.iter_rows(min_row=2):
            winner = row[winner_col].value
            if winner in titles:
                row[titles.index(winner)].fill = BEST_FILL


def paint(worksheet, title, first_row, last_row, top=30):
    titles = [str(c.value or "") for c in worksheet[1]]
    if title not in titles:
        return
    letter = get_column_letter(titles.index(title) + 1)
    worksheet.conditional_formatting.add(
        f"{letter}{first_row}:{letter}{last_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                       mid_type="num", mid_value=top / 3, mid_color="FFEB84",
                       end_type="num", end_value=top, end_color="F8696B"))


def main(first, second, country):
    if not os.path.exists(SRC):
        raise SystemExit(f"Нет файла {SRC} — сначала запустите parse_price_lists.py")

    df = load(first, second, country)
    if df.empty:
        raise SystemExit(f"Нет позиций для {first} и {second} в стране {country}")
    df["Бренд"] = brand_of(df)

    df, conflicts = unify_packs(df)
    if conflicts:
        print(f"Разная фасовка по одному штрихкоду: {len(conflicts)} позиций")

    column = "Цена за штуку (сводно)"
    table = positions(df, first, second, column)
    shared = table[table[first].notna() & table[second].notna()].copy()
    shared = diff_columns(shared, first, second)
    shared = shared.sort_values(["Бренд", "Переплата, %"], ascending=[True, False])

    only_first = table[table[first].notna() & table[second].isna()].copy()
    only_second = table[table[second].notna() & table[first].isna()].copy()

    brands = by_brand(shared, first, second)
    basket_first = shared[first].sum()
    basket_second = shared[second].sum()
    cheaper = first if basket_first < basket_second else second
    gap = abs(basket_first - basket_second) / max(basket_first, basket_second) * 100

    total = pd.DataFrame([
        {"Показатель": "Общих позиций по штрихкоду", "Значение": len(shared)},
        {"Показатель": f"Позиций только у {first}", "Значение": len(only_first)},
        {"Показатель": f"Позиций только у {second}", "Значение": len(only_second)},
        {"Показатель": f"Дешевле у {first}, позиций",
         "Значение": int((shared["Дешевле у"] == first).sum())},
        {"Показатель": f"Дешевле у {second}, позиций",
         "Значение": int((shared["Дешевле у"] == second).sum())},
        {"Показатель": "Медиана разницы по позициям, %",
         "Значение": round(shared["Дешевле на, %"].median(), 1)},
        {"Показатель": "Средняя разница по позициям, %",
         "Значение": round(shared["Дешевле на, %"].mean(), 1)},
        {"Показатель": f"Корзина общих позиций, {first}, KRW", "Значение": round(basket_first)},
        {"Показатель": f"Корзина общих позиций, {second}, KRW", "Значение": round(basket_second)},
        {"Показатель": "Корзина дешевле у", "Значение": cheaper},
        {"Показатель": "Корзина дешевле на, %", "Значение": round(gap, 1)},
        {"Показатель": "Разница корзины, руб",
         "Значение": round(abs(basket_first - basket_second) * RATE * IMPORT)},
        {"Показатель": "Позиций с расхождением 30%+ (на сверку)",
         "Значение": int((shared["Проверить у менеджера"] == "да").sum())},
    ])

    slug = slugify(f"{first}_{second}")
    out_path = os.path.join(OUT_DIR, f"pair_{slug}_{country.lower()}.xlsx")
    with pd.ExcelWriter(out_path) as writer:
        total.to_excel(writer, sheet_name="ИТОГО", index=False)
        brands.to_excel(writer, sheet_name="ПО БРЕНДАМ", index=False)
        coverage = brand_coverage(df, first, second)
        coverage.to_excel(writer, sheet_name="БРЕНДЫ", index=False)
        shared.to_excel(writer, sheet_name="ПОЗИЦИИ")
        only_first.to_excel(writer, sheet_name=f"ТОЛЬКО {first}"[:31])
        only_second.to_excel(writer, sheet_name=f"ТОЛЬКО {second}"[:31])

        book = writer.book
        format_sheet(book["ИТОГО"])
        format_sheet(book["ПО БРЕНДАМ"],
                     percent_titles=("Медиана разницы, %", "Корзина дешевле на, %"))
        paint(book["ПО БРЕНДАМ"], "Медиана разницы, %", 2, len(brands) + 1)
        format_sheet(book["ПОЗИЦИИ"],
                     percent_titles=("Дешевле на, %", "Переплата, %"))
        paint(book["ПОЗИЦИИ"], "Переплата, %", 2, len(shared) + 1, top=50)
        for name in ("БРЕНДЫ", f"ТОЛЬКО {first}"[:31], f"ТОЛЬКО {second}"[:31]):
            format_sheet(book[name])

    print(f"\n{first} против {second}, {country}, цены в KRW за штуку")
    print(total.to_string(index=False))
    print("\nПо брендам:")
    print(brands.to_string(index=False))
    print(f"\nФайл: {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Сравнение двух поставщиков по всем брендам")
    parser.add_argument("first")
    parser.add_argument("second")
    parser.add_argument("--country", default="KR")
    args = parser.parse_args()
    main(args.first, args.second, args.country)
