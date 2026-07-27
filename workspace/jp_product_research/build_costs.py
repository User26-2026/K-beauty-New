#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Консолидация себестоимости по японским прайсам.
себестоимость_руб = цена_JPY * КУРС * 1.4
"""
import openpyxl
from openpyxl import Workbook

UP = "data/price_lists/japan/"
RATE = 0.476   # JPY->RUB на 27.07.2026 (ЦБ РФ)
K = 1.4        # логистика + пошлина + приемка

def rub(jpy):
    try:
        return round(float(jpy) * RATE * K)
    except (TypeError, ValueError):
        return None

def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None

rows = []  # (source, brand/series, name, volume, jan, cost_jpy, cost_rub, minlot)

def load(f, sheet=None):
    wb = openpyxl.load_workbook(UP+f, read_only=True, data_only=True)
    return wb[sheet] if sheet else wb.active

# ---- 1) KOR RUSSIAN PRICE (два блока в строке) ----
ws = load("KOR_russian_price_2026-07-14.xlsx")
data = [[("" if c is None else str(c)).strip() for c in r] for r in ws.iter_rows(values_only=True)]
def clean(s): return " ".join(s.split())
for r in data:
    for base in (1, 8):  # левый блок col2.., правый блок col9..
        if len(r) <= base+5: continue
        jan, name, vol, lot, retail, cost = r[base:base+6]
        if jan.isdigit() and len(jan) >= 12 and name:
            rows.append(("KOR", "KOR/多ブランド", clean(name), vol, jan, num(cost), rub(cost), lot))

# ---- 1b) KOR NEW ----
ws = load("KOR_new_russian_price.xlsx")
for r in list(ws.iter_rows(values_only=True))[1:]:
    jan, name, price, lot = (str(x).strip() if x is not None else "" for x in r[:4])
    if jan.isdigit() and name:
        rows.append(("KOR-NEW", "KOR", clean(name), "", jan, num(price), rub(price), lot))

# ---- 2) MANDOM (LUCIDO-L / Bifesta) ----
ws = load("Mandom_LucidoL_Bifesta_2026-07.xlsx")
for r in list(ws.iter_rows(values_only=True))[3:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 11: continue
    jan, brand, name_en, vol, price = r[2], r[5], r[7], r[8], r[10]
    name = name_en or r[6]
    if jan.isdigit() and len(jan) >= 12 and name:
        rows.append(("MANDOM", brand, clean(name), vol, jan, num(price), rub(price), ""))

# ---- 3) COSME STATION (納入価核 = стоимость) ----
ws = load("CosmeStation_2026.xlsx")
series = ""
for r in list(ws.iter_rows(values_only=True))[5:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 25: continue
    if r[0]: series = r[0]
    name, jan, vol = r[1], r[3], r[6]
    teika, nouka, biko = r[22], r[23], r[24]  # 定価 / 納入価核 / 備考
    if jan.isdigit() and len(jan) >= 12 and name and num(nouka):
        rows.append(("COSME-ST", series, clean(name), vol, jan, num(nouka), rub(nouka), ""))

# ---- 4) DOSHISHA ----
ws = load("Doshisha.xlsx", "Price")
for r in list(ws.iter_rows(values_only=True))[6:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 8: continue
    jan, name_jp, name_ru, cap, price = r[1], r[3], r[4], r[5], r[7]
    name = name_ru or name_jp
    if jan.isdigit() and len(jan) >= 12 and name and num(price):
        rows.append(("DOSHISHA", "DOSHISHA", clean(name), cap, jan, num(price), rub(price), ""))

# ---- 5) FINE (БАДы/еда — отдельно помечаем) ----
ws = load("Fine_supplements.xlsx")
for r in list(ws.iter_rows(values_only=True))[7:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 8: continue
    maker, jan, name_jp, name_en, name_ru, ppc, price = r[1], r[2], r[3], r[4], r[5], r[6], r[7]
    name = name_ru or name_en or name_jp
    if jan.isdigit() and len(jan) >= 12 and name and num(price):
        rows.append(("FINE-БАД", maker, clean(name), "", jan, num(price), rub(price), ""))

# ---- Запись ----
wb = Workbook()
ws = wb.active; ws.title = "Себестоимость"
hdr = ["Источник","Бренд/серия","Наименование","Объём","JAN","Цена JPY","Себест. RUB","Мин.лот"]
ws.append(hdr)
for row in rows:
    ws.append(list(row))
out = "outputs/japan_sebestoimost_2026-07-27.xlsx"
wb.save(out)

# статистика
from collections import Counter
cnt = Counter(r[0] for r in rows)
print("Курс JPY->RUB на 27.07.2026 =", RATE, "| множитель", K)
print("Всего SKU с ценой:", len(rows))
for k,v in cnt.items(): print(f"  {k}: {v}")
print("Файл:", out)
# TSV превью top по каждому источнику
EOF_MARK = None
