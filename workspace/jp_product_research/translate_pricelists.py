#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Перевод японских прайсов на русский для чтения.
Читает исходные прайсы, добавляет колонку 'Наименование (рус)',
объём, JAN, цену JPY и себестоимость в рублях. Один лист на поставщика.
"""
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

UP = "data/price_lists/japan/"
RATE = 0.476
K = 1.4
def rub(j):
    try: return round(float(j)*RATE*K)
    except: return None
def num(x):
    try: return float(x)
    except: return None
def clean(s): return " ".join(str(s).split())

# --- словарь терминов (регистронезависимо, длинные фразы первыми) ---
EN = [
 ("botanical all in one gel","ботанический гель всё-в-одном"),
 ("moisture keep face all-in-one gel","гель всё-в-одном для лица, удержание влаги"),
 ("all in one gel","гель всё-в-одном"),("all-in-one gel","гель всё-в-одном"),("all-ione gel","гель всё-в-одном"),
 ("all in one essence","эссенция всё-в-одном"),("all in one mask","маска всё-в-одном"),
 ("skin lotion","лосьон для лица"),("power skin serum","укрепляющая сыворотка"),
 ("facial treatment mask","уходовая тканевая маска"),("facial mask","тканевая маска для лица"),
 ("white face mask","тканевая маска для лица, осветляющая"),("face mask","тканевая маска для лица"),
 ("essence mask","эссенц-маска"),("eye sheet","патчи под глаза"),("eye lash serum","сыворотка для ресниц"),
 ("eyes all in one essence","эссенция всё-в-одном для век"),
 ("honey oil hair treatment","медовое масло для волос"),("oil hair treatment","масло для волос"),
 ("hair rescue hair treatment oil","восстанавливающее масло для волос"),
 ("hair treatment oil","масло-уход для волос"),("hair oil","масло для волос"),
 ("hair milk","молочко для волос"),("hair mask","маска для волос"),("hair wax","воск для волос"),
 ("hair jam","гель-желе для укладки"),("hair repair treatment oil","восстанавливающее масло для волос"),
 ("oil control spray","спрей, контроль жирности"),
 ("premium pure oil","премиальное чистое масло"),("argan rich oil","аргановое масло"),
 ("micellar cleansing water","мицеллярная вода"),("micellar eye makeup remover","мицеллярное средство для снятия макияжа с глаз"),
 ("micellar cleansing sheet","мицеллярные салфетки"),("cleansing sheet","очищающие салфетки"),
 ("serum cleansing oil refill","гидрофильное масло-сыворотка, рефилл"),("serum cleansing oil","гидрофильное масло-сыворотка"),
 ("foaming whip","пенка-мусс для умывания"),("ampoule mist","ампульный мист"),("ampoule milk","ампульное молочко"),
 ("lip scrub balm","скраб-бальзам для губ"),("lip serum mask","сывороточная маска для губ"),("lip serum","сыворотка для губ"),
 ("high class moist essence","увлажняющая эссенция премиум"),("moist essence","увлажняющая эссенция"),
 ("white moist essence","увлажняющая эссенция, осветляющая"),("white deep serum","глубокая осветляющая сыворотка"),
 ("deep serum","глубокая сыворотка"),("pure essence","чистая эссенция"),("delivery serum","сыворотка-транспорт"),
 ("delivery jelly","желе-транспорт"),("power jelly","укрепляющее желе"),("delivery face mask","тканевая маска-транспорт"),
 ("power face mask","укрепляющая тканевая маска"),("uv all in one gel","солнцезащитный гель всё-в-одном"),
 ("spicule foundation","тональная основа со спикулами"),("foundation","тональная основа"),
 ("tansan mist","карбонизированный мист"),("stick balm","стик-бальзам"),
 ("graphene","графеновая"),("micro current face mask","микротоковая тканевая маска"),
 ("straight hair iron","выпрямитель для волос"),("moroccan","марокканское"),("huile pour les cheveux","масло для волос"),
 ("bio total oil","биомасло для волос"),("science delivery","система доставки"),("secret science","secret science"),
 ("shiratama suhada","гладкая кожа «сиратама»"),("uruoi power","увлажняющая сила"),("moisture keep","удержание влаги"),
 # общие токены
 ("brightup","осветление"),("bright up","осветление"),("serum moist","увлажнение"),("deep clear","глубокое очищение"),
 ("acne care","уход за проблемной кожей"),("perfect clear","идеальное очищение"),("perfect glow","идеальное сияние"),
 ("premium","премиум"),("night & vitamin","ночной уход + витамины"),("retinol&vitamin","ретинол+витамины"),
 ("retinol＆vitamin","ретинол+витамины"),("vitamin","витамин"),("retinol","ретинол"),("white","осветляющий"),
 ("stem cell","стволовые клетки"),("exosome","экзосомы"),("exosomes","экзосомы"),("glutathione","глутатион"),
 ("pdrn","PDRN"),("nmn","NMN"),("bio liposome","биолипосомы"),("ips cells","iPS-клетки"),("ips cell","iPS-клетка"),
 ("human","человеческие"),("science",""),("pure","чистый"),("perfect","идеальный"),("high capacity","большой объём"),
 ("shampoo & treatment","шампунь и маска"),("shampoo","шампунь"),("treatment","маска-уход"),("re: repair","восстановление"),
 ("volume","объём"),("nuance","нюанс"),("edgy move","эффект «эджи»"),("arrange fix","фиксация укладки"),
 ("mini model","мини-формат"),("trial","пробник"),("1day","1 день"),("1da","1 день"),("night","ночной"),
 ("frizz care","уход за пушащимися волосами"),("sheer gloss","лёгкий блеск"),("rich moisture","насыщенное увлажнение"),
 ("rich moistur","насыщенное увлажнение"),("oil control","контроль жирности"),("edgy move","эффект «эджи»"),
 ("hair treatment","уход для волос"),("hair repair","восстановление волос"),("hair mask","маска для волос"),
 ("hair water","вода для укладки"),("hair","для волос"),("mask","маска"),("water","вода"),("milk","молочко"),
 ("cream","крем"),("wax","воск"),("spray","спрей"),("foam","пенка"),("lotion","лосьон"),("serum","сыворотка"),
 ("essence","эссенция"),("gel","гель"),("oil","масло"),("balm","бальзам"),("refill","рефилл"),("control","контроль"),
 ("care","уход"),("repair","восстановление"),("deep","глубокий"),("clear","очищение"),("moist","увлажнение"),
 ("acne","проблемная кожа"),("tight","плотная фиксация"),("flow","подвижная укладка"),("active","активная"),
 ("mat ","матовый "),("arrange","укладка"),("volume up","для объёма"),("long","для длины"),(" & "," и "),
 (" with "," с "),(" and "," и "),(" for "," для "),(" the "," "),("＆"," и "),("home",""),
]
# японский словарь (Cosme Station)
JP = [
 ("リンスインシャンプー","шампунь-кондиционер 2в1"),("スカルプケア薬用シャンプー詰替","лечебный шампунь для кожи головы, рефилл"),
 ("スカルプケア薬用シャンプー","лечебный шампунь для кожи головы"),("スカルプシャンプー","шампунь для кожи головы"),
 ("トニックリンスインシャンプー","тонизирующий шампунь-кондиционер 2в1"),
 ("シャンプー詰替","шампунь, рефилл"),("コンディショナー詰替","кондиционер, рефилл"),("ボディソープ詰替","гель для душа, рефилл"),
 ("リペアシャンプー","восстанавливающий шампунь"),("リペアコンディショナー","восстанавливающий кондиционер"),
 ("シャンプー","шампунь"),("コンディショナー","кондиционер"),("ボディソープ","гель для душа"),("ボディローション","лосьон для тела"),
 ("ボディミルク","молочко для тела"),("ボディクリーム","крем для тела"),("ボディパウダー","пудра для тела"),
 ("スタイリングウォーター","стайлинг-вода"),("トリートメント","маска-уход"),("ヘアクリーム","крем для волос"),
 ("ハンドクリーム","крем для рук"),("ハンド＆ネイルクリーム","крем для рук и ногтей"),("ハンドソープ詰替","мыло для рук, рефилл"),
 ("尿素ハンドクリーム","крем для рук с мочевиной"),("ハトムギハンドクリーム","крем для рук с коиксом"),
 ("洗顔フォーム","пенка для умывания"),("洗顔ジェル","гель для умывания"),("マッサージ洗顔ジェル","массажный гель для умывания"),
 ("Wクレンジング洗顔フォーム","пенка для умывания + демакияж"),("スキンクリーム","крем для лица"),("化粧水","тонер для лица"),
 ("スパミストウォーター","спа-мист"),("ミストローション","мист-лосьон"),("ミスト","мист"),("スプレー","спрей"),
 ("フェイシャルマスク","тканевая маска для лица"),("マスク","маска"),("あぶらとり紙","матирующие салфетки"),
 ("オードトワレ","туалетная вода"),("ルームフレグランス","аромат для дома"),("ミニオードトワレ","туалетная вода мини"),
 ("マスカラ","тушь для ресниц"),("ロング＆ボリュームアップマスカラ","тушь для длины и объёма"),
 ("馬油","лошадиное масло"),("椿油","масло камелии"),("柿渋","экстракт хурмы"),("ゴートミルク","козье молоко"),
 ("はちみつ","мёд"),("ハニー","мёд"),("モモ","персик"),("あろえ","алоэ"),("アロエ","алоэ"),("ティーツリー","чайное дерево"),
 ("プロポリス","прополис"),("モイスチャー","увлажнение"),("セラミド","керамиды"),("レチノール","ретинол"),
 ("ビタミンC","витамин C"),("酵素","ферменты"),("クール","охлаждающий"),("フローズンスプレー","охлаждающий спрей"),
 ("リフレッシュ","освежающий"),("ブラック","чёрный"),("シルク","шёлк"),("金","золото"),("炭","уголь"),
 ("ブラックミント","чёрная мята"),("グレープフルーツ","грейпфрут"),("さくらんぼの香り","аромат вишни"),
 ("ホワイトムスク","белый мускус"),("センシュアルムスク","чувственный мускус"),("アンバーオーシャン","янтарный океан"),
 ("シトラスウッド","цитрус и дерево"),("男性","мужской"),("詰替","рефилл"),("の香り","— аромат"),("入","шт"),
 ("シャボン","мыльный аромат"),("フローラル","цветочный"),("リッチ","насыщенный"),("スウィート","сладкий"),
 ("実花草爽","серия «дзикка сосо»"),("恵肌小町","серия «мэгихада»"),
 ("ジャスミンブーケ","жасминовый букет"),("ブルームオスマンサス","османтус в цвету"),("ブルームポワール","груша в цвету"),
 ("オスマンサス","османтус"),("ジャスミン","жасмин"),("フリージア","фрезия"),("リリー","лилия"),("イベリス","иберис"),
 ("エレガント","элегантный"),("ムスク","мускус"),("ローズ","роза"),("ロゼ","розе"),("ピーチマジック","персиковая магия"),
 ("チャーミングベリー","очаровательная ягода"),("オレンジハピネス","апельсиновое счастье"),("ピーチ","персик"),
 ("クリーンペタル","чистый лепесток"),("リッチブロッサム","насыщенный цвет"),("ピュアシャボン","чистое мыло"),
 ("シャンパン","шампанское"),("ミニョン","миньон"),("ルナ","луна"),("ヴェール","вуаль"),("サボン","мыло"),("ブラン","белый"),
 ("ブラックティー","чёрный чай"),("ホワイトティー","белый чай"),("カフェ","кафе"),("グレイスミモザ","грациозная мимоза"),
 ("クラシックペアー","классическая груша"),("メルティスウィート","тающая сладость"),("シアーブルーム","лёгкий цвет"),
 ("ルミナスアクア","сияющая вода"),("ギルティローズ","страстная роза"),("プレミアム","премиум"),("プチ","мини"),
 ("ラメ","с блёстками"),("POPNPON","POPNPON"),("ラウンド","круглая"),("ハート","сердце"),("マカロン","макарон"),
 ("チェリー","вишня"),("スウィートマカロンの香り","аромат сладкого макарона"),("ブルーム","в цвету"),("ポワール","груша"),
 ("スカルプケア","уход за кожей головы"),("薬用","лечебный"),("スパ","спа"),("CICA","CICA"),("SPA","спа"),
]
import re as _re
JP = [
 ("シャボンフローラル","мыльно-цветочный"),("エレガントリッチ","элегантный насыщенный"),("フローラルリッチ","цветочный насыщенный"),
 ("スウィートフリージア","сладкая фрезия"),("スウィートリリー","сладкая лилия"),("スウィートイベリス","сладкий иберис"),
 ("ブラウン","коричневый"),("ブラック","чёрный"),
] + JP
def strip_marks(s):
    s = _re.sub(r"[（(][^）)]*[）)]", "", s)        # (MM002), （…）
    s = _re.sub(r"【[^】]*】", "", s)                 # 【リニューアル】
    s = _re.sub(r"※.*$", "", s)                      # ※リニューアル …
    s = s.replace("リニューアル","")
    return " ".join(s.split())
def tr(name, table):
    s = strip_marks(name)
    for a,b in sorted(table, key=lambda x:-len(x[0])):
        s = re.sub(re.escape(a), b, s, flags=re.IGNORECASE)
    return clean(s)

wb_out = Workbook(); wb_out.remove(wb_out.active)
HDR = ["Наименование (рус)","Оригинал","Объём","JAN","Цена JPY","Себест. RUB","Мин.лот"]
def sheet(title):
    ws = wb_out.create_sheet(title[:31])
    ws.append(HDR)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="4472C4")
    return ws
def widths(ws, w):
    for i,x in enumerate(w,1): ws.column_dimensions[chr(64+i)].width = x

def load(f, sh=None):
    wb = openpyxl.load_workbook(UP+f, read_only=True, data_only=True)
    return wb[sh] if sh else wb.active

# 1) KOR
ws = sheet("KOR");
src = load("KOR_russian_price_2026-07-14.xlsx")
data = [[("" if c is None else str(c)).strip() for c in r] for r in src.iter_rows(values_only=True)]
for r in data:
    for base in (1,8):
        if len(r) <= base+5: continue
        jan,name,vol,lot,retail,cost = r[base:base+6]
        if jan.isdigit() and len(jan)>=12 and name:
            ws.append([tr(clean(name),EN), clean(name), vol, jan, num(cost), rub(cost), lot])
widths(ws,[46,40,10,16,10,12,8])

# 1b) KOR NEW
for r in list(load("KOR_new_russian_price.xlsx").iter_rows(values_only=True))[1:]:
    jan,name,price,lot = (str(x).strip() if x is not None else "" for x in r[:4])
    if jan.isdigit() and name:
        ws.append([tr(clean(name),EN), clean(name), "", jan, num(price), rub(price), lot])

# 2) MANDOM
ws = sheet("Mandom (LUCIDO-L, Bifesta)")
for r in list(load("Mandom_LucidoL_Bifesta_2026-07.xlsx").iter_rows(values_only=True))[3:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 11: continue
    jan,brand,name_en,vol,price = r[2],r[5],r[7],r[8],r[10]
    name = name_en or r[6]
    if jan.isdigit() and len(jan)>=12 and name:
        rus = tr(clean(name_en), EN) if name_en else clean(r[6])
        ws.append([clean(rus), clean(name), vol, jan, num(price), rub(price), ""])
widths(ws,[48,42,10,16,10,12,8])

# 3) COSME STATION
ws = sheet("Cosme Station")
series=""
for r in list(load("CosmeStation_2026.xlsx").iter_rows(values_only=True))[5:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 25: continue
    if r[0]: series = r[0]
    name,jan,vol,nouka = r[1],r[3],r[6],r[23]
    if jan.isdigit() and len(jan)>=12 and name and num(nouka):
        rus = tr(clean(name), JP)
        ser = tr(series, JP)
        full = f"{rus} [{ser}]" if ser and ser not in rus else rus
        ws.append([clean(full), clean(f"{series} {name}"), vol, jan, num(nouka), rub(nouka), ""])
widths(ws,[50,40,10,16,10,12,8])

# 4) DOSHISHA (уже рус)
ws = sheet("Doshisha")
for r in list(load("Doshisha.xlsx","Price").iter_rows(values_only=True))[6:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 8: continue
    jan,name_jp,name_ru,cap,price = r[1],r[3],r[4],r[5],r[7]
    if jan.isdigit() and len(jan)>=12 and (name_ru or name_jp) and num(price):
        ws.append([clean(name_ru or tr(name_jp,JP)), clean(name_jp), cap, jan, num(price), rub(price), ""])
widths(ws,[50,36,10,16,10,12,8])

# 5) FINE (уже рус)
ws = sheet("Fine (БАД)")
for r in list(load("Fine_supplements.xlsx").iter_rows(values_only=True))[7:]:
    r = [("" if c is None else str(c)).strip() for c in r]
    if len(r) < 8: continue
    maker,jan,name_jp,name_en,name_ru,ppc,price = r[1],r[2],r[3],r[4],r[5],r[6],r[7]
    if jan.isdigit() and len(jan)>=12 and (name_ru or name_en) and num(price):
        ws.append([clean(name_ru or name_en), clean(name_en or name_jp), "", jan, num(price), rub(price), ""])
widths(ws,[52,40,10,16,10,12,8])

# нормализация: имя — только японские счётчики; объём — ещё и g/ml
def nname(x):
    if x is None: return x
    s = str(x)
    for a,b in [("枚"," шт"),("入",""),("台"," шт"),("Sheets"," шт"),("sheets"," шт"),("Sheet"," шт")]:
        s = s.replace(a,b)
    return " ".join(s.split())
def nvol(x):
    if x is None: return x
    s = nname(x)
    for a,b in [("g"," г"),("ml"," мл"),("ｇ"," г"),("ｍｌ"," мл")]:
        s = s.replace(a,b)
    return " ".join(s.split())
for ws in wb_out.worksheets:
    for row in ws.iter_rows(min_row=2):
        row[0].value = nname(row[0].value)
        if row[2].value: row[2].value = nvol(row[2].value)

out = "outputs/japan_price_lists_RU_2026-07-27.xlsx"
wb_out.save(out)
print("Сохранено:", out)
for ws in wb_out.worksheets:
    print(f"  {ws.title}: {ws.max_row-1} позиций")
