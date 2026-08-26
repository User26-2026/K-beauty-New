"""Ставки Яндекс Маркета по категориям для каждого SKU юнитки.

Источник ставок: data/yandex_market/marketplace_services_rates_01072026.xlsx
(официальный файл Маркета «Ставки за услуги маркетплейса», действует с 01.07.2026),
срез по красоте — outputs/yandex_market/ym_rates_beauty_01072026.tsv.

Категория товара определяется по названию: правила ниже, первое совпадение выигрывает.
Правила намеренно простые и читаемые — их проверяют глазами по листу «Тарифы».
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
RATES_TSV = ROOT / "outputs" / "yandex_market" / "ym_rates_beauty_01072026.tsv"
PREFIX = "Товары для красоты > Косметика, парфюмерия и уход > "

# (регулярка по названию, путь категории Маркета). Порядок важен: сначала узкие правила.
RULES: list[tuple[str, str]] = [
    # уход за кожей вокруг глаз
    (r"eye patch|патч|eye cream|крем для глаз|eye serum|rolling eye",
     "Уход за лицом > Уход за кожей вокруг глаз"),
    # губы
    (r"lip sleeping|бальзам для губ|для губ", "Уход за лицом > Уход за кожей губ"),
    # макияж
    (r"concealer|консилер", "Макияж > Для лица > Корректоры и консилеры"),
    (r"glow base|база под макияж|makeup base|праймер",
     "Макияж > Для лица > Основа и фиксаторы для макияжа"),
    (r"foundation|тональн|bb cream|бб.?крем", "Макияж > Для лица > Тональные средства"),
    # волосы
    (r"shampoo\+balsam|шампунь.?\+.?бальзам|shampoo & conditioner|shampoo and conditioner",
     "Наборы > Косметические наборы для ухода за волосами"),
    (r"shampoo|шампун", "Уход за волосами > Шампуни"),
    (r"hair mask|маска для волос|hair treatment|для волос",
     "Уход за волосами > Маски и сыворотки"),
    # очищение лица
    (r"cleanser|cleansing foam|cleansing bar|пенка|гель для умывания|мыло для лица|"
     r"пена для умывания|bubble foam|gel cleanser",
     "Уход за лицом > Очищение и снятие макияжа > Средства для умывания"),
    (r"cleansing oil|cleansing balm|гидрофильн|для снятия макияжа|бальзам щербет",
     "Уход за лицом > Очищение и снятие макияжа > Средства для снятия макияжа"),
    # скрабы и пилинги
    (r"scrub|скраб|пилинг|peeling", "Уход за лицом > Скрабы и пилинги > Скрабы для лица"),
    # маски для лица
    (r"mask pack|face mask|тканев|маска для лица|маски|sleeping pack|ночная маска",
     "Уход за лицом > Маски > Маски для лица"),
    # все остальное по уходу за лицом: кремы, сыворотки, тонеры, эссенции, ампулы, пэды
    (r"cream|крем|serum|сыворотк|ampoule|ампул|essence|эссенц|toner|тонер|тоник|"
     r"lotion|лосьон|pad|пэд|booster|бустер|shot|retinol|ретинол",
     "Уход за лицом > Кремы и сыворотки > Средства для ухода за лицом"),
]

FALLBACK = "Уход за лицом > Кремы и сыворотки > Средства для ухода за лицом"


def load_rates() -> dict[str, tuple[float, float]]:
    """{короткий путь категории: (ставка FBY, ставка FBS)} в долях."""
    rates: dict[str, tuple[float, float]] = {}
    with RATES_TSV.open(encoding="utf-8") as fh:
        for row in csv.DictReader(fh, delimiter="\t"):
            path = row["Путь"]
            if not path.startswith(PREFIX):
                continue
            rates.setdefault(
                path[len(PREFIX):], (float(row["FBY %"]) / 100, float(row["FBS %"]) / 100)
            )
    return rates


def category_for(name: str) -> str:
    low = name.lower()
    for pattern, path in RULES:
        if re.search(pattern, low):
            return path
    return FALLBACK


def rate_for(name: str, rates: dict[str, tuple[float, float]], model: str) -> tuple[str, float]:
    """(путь категории, ставка) для схемы 'FBY' (наш FBO) или 'FBS'."""
    path = category_for(name)
    fby, fbs = rates[path]
    return path, (fby if model == "FBY" else fbs)


if __name__ == "__main__":
    import openpyxl

    rates = load_rates()
    wb = openpyxl.load_workbook(
        ROOT / "outputs" / "ym" / "ym_unitka_fbo_fbs_2026-08-17.xlsx", data_only=True
    )
    ws = wb["Юнитка FBO"]
    counts: dict[str, int] = {}
    for row in range(13, 153):
        name = ws.cell(row, 1).value
        if not isinstance(name, str) or ws.cell(row, 4).value is None:
            continue
        path, fby = rate_for(name, rates, "FBY")
        _, fbs = rate_for(name, rates, "FBS")
        counts[path] = counts.get(path, 0) + 1
        print(f"{fby:.3f} / {fbs:.3f} | {path[:55]:<55} | {name[:60]}")
    print("\nпо категориям:")
    for path, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"  {n:>3}  {path}  {rates[path][0]:.0%} / {rates[path][1]:.0%}")
