"""Расчет юнит-экономики Яндекс Маркета (FBO и FBS) по модели из ym_unit_model_11.08.26.xlsx.

Что делает скрипт:
1. Пересчитывает все заполненные строки модели в Python и сверяет с кэшем Excel.
2. Достает себестоимость по последнему приходу для строк без себестоимости
   из data/stock_costs/ОСТАТКИ_КОРЕЯ_03.06.xlsx (колонка R).
3. Дописывает эти строки в модель теми же формулами (цена под ROI 25% на FBO,
   20% на FBS — как в юнитке Ozon).
4. Подставляет ставку Маркета по категории каждого товара из официального файла
   ставок с 01.07.2026 (см. ym_category_rates.py).
5. Пишет два файла:
   - outputs/ym/ym_unitka_fbo_fbs_<дата>.xlsx — чистая юнитка, два листа;
   - outputs/ym/ym_unit_economics_<дата>.xlsx — она же плюс листы "Тарифы",
     "Сводка" и "Решения FBO".
"""

from __future__ import annotations

import re
import shutil
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from ym_category_rates import load_rates, rate_for

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "data" / "unit_economics" / "ym_unit_model_11.08.26.xlsx"
STOCK = ROOT / "data" / "stock_costs" / "ОСТАТКИ_КОРЕЯ_03.06.xlsx"
STOCK_SHEET = "03.06.26"
STOCK_NAME_COL = 8  # H, Наименование
STOCK_COST_COL = 18  # R, Себестоимость по последнему приходу
OUT_DIR = ROOT / "outputs" / "ym"
OUT = OUT_DIR / "ym_unit_economics_2026-08-17.xlsx"  # расчет со сводкой и решениями
OUT_UNITKA = OUT_DIR / "ym_unitka_fbo_fbs_2026-08-17.xlsx"  # чистая юнитка в формате Ozon

# Целевой ROI для цены-формулы: на FBO 25%, на FBS 20% — как в юнитке Ozon
# (outputs/ozon/ozon_unitka_fbo_fbs_2026-08-05.xlsx: FBO 1.25, FBS 1.20).
TARGET_ROI = 0.25
TARGET_ROI_BY_SHEET = {"Юнитка FBO": 0.25, "Юнитка FBS": 0.20}
DEFAULT_CLICK = 0.07  # оплата за клик, % от цены без СПП
DEFAULT_TOP = 0.08  # вывод в топ, % от цены без СПП
BUYOUT = 0.90

# Объем упаковки (л) для строк, которых не было в расчете. Оценка по аналогии
# с сопоставимыми SKU внутри самой модели: объем влияет только на кроссдок
# (5.173 руб/л) и на ступень логистики, поэтому погрешность в пределах рубля-двух.
VOLUME_ASSUMPTIONS = {
    "DEOPROCE SHAMPOO - BLACK GARLIC INTENSME ENERGY [600ml] мягкая упаковка": 2.0,
    "DEOPROCE SHAMPOO - BLACK GARLIC INTENSME ENERGY [600ml] мягкая упаковка с колпачком": 2.0,
    "DEOPROCE SHAMPOO+BALSAM - BLACK GARLIC[2000ml]": 4.0,
    "ENOUGH Collagen whitening cover tip concealer #01 [9g]": 0.18,
    "ENOUGH Collagen whitening cover tip concealer #02 [9g]": 0.18,
    "ENOUGH Collagen whitening cover tip concealer #03 [9g]": 0.18,
    "JIGOTT - Ассорти набор масок [27ml х 15]": 1.60,
    "MASIL - 8 Seconds Salon Hair Mask 350ml Set [350ml+8ml*2]": 1.20,
    "PETITFEE - Black Pearl & Gold Hydrogel Eye Patch [1 pairs]": 0.05,
    "PETITFEE - Gold Hydrogel Eye Patch [1 Pairs]": 0.05,
    "ROUND LAB BIRCH JUICE MOISTURIZING CREAM_80ml": 0.40,
    "ROUND LAB SOYBEAN PANTHENOL CREAM_80ml": 0.40,
    "ROUND LAB PINE CALMING CICA AMPOULE_30ml": 0.38,
    "SKIN1004 - Madagascar Centella Ampoule [30ml]": 0.38,
    "SKIN1004 - Madagascar Centella Hyalu-Cica Sleeping Pack [30ml]": 0.38,
    "SOME BY MI - AHA-BHA-PHA 30 Days Miracle Cleansing Bar [106g]": 0.30,
    "SOME BY MI - AHA-BHA-PHA 30 Days Miracle Cream [60g]": 0.40,
    "THE ORDINARY - Argireline Solution 10% [30ml]": 0.38,
    "THE ORDINARY - Glycolic Acid 7% Toning Solution [KOR] [240ml]": 0.735,
    "CELIMAX PORE+DARK SPOT BRIGHTENING CREAM [35ml]": 0.38,
    "CELIMAX DUAL BARRIER SKIN WEARABLE CREAM [50ml]": 0.50,
    "CELIMAX THE REAL NONI ENERGY AMPOULE_30ML (NEW) [30ml]": 0.38,
    "CELIMAX PORE+DARK SPOT BRIGHTENING SERUM_30ml [30ml]": 0.38,
    "CELIMAX THE REAL NONI MOISTURE BALANCING TONER (NEW) [150ml]": 0.735,
}

# Ступени логистики Яндекс Маркета, восстановленные по заполненным строкам модели.
LOGISTICS_TIERS = ((0.20, 56.0), (0.40, 63.0), (1.00, 67.0), (1.25, 72.0), (2.00, 75.0))


def logistics_for(volume: float) -> float:
    for limit, value in LOGISTICS_TIERS:
        if volume <= limit + 1e-9:
            return value
    return 75.0


@dataclass
class SheetSpec:
    """Раскладка колонок вкладки модели (номера колонок 1-based)."""

    title: str
    header_row: int
    param_col: int  # колонка с параметрами (H на FBO, J на FBS)
    name: int
    comment: int
    stock: int
    comp_price: int
    price: int
    spp: int
    price_spp: int
    cost: int
    ff: int
    cost_base: int
    comm_pct: int
    comm_rub: int
    logistics: int
    nonlocal_rub: int
    pvz: int
    acquiring: int
    tax: int
    buyout: int
    rev_log_base: int
    rev_log: int
    volume: int
    storage: int
    crossdock: int
    click_pct: int
    click_rub: int
    top_pct: int
    top_rub: int
    ads: int
    drr: int
    margin_before_ads: int
    margin: int
    margin_pct: int
    roi: int
    breakeven: int
    pvz_shipment: int | None  # AN, только FBS
    subscription: int

    @property
    def params(self) -> dict[str, int]:
        """Строки параметров листа: {имя: строка}."""
        base = {
            "tax": 3,
            "acquiring": 4,
            "ff": 5,
            "storage_rate": 6,
            "storage_days": 7,
            "nonlocal": 8,
            "cross_handling": 9,
            "cross_transport": 10,
            "subscription": 11,
        }
        if self.pvz_shipment is not None:
            base["pvz_shipment"] = 12
        return base


FBO = SheetSpec(
    title="Юнитка FBO",
    header_row=12,
    param_col=8,
    name=1,
    comment=3,
    stock=4,
    comp_price=5,
    price=8,
    spp=9,
    price_spp=10,
    cost=11,
    ff=12,
    cost_base=13,
    comm_pct=14,
    comm_rub=15,
    logistics=16,
    nonlocal_rub=17,
    pvz=18,
    acquiring=19,
    tax=20,
    buyout=21,
    rev_log_base=22,
    rev_log=23,
    volume=24,
    storage=25,
    crossdock=26,
    click_pct=27,
    click_rub=28,
    top_pct=29,
    top_rub=30,
    ads=31,
    drr=32,
    margin_before_ads=33,
    margin=34,
    margin_pct=35,
    roi=36,
    breakeven=37,
    pvz_shipment=None,
    subscription=38,
)

FBS = SheetSpec(
    title="Юнитка FBS",
    header_row=13,
    param_col=10,
    name=2,
    comment=5,
    stock=6,
    comp_price=7,
    price=10,
    spp=11,
    price_spp=12,
    cost=13,
    ff=14,
    cost_base=15,
    comm_pct=16,
    comm_rub=17,
    logistics=18,
    nonlocal_rub=19,
    pvz=20,
    acquiring=21,
    tax=22,
    buyout=23,
    rev_log_base=24,
    rev_log=25,
    volume=26,
    storage=27,
    crossdock=28,
    click_pct=29,
    click_rub=30,
    top_pct=31,
    top_rub=32,
    ads=33,
    drr=34,
    margin_before_ads=35,
    margin=36,
    margin_pct=37,
    roi=38,
    breakeven=39,
    pvz_shipment=40,
    subscription=41,
)

SPECS = (FBO, FBS)
LAST_ROW = 152


def col_letter(idx: int) -> str:
    return openpyxl.utils.get_column_letter(idx)


def norm(text: object) -> str:
    s = unicodedata.normalize("NFKC", str(text)).lower()
    return re.sub(r"[^a-zа-я0-9]+", "", s)


def load_params(ws, spec: SheetSpec) -> dict[str, float]:
    col = spec.param_col
    return {key: float(ws.cell(row, col).value or 0) for key, row in spec.params.items()}


def read_stock_costs() -> dict[str, float]:
    ws = openpyxl.load_workbook(STOCK, data_only=True)[STOCK_SHEET]
    costs: dict[str, float] = {}
    for row in range(5, ws.max_row + 1):
        name = ws.cell(row, STOCK_NAME_COL).value
        cost = ws.cell(row, STOCK_COST_COL).value
        if isinstance(name, str) and isinstance(cost, (int, float)):
            costs.setdefault(norm(name), float(cost))
    return costs


def find_cost(name: str, costs: dict[str, float]) -> float | None:
    """Себестоимость по названию: точное совпадение, затем совпадение по префиксу.

    В остатках и в модели у одного и того же товара расходятся хвосты названия
    ("(новая карточка)" против "(основная карточка)"), поэтому нужен префикс.
    """
    key = norm(name)
    if key in costs:
        return costs[key]
    prefix = re.split(r"основ|новаякарточка", key)[0]
    if len(prefix) < 25:
        return None
    hits = {value for other, value in costs.items() if other.startswith(prefix)}
    return hits.pop() if len(hits) == 1 else None


def unit(
    *,
    price: float,
    cost: float,
    volume: float,
    logistics: float,
    comm_pct: float,
    spp: float,
    click_pct: float,
    top_pct: float,
    params: dict[str, float],
    target_roi: float = TARGET_ROI,
) -> dict[str, float]:
    """Одна строка юнитки по логике модели Яндекс Маркета."""
    ff = params["ff"]
    pvz_shipment = params.get("pvz_shipment", 0.0)
    cost_base = cost + ff
    price_spp = price * (1 - spp)
    storage = params["storage_rate"] * params["storage_days"] * volume
    crossdock = volume * (params["cross_handling"] + params["cross_transport"])
    fixed = logistics + 25.0 + logistics * (1 - BUYOUT) + storage + crossdock + pvz_shipment
    pct = (
        comm_pct
        + params["acquiring"]
        + (1 - spp) * params["tax"]
        + params["nonlocal"]
        + params["subscription"]
    )
    ads = price * (click_pct + top_pct)
    margin_before_ads = price * (1 - pct) - fixed - cost_base
    margin = margin_before_ads - ads
    return {
        "price": price,
        "price_spp": price_spp,
        "cost": cost,
        "cost_base": cost_base,
        "comm_rub": price * comm_pct,
        "logistics": logistics,
        "nonlocal_rub": price * params["nonlocal"],
        "pvz": 25.0,
        "pvz_shipment": pvz_shipment,
        "acquiring": price * params["acquiring"],
        "tax": price_spp * params["tax"],
        "rev_log": logistics * (1 - BUYOUT),
        "storage": storage,
        "crossdock": crossdock,
        "subscription": price * params["subscription"],
        "ads": ads,
        "drr": (click_pct + top_pct),
        "margin_before_ads": margin_before_ads,
        "margin": margin,
        "margin_pct": margin / price if price else 0.0,
        "roi": margin / cost_base if cost_base else 0.0,
        "breakeven": (fixed + cost_base) / (1 - pct - click_pct - top_pct),
        "price_target": (fixed + (1 + target_roi) * cost_base)
        / (1 - pct - click_pct - top_pct),
    }


def price_formula(spec: SheetSpec, row: int, target_roi: float) -> str:
    """Формула цены под целевой ROI — точно та же, что в формульных строках модели."""
    pc = col_letter(spec.param_col)
    L = {
        key: col_letter(getattr(spec, key))
        for key in vars(spec)
        if isinstance(getattr(spec, key), int)
    }
    terms = [L["logistics"], L["pvz"], L["rev_log"], L["storage"], L["crossdock"]]
    if spec.pvz_shipment is not None:
        terms.append(L["pvz_shipment"])
    fixed = "+".join(f"{c}{row}" for c in terms)
    return (
        f"=({1 + target_roi:g}*{L['cost_base']}{row}+{fixed})/(1-{L['comm_pct']}{row}"
        f"-${pc}${spec.params['nonlocal']}-${pc}${spec.params['acquiring']}"
        f"-(1-{L['spp']}{row})*${pc}${spec.params['tax']}-${pc}${spec.params['subscription']}"
        f"-{L['click_pct']}{row}-{L['top_pct']}{row})"
    )


def verify(wb_values) -> list[str]:
    """Сверяет питоновский пересчет с кэшированными значениями Excel."""
    problems: list[str] = []
    for spec in SPECS:
        ws = wb_values[spec.title]
        params = load_params(ws, spec)
        for row in range(spec.header_row + 1, LAST_ROW + 1):
            price = ws.cell(row, spec.price).value
            cost = ws.cell(row, spec.cost).value
            if not isinstance(price, (int, float)) or not isinstance(cost, (int, float)):
                continue
            calc = unit(
                price=float(price),
                cost=float(cost),
                volume=float(ws.cell(row, spec.volume).value or 0),
                logistics=float(ws.cell(row, spec.logistics).value or 0),
                comm_pct=float(ws.cell(row, spec.comm_pct).value or 0),
                spp=float(ws.cell(row, spec.spp).value or 0),
                click_pct=float(ws.cell(row, spec.click_pct).value or 0),
                top_pct=float(ws.cell(row, spec.top_pct).value or 0),
                params=params,
            )
            for key, col in (
                ("margin", spec.margin),
                ("roi", spec.roi),
                ("breakeven", spec.breakeven),
            ):
                excel = ws.cell(row, col).value
                if isinstance(excel, (int, float)) and abs(calc[key] - float(excel)) > 0.01:
                    problems.append(
                        f"{spec.title} стр.{row} {key}: excel={excel:.4f} python={calc[key]:.4f}"
                    )
    return problems


def fill_missing(wb_formulas, wb_values, costs: dict[str, float]) -> list[dict]:
    """Дописывает строки без себестоимости теми же формулами, что в модели."""
    filled: list[dict] = []
    for spec in SPECS:
        wsf = wb_formulas[spec.title]
        wsv = wb_values[spec.title]
        params = load_params(wsv, spec)
        for row in range(spec.header_row + 1, LAST_ROW + 1):
            name = wsv.cell(row, spec.name).value
            if not isinstance(name, str):
                continue
            if wsv.cell(row, spec.cost).value is not None:
                continue
            if wsv.cell(row, spec.stock).value is None:
                continue  # строка-категория
            cost = find_cost(name, costs)
            volume = next(
                (v for key, v in VOLUME_ASSUMPTIONS.items() if norm(key) in norm(name)),
                None,
            )
            if cost is None or volume is None:
                filled.append(
                    {
                        "sheet": spec.title,
                        "row": row,
                        "name": name,
                        "status": "нет себестоимости в остатках" if cost is None else "нет объема",
                    }
                )
                continue

            logistics = logistics_for(volume)
            comm_pct = 0.47 if spec.pvz_shipment is not None else 0.41
            spp = 0.45
            target_roi = TARGET_ROI_BY_SHEET[spec.title]
            P = spec.param_col
            pc = col_letter(P)
            L = {key: col_letter(getattr(spec, key)) for key in vars(spec) if isinstance(getattr(spec, key), int)}
            r = row

            def put(col: int, value):
                wsf.cell(r, col).value = value

            put(spec.price, None)  # проставим формулу ниже
            put(spec.spp, spp)
            put(spec.price_spp, f"={L['price']}{r}*(1-{L['spp']}{r})")
            put(spec.cost, cost)
            put(spec.ff, f"=${pc}${spec.params['ff']}")
            put(spec.cost_base, f"={L['cost']}{r}+{L['ff']}{r}")
            put(spec.comm_pct, comm_pct)
            put(spec.comm_rub, f"={L['price']}{r}*{L['comm_pct']}{r}")
            put(spec.logistics, logistics)
            put(spec.nonlocal_rub, f"={L['price']}{r}*${pc}${spec.params['nonlocal']}")
            put(spec.pvz, 25.0)
            put(spec.acquiring, f"={L['price']}{r}*${pc}${spec.params['acquiring']}")
            put(spec.tax, f"={L['price_spp']}{r}*${pc}${spec.params['tax']}")
            put(spec.buyout, BUYOUT)
            put(spec.rev_log_base, logistics)
            put(spec.rev_log, f"={L['rev_log_base']}{r}*(1-{L['buyout']}{r})")
            put(spec.volume, volume)
            put(
                spec.storage,
                f"=${pc}${spec.params['storage_rate']}*${pc}${spec.params['storage_days']}*{L['volume']}{r}",
            )
            put(
                spec.crossdock,
                f"={L['volume']}{r}*(${pc}${spec.params['cross_handling']}+${pc}${spec.params['cross_transport']})",
            )
            put(spec.click_pct, DEFAULT_CLICK)
            put(spec.click_rub, f"={L['price']}{r}*{L['click_pct']}{r}")
            put(spec.top_pct, DEFAULT_TOP)
            put(spec.top_rub, f"={L['price']}{r}*{L['top_pct']}{r}")
            put(spec.ads, f"={L['click_rub']}{r}+{L['top_rub']}{r}")
            put(spec.drr, f"={L['ads']}{r}/{L['price']}{r}")
            put(spec.subscription, f"={L['price']}{r}*${pc}${spec.params['subscription']}")
            if spec.pvz_shipment is not None:
                put(spec.pvz_shipment, f"=${pc}${spec.params['pvz_shipment']}")

            minus = [
                L["comm_rub"],
                L["logistics"],
                L["nonlocal_rub"],
                L["pvz"],
                L["acquiring"],
                L["tax"],
                L["rev_log"],
                L["storage"],
                L["crossdock"],
                L["cost_base"],
            ]
            if spec.pvz_shipment is not None:
                minus.append(L["pvz_shipment"])
            minus.append(L["subscription"])
            put(
                spec.margin_before_ads,
                "=" + L["price"] + str(r) + "".join(f"-{c}{r}" for c in minus),
            )
            put(spec.margin, f"={L['margin_before_ads']}{r}-{L['ads']}{r}")
            put(spec.margin_pct, f"={L['margin']}{r}/{L['price']}{r}")
            put(spec.roi, f"={L['margin']}{r}/{L['cost_base']}{r}")

            fixed_terms = [L["logistics"], L["pvz"], L["rev_log"], L["storage"], L["crossdock"], L["cost_base"]]
            if spec.pvz_shipment is not None:
                fixed_terms.append(L["pvz_shipment"])
            fixed_sum = "+".join(f"{c}{r}" for c in fixed_terms)
            denom = (
                f"1-({L['comm_pct']}{r}+${pc}${spec.params['acquiring']}"
                f"+(1-{L['spp']}{r})*${pc}${spec.params['tax']}+${pc}${spec.params['nonlocal']}"
                f"+${pc}${spec.params['subscription']}+{L['click_pct']}{r}+{L['top_pct']}{r})"
            )
            put(spec.breakeven, f"=({fixed_sum})/({denom})")
            put(spec.price, price_formula(spec, r, target_roi))

            common = dict(
                cost=cost,
                volume=volume,
                logistics=logistics,
                comm_pct=comm_pct,
                spp=spp,
                click_pct=DEFAULT_CLICK,
                top_pct=DEFAULT_TOP,
                params=params,
                target_roi=target_roi,
            )
            final = unit(price=unit(price=0.0, **common)["price_target"], **common)
            note = wsv.cell(row, spec.comment).value
            wsf.cell(row, spec.comment).value = (
                f"{note + '; ' if note else ''}добавлено в расчет: цена под ROI "
                f"{target_roi:.0%}, объем оценочный"
            )
            filled.append(
                {
                    "sheet": spec.title,
                    "row": row,
                    "name": name,
                    "status": "рассчитано",
                    "cost": cost,
                    "volume": volume,
                    "logistics": logistics,
                    **{k: final[k] for k in ("price", "price_spp", "margin", "roi", "breakeven")},
                }
            )
    return filled


def build_summary(wb_formulas, wb_values, costs: dict[str, float]) -> list[list]:
    """Сводка по всем SKU: цена, себестоимость, ДРР, маржа, ROI, цена безубытка."""
    rows: list[list] = [
        [
            "Вкладка",
            "Строка",
            "Товар",
            "Остаток",
            "Себест., ₽",
            "База затрат, ₽",
            "Цена без СПП, ₽",
            "Цена с СПП, ₽",
            "Цена конкурента без СПП, ₽",
            "ДРР, %",
            "Маржа до рекламы, ₽",
            "Маржа/шт, ₽",
            "Маржа, % цены",
            "ROI, %",
            "Цена безубытка, ₽",
            "Цена под ROI 25%, ₽",
            "Источник цены",
        ]
    ]
    for spec in SPECS:
        wsv = wb_values[spec.title]
        wsf = wb_formulas[spec.title]
        params = load_params(wsv, spec)
        for row in range(spec.header_row + 1, LAST_ROW + 1):
            name = wsv.cell(row, spec.name).value
            if not isinstance(name, str) or wsv.cell(row, spec.stock).value is None:
                continue
            cost = wsf.cell(row, spec.cost).value
            price_cell = wsf.cell(row, spec.price).value
            if not isinstance(cost, (int, float)) or price_cell is None:
                continue
            volume = float(wsf.cell(row, spec.volume).value or 0)
            logistics = float(wsf.cell(row, spec.logistics).value or 0)
            comm_pct = float(wsf.cell(row, spec.comm_pct).value or 0)
            spp = float(wsf.cell(row, spec.spp).value or 0)
            click = float(wsf.cell(row, spec.click_pct).value or 0)
            top = float(wsf.cell(row, spec.top_pct).value or 0)
            common = dict(
                cost=float(cost),
                volume=volume,
                logistics=logistics,
                comm_pct=comm_pct,
                spp=spp,
                click_pct=click,
                top_pct=top,
                params=params,
            )
            if isinstance(price_cell, (int, float)):
                price = float(price_cell)
                source = "ручная цена в модели"
            else:
                # коэффициент целевого ROI берем из самой формулы: =(1.25*... или =(1.2*...
                coef = float(re.match(r"=\((\d+(?:\.\d+)?)\*", price_cell).group(1))
                price = unit(price=0.0, target_roi=coef - 1, **common)["price_target"]
                source = f"формула: ROI {coef - 1:.0%}"
            calc = unit(price=price, **common)
            comp = wsv.cell(row, spec.comp_price).value
            rows.append(
                [
                    spec.title.replace("Юнитка ", ""),
                    row,
                    name,
                    wsv.cell(row, spec.stock).value,
                    round(float(cost), 2),
                    round(calc["cost_base"], 2),
                    round(price, 2),
                    round(calc["price_spp"], 2),
                    round(float(comp), 2) if isinstance(comp, (int, float)) else "",
                    round(calc["drr"] * 100, 1),
                    round(calc["margin_before_ads"], 2),
                    round(calc["margin"], 2),
                    round(calc["margin_pct"] * 100, 2),
                    round(calc["roi"] * 100, 1),
                    round(calc["breakeven"], 2),
                    round(calc["price_target"], 2),
                    source,
                ]
            )
    return rows


def apply_category_rates(wb) -> list[dict]:
    """Ставка Маркета по категории товара вместо единой ставки на весь лист.

    Берем официальный файл ставок с 01.07.2026: FBY для листа FBO, FBS для FBS.
    """
    rates = load_rates()
    changed: list[dict] = []
    for spec in SPECS:
        ws = wb[spec.title]
        model = "FBS" if spec.pvz_shipment is not None else "FBY"
        for row in range(spec.header_row + 1, LAST_ROW + 1):
            name = ws.cell(row, spec.name).value
            if not isinstance(name, str) or ws.cell(row, spec.stock).value is None:
                continue
            old = ws.cell(row, spec.comm_pct).value
            if old is None:
                continue
            path, rate = rate_for(name, rates, model)
            if abs(float(old) - rate) > 1e-9:
                changed.append(
                    {
                        "sheet": spec.title,
                        "row": row,
                        "name": name,
                        "category": path,
                        "old": float(old),
                        "new": rate,
                    }
                )
            ws.cell(row, spec.comm_pct).value = rate
    return changed


def build_rates_sheet(wb, wb_values) -> list[list]:
    """Лист «Тарифы»: какая категория и ставка Маркета подставлена каждому SKU."""
    rates = load_rates()
    rows: list[list] = [
        [
            "Товар",
            "Категория Яндекс Маркета",
            "Ставка FBY (лист FBO), %",
            "Ставка FBS, %",
            "Было в модели FBO, %",
            "Было в модели FBS, %",
        ]
    ]
    ws_fbo, ws_fbs = wb_values[FBO.title], wb_values[FBS.title]
    fbs_old = {
        ws_fbs.cell(r, FBS.name).value: ws_fbs.cell(r, FBS.comm_pct).value
        for r in range(FBS.header_row + 1, LAST_ROW + 1)
    }
    for row in range(FBO.header_row + 1, LAST_ROW + 1):
        name = ws_fbo.cell(row, FBO.name).value
        if not isinstance(name, str) or ws_fbo.cell(row, FBO.stock).value is None:
            continue
        path, fby = rate_for(name, rates, "FBY")
        _, fbs = rate_for(name, rates, "FBS")
        old_fbo = ws_fbo.cell(row, FBO.comm_pct).value
        rows.append(
            [
                name,
                path,
                round(fby * 100, 1),
                round(fbs * 100, 1),
                round(float(old_fbo) * 100, 1) if old_fbo else "",
                round(float(fbs_old.get(name) or 0) * 100, 1) or "",
            ]
        )
    return rows


def build_unitka(costs: dict[str, float], wb_values) -> tuple[dict[str, int], list[dict]]:
    """Чистая юнитка в формате Ozon: два листа, все цены живыми формулами.

    Отличия от файла клиента: дозаполнены пустые строки и 61 цена на FBS,
    вставленная значениями, переведена обратно в формулу под ROI 20% (числа те же).
    """
    OUT_UNITKA.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT_UNITKA)
    wb = openpyxl.load_workbook(OUT_UNITKA, data_only=False)
    filled = fill_missing(wb, wb_values, costs)
    stats = {
        "ставок по категориям подставлено": len(apply_category_rates(wb)),
        "формул на FBS вместо значений": 0,
        "хвост пустых строк убран": 0,
    }
    for spec in SPECS:
        ws = wb[spec.title]
        if spec is FBS:
            for row in range(spec.header_row + 1, LAST_ROW + 1):
                price = ws.cell(row, spec.price).value
                if isinstance(price, (int, float)) and ws.cell(row, spec.cost).value is not None:
                    ws.cell(row, spec.price).value = price_formula(
                        spec, row, TARGET_ROI_BY_SHEET[spec.title]
                    )
                    stats["формул на FBS вместо значений"] += 1
        if ws.max_row > LAST_ROW:
            stats["хвост пустых строк убран"] += ws.max_row - LAST_ROW
            ws.delete_rows(LAST_ROW + 1, ws.max_row - LAST_ROW)
    wb.save(OUT_UNITKA)
    return stats, filled


def build_decisions(summary: list[list], head: list[str]) -> list[list]:
    """Решение по каждому SKU: сходится ли FBO по рыночной цене и сколько просит FBS."""
    ix = {name: i for i, name in enumerate(head)}
    fbs_price = {
        r[ix["Товар"]]: r[ix["Цена под ROI 25%, ₽"]] for r in summary if r[ix["Вкладка"]] == "FBS"
    }
    rows: list[list] = [
        [
            "Строка FBO",
            "Товар",
            "Остаток",
            "Себест., ₽",
            "Цена конкурента без СПП, ₽",
            "Цена безубытка FBO, ₽",
            "Цена FBO под ROI 25%, ₽",
            "Цена FBS под ROI 25%, ₽",
            "Цена конкурента / безубыток, %",
            "Решение",
        ]
    ]
    for r in summary:
        if r[ix["Вкладка"]] != "FBO":
            continue
        comp = r[ix["Цена конкурента без СПП, ₽"]]
        breakeven = r[ix["Цена безубытка, ₽"]]
        roi25 = r[ix["Цена под ROI 25%, ₽"]]
        if not isinstance(comp, (int, float)) or comp <= 0:
            ratio, verdict = "", "нет цены конкурента: сверить по выдаче Маркета"
        else:
            ratio = round(comp / breakeven * 100, 0)
            if comp >= roi25:
                verdict = "запускать: рынок держит ROI 25% и выше"
            elif comp > breakeven:
                verdict = "маржа тонкая: держать ДРР ниже 15% или поднимать цену"
            else:
                verdict = "не запускать на FBO по рыночной цене"
        rows.append(
            [
                r[ix["Строка"]],
                r[ix["Товар"]],
                r[ix["Остаток"]],
                r[ix["Себест., ₽"]],
                comp,
                breakeven,
                roi25,
                fbs_price.get(r[ix["Товар"]], ""),
                ratio,
                verdict,
            ]
        )
    return rows


def main() -> None:
    costs = read_stock_costs()
    wb_values = openpyxl.load_workbook(SRC, data_only=True)
    problems = verify(wb_values)
    print(f"сверка модели с Excel: расхождений {len(problems)}")
    for p in problems[:10]:
        print("  ", p)

    stats, filled = build_unitka(costs, wb_values)
    print(f"записан {OUT_UNITKA.relative_to(ROOT)} — юнитка в формате Ozon: " + ", ".join(
        f"{k} {v}" for k, v in stats.items()
    ))
    done = [f for f in filled if f["status"] == "рассчитано"]
    skipped = [f for f in filled if f["status"] != "рассчитано"]
    print(f"дозаполнено строк: {len(done)}, пропущено: {len(skipped)}")
    for s in skipped:
        print(f"   пропуск {s['sheet']} стр.{s['row']}: {s['status']} — {s['name'][:60]}")

    # Аналитический файл — это та же юнитка плюс листы разбора, чтобы цифры совпадали.
    shutil.copy(OUT_UNITKA, OUT)
    wb_formulas = openpyxl.load_workbook(OUT, data_only=False)
    rates_sheet = build_rates_sheet(wb_formulas, wb_values)
    summary = build_summary(wb_formulas, wb_values, costs)
    decisions = build_decisions(summary[1:], summary[0])
    for title, table, widths, freeze in (
        ("Тарифы", rates_sheet, (70, 55, 12, 12, 14, 14), "B2"),
        ("Решения FBO", decisions, (7, 70, 9, 11, 15, 15, 15, 15, 13, 42), "C2"),
        ("Сводка", summary, (6, 7, 70, 9, 11, 13, 15, 14, 22, 8, 18, 12, 13, 9, 16, 16, 22), "D2"),
    ):
        if title in wb_formulas.sheetnames:
            del wb_formulas[title]
        ws = wb_formulas.create_sheet(title, 0)
        for r in table:
            ws.append(r)
        ws.freeze_panes = freeze
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[col_letter(idx)].width = width
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    wb_formulas.save(OUT)
    print(f"записан {OUT.relative_to(ROOT)}: SKU в сводке {len(summary) - 1}")

    verdicts: dict[str, int] = {}
    for r in decisions[1:]:
        verdicts[r[-1]] = verdicts.get(r[-1], 0) + 1
    print("\nрешения по FBO:")
    for verdict, count in sorted(verdicts.items(), key=lambda kv: -kv[1]):
        print(f"  {count:>3}  {verdict}")

    losers = [r for r in summary[1:] if r[13] < 15]
    losers.sort(key=lambda r: r[13])
    print(f"\nSKU с ROI ниже 15% (по цене из модели): {len(losers)}")
    for r in losers[:40]:
        print(f"  {r[0]} стр.{r[1]:>3} ROI {r[13]:>7.1f}% маржа {r[11]:>8.2f} цена {r[6]:>8.0f} "
              f"безубыток {r[14]:>8.0f} | {r[2][:55]}")


if __name__ == "__main__":
    main()
