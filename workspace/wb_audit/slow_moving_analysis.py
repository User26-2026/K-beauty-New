"""
Анализ неходовых дорогих товаров на WB по оборачиваемости.

Логика:
- Остатки берем из свежего файла остатков (колонки 'Артикул WB',
  'Всего находится на складах', 'Себестоимость', 'Продажная цена').
- Скорость продаж считаем по data/wb_api/sales.json за последние 90 дней,
  сопоставляя по 'Артикул WB' (nmId).
- Оборачиваемость (дни) = остаток / средние продажи в день.
- Неходовой = 0 продаж за 90 дней ИЛИ оборот > SLOW_DOI дней.
- Дорогой = продажная цена >= PRICE_MIN.
- Дубли карточек КРОНА (0 продаж) помечаются отдельно: продажи по ним
  идут через основную карточку на другой организации.

Использование:
    python slow_moving_analysis.py <остатки.xlsx> [--out outputs/файл.xlsx]
"""
import json, sys, argparse, datetime as dt
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRICE_MIN = 1500   # порог "дорогой", руб (примерно медиана ассортимента)
SLOW_DOI = 180     # порог "неходовой" по оборачиваемости, дней
ROOT = Path(__file__).resolve().parents[2]


def read_stock(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    stock = {}
    for sh in wb.worksheets:
        hdr = list(next(sh.iter_rows(values_only=True)))
        def idx(name):
            return hdr.index(name) if name in hdr else None
        ix = {k: idx(v) for k, v in {
            'brand': 'Бренд', 'subject': 'Предмет', 'art': 'Артикул продавца',
            'nm': 'Артикул WB', 'stock': 'Всего находится на складах',
            'cost': 'Себестоимость', 'price': 'Продажная цена'}.items()}
        if ix['nm'] is None:
            continue
        for r in list(sh.iter_rows(values_only=True))[1:]:
            def g(i):
                return r[i] if i is not None and i < len(r) else None
            nm = g(ix['nm'])
            if nm is None:
                continue
            try:
                nm = int(nm)
            except (TypeError, ValueError):
                continue
            stock[nm] = {'entity': sh.title, 'brand': g(ix['brand']),
                         'subject': g(ix['subject']), 'art': g(ix['art']), 'nm': nm,
                         'stock': g(ix['stock']) or 0, 'cost': g(ix['cost']) or 0,
                         'price': g(ix['price']) or 0}
    return stock


def read_sales(today, days=90):
    sales = json.load(open(ROOT / 'data/wb_api/sales.json'))
    sold = defaultdict(int)
    sold30 = defaultdict(int)
    for s in sales:
        nm = s.get('nmId')
        age = (today - dt.date.fromisoformat(s['date'][:10])).days
        q = 1 if (s.get('forPay', 0) or 0) >= 0 else -1  # возвраты вычитаем
        if age <= days:
            sold[nm] += q
        if age <= 30:
            sold30[nm] += q
    return sold, sold30


def analyze(stock, sold, sold30):
    rows = []
    for nm, st in stock.items():
        s90 = sold.get(nm, 0)
        rate = s90 / 90.0
        stk = st['stock'] or 0
        doi = (stk / rate) if rate > 0 else None
        rows.append({**st, 's30': sold30.get(nm, 0), 's90': s90, 'doi': doi,
                     'frozen': stk * (st['cost'] or 0)})
    return rows


def is_slow(r):
    if r['stock'] <= 0:
        return False
    if r['s90'] == 0:
        return True
    return r['doi'] is not None and r['doi'] > SLOW_DOI


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('stock_file')
    ap.add_argument('--out', default=str(ROOT / 'outputs' / 'slow_moving.xlsx'))
    ap.add_argument('--today', default=str(dt.date.today()))
    a = ap.parse_args()
    today = dt.date.fromisoformat(a.today)

    stock = read_stock(a.stock_file)
    sold, sold30 = read_sales(today)
    rows = analyze(stock, sold, sold30)

    target = [r for r in rows if r['price'] >= PRICE_MIN and is_slow(r)]
    for r in target:
        if r['entity'] == 'КРОНА' and r['s90'] == 0:
            r['flag'] = 'Дубль-карточка (продажи идут через основную)'
        elif r['s90'] == 0:
            r['flag'] = 'Мертвый сток'
        else:
            r['flag'] = 'Медленный оборот'
    target.sort(key=lambda r: (0 if r['s90'] == 0 else 1, -r['frozen']))

    Path(a.out).parent.mkdir(parents=True, exist_ok=True)
    build_xlsx(rows, target, a.out, today)
    print(f'SKU всего: {len(rows)} | отобрано: {len(target)} | '
          f'заморожено: {sum(r["frozen"] for r in target):,.0f} руб')
    print('Сохранено:', a.out)


def build_xlsx(rows, target, out, today):
    wb = Workbook()
    H = Font(bold=True, color='FFFFFF'); HF = PatternFill('solid', fgColor='2F5496')
    red = PatternFill('solid', fgColor='FFC7CE'); yel = PatternFill('solid', fgColor='FFEB9C')
    ora = PatternFill('solid', fgColor='FCE4D6')
    thin = Border(*[Side(style='thin', color='D9D9D9')] * 4)
    ctr = Alignment(horizontal='center')

    ws = wb.active; ws.title = 'Итоги'
    ws['A1'] = 'Анализ неходовых дорогих товаров на WB'; ws['A1'].font = Font(bold=True, size=14)
    tot = sum(r['frozen'] for r in target)
    info = [('Критерий', f'цена ≥ {PRICE_MIN} ₽ И (0 продаж 90 дн ИЛИ оборот > {SLOW_DOI} дн)'),
            ('Всего SKU', len(rows)), ('Отобрано', len(target)),
            ('Заморожено, ₽', f'{tot:,.0f}'.replace(',', ' ')),
            ('Дата остатков / отчета', str(today))]
    for i, (k, v) in enumerate(info, 3):
        ws.cell(i, 1, k).font = Font(bold=True); ws.cell(i, 2, v)
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 65

    ws = wb.create_sheet('Неходовые дорогие')
    cols = [('#', 5), ('Бренд', 16), ('Товар', 46), ('Предмет', 16), ('Артикул WB', 13),
            ('Орг', 10), ('Остаток, шт', 11), ('Прод. 90д', 10), ('Прод. 30д', 10),
            ('Оборот, дн', 12), ('Прод. цена, ₽', 12), ('Себест., ₽', 11),
            ('Заморожено, ₽', 14), ('Пометка', 42)]
    for j, (c, w) in enumerate(cols, 1):
        cell = ws.cell(1, j, c); cell.font = H; cell.fill = HF; cell.alignment = ctr
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, r in enumerate(target, 1):
        doi = 'нет продаж' if r['s90'] == 0 else round(r['doi'])
        vals = [i, r['brand'] or '—', r['art'], r['subject'], r['nm'], r['entity'],
                int(r['stock']), r['s90'], r['s30'], doi, round(r['price']),
                round(r['cost']), round(r['frozen']), r['flag']]
        fill = red if r['s90'] == 0 and 'Дубль' not in r['flag'] else (
            ora if 'Дубль' in r['flag'] else yel)
        for j, v in enumerate(vals, 1):
            c = ws.cell(i + 1, j, v); c.border = thin; c.fill = fill
            if j in (7, 8, 9, 10, 11, 12, 13):
                c.alignment = ctr
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{len(target)+1}'

    ws = wb.create_sheet('Все SKU')
    cols2 = [('Бренд', 16), ('Товар', 46), ('Предмет', 16), ('Артикул WB', 13), ('Орг', 10),
             ('Остаток, шт', 11), ('Прод. 90д', 10), ('Прод. 30д', 10), ('Оборот, дн', 12),
             ('Прод. цена, ₽', 12), ('Себест., ₽', 11), ('Заморожено, ₽', 14)]
    for j, (c, w) in enumerate(cols2, 1):
        cell = ws.cell(1, j, c); cell.font = H; cell.fill = HF; cell.alignment = ctr
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, r in enumerate(sorted(rows, key=lambda r: -r['frozen']), 1):
        doi = 'нет продаж' if r['s90'] == 0 and r['stock'] > 0 else (
            round(r['doi']) if r['doi'] is not None else '—')
        vals = [r['brand'] or '—', r['art'], r['subject'], r['nm'], r['entity'],
                int(r['stock']), r['s90'], r['s30'], doi, round(r['price']),
                round(r['cost']), round(r['frozen'])]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i + 1, j, v); c.border = thin
            if j in (6, 7, 8, 9, 10, 11, 12):
                c.alignment = ctr
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols2))}{len(rows)+1}'
    wb.save(out)


if __name__ == '__main__':
    main()
