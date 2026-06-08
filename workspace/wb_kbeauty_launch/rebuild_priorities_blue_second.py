import re
from pathlib import Path

import openpyxl


SRC = Path("/Users/nikita/Desktop/Новая папка 2/outputs/wb_kbeauty_launch/WB_Kbeauty_roadmap_запуск_новых_товаров_2026-06-04.xlsx")
NELIQ = Path("/Users/nikita/Downloads/Нелеквиды.xlsx")
OUT = Path("/Users/nikita/Desktop/Новая папка 2/outputs/wb_kbeauty_launch/WB_Kbeauty_roadmap_приоритеты_зеленый_синий_желтый_2026-06-04.xlsx")


def norm_id(value):
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).strip().replace("\xa0", "")
    match = re.search(r"\d{6,}", text)
    return match.group(0) if match else ""


def norm_text(value):
    if value is None:
        return ""
    text = str(value).lower().replace("\xa0", " ")
    text = re.sub(r"[^a-zа-я0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def read_neliq_sheet1():
    workbook = openpyxl.load_workbook(NELIQ, data_only=True)
    sheet = workbook["Лист1"]
    by_id = {}
    names = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not any(value not in (None, "") for value in row):
            continue
        if row and str(row[0]).strip().lower() in ("коротич", "артикул wb"):
            continue
        item_id = norm_id(row[0] if len(row) > 0 else None)
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        comment = str(row[11] or "").strip() if len(row) > 11 else ""
        if not item_id and not name:
            continue
        item = {"sheet": "Лист1", "row": row_idx, "id": item_id, "name": name, "comment": comment}
        if item_id:
            by_id.setdefault(item_id, []).append(item)
        normalized_name = norm_text(name)
        if normalized_name:
            names.append((normalized_name, item))
    return by_id, names


def find_matches(sheet, by_id, names):
    matches = []
    matched_rows = set()
    for row_idx in range(2, sheet.max_row + 1):
        item_id = norm_id(sheet.cell(row_idx, 4).value)
        if item_id and item_id in by_id:
            matches.append({"row": row_idx, "type": "Артикул WB", "neliq": by_id[item_id][0]})
            matched_rows.add(row_idx)

    for row_idx in range(2, sheet.max_row + 1):
        if row_idx in matched_rows:
            continue
        launch_name = norm_text((sheet.cell(row_idx, 5).value or "") + " " + (sheet.cell(row_idx, 6).value or ""))
        if not launch_name:
            continue
        for neliq_name, neliq_item in names:
            if len(neliq_name) >= 15 and (neliq_name in launch_name or launch_name in neliq_name):
                matches.append({"row": row_idx, "type": "Название", "neliq": neliq_item})
                matched_rows.add(row_idx)
                break
    return matches


def action_for_status(status):
    if status == "Приоритет 1":
        return "Запускать в первую очередь"
    if status == "Приоритет 2":
        return "Запускать после приоритета 1"
    return "Запускать после приоритета 2"


def main():
    by_id, names = read_neliq_sheet1()
    workbook = openpyxl.load_workbook(SRC)
    launch = workbook["Запуск"]
    no_launch = workbook["Не запускаем"]
    matches = find_matches(launch, by_id, names)

    if "Исключено неликвиды" in workbook.sheetnames:
        del workbook["Исключено неликвиды"]
    excluded = workbook.create_sheet("Исключено неликвиды", 2)
    excluded_headers = [
        "Очередь",
        "Старый статус",
        "Бренд",
        "Артикул WB",
        "Наименование",
        "Русское название",
        "Ед.изм.",
        "Себестоимость, ₽",
        "Остаток доступный",
        "Остаток WB",
        "Группа из файла",
        "Действие",
        "Источник неликвидов",
        "Строка неликвидов",
        "Совпадение",
        "Комментарий из неликвидов",
    ]
    excluded.append(excluded_headers)

    for match in matches:
        row_idx = match["row"]
        neliq_item = match["neliq"]
        excluded.append(
            [launch.cell(row_idx, col_idx).value for col_idx in range(1, 13)]
            + [neliq_item["sheet"], neliq_item["row"], match["type"], neliq_item["comment"]]
        )
        no_launch.append([
            "Не запускаем",
            launch.cell(row_idx, 3).value,
            launch.cell(row_idx, 4).value,
            launch.cell(row_idx, 5).value,
            launch.cell(row_idx, 6).value,
            launch.cell(row_idx, 8).value,
            launch.cell(row_idx, 9).value,
            launch.cell(row_idx, 10).value,
            launch.cell(row_idx, 11).value,
            f"Исключено: есть в файле Нелеквиды.xlsx (Лист1, строка {neliq_item['row']})",
            launch.cell(row_idx, 13).value or "",
        ])

    # Collect launch rows that are not excluded, then remap priority:
    # old green = priority 1, old blue = priority 2, old yellow = priority 3.
    excluded_row_nums = {match["row"] for match in matches}
    launch_rows = []
    for row_idx in range(2, launch.max_row + 1):
        if row_idx in excluded_row_nums:
            continue
        values = [launch.cell(row_idx, col_idx).value for col_idx in range(1, launch.max_column + 1)]
        old_status = values[1]
        if old_status == "Приоритет 2":
            new_status = "Приоритет 3"
        elif old_status == "Приоритет 3":
            new_status = "Приоритет 2"
        else:
            new_status = old_status
        values[1] = new_status
        values[11] = action_for_status(new_status)
        launch_rows.append(values)

    priority_rank = {"Приоритет 1": 1, "Приоритет 2": 2, "Приоритет 3": 3}
    launch_rows.sort(key=lambda values: (priority_rank.get(values[1], 99), values[0] or 9999))

    if launch.max_row > 1:
        launch.delete_rows(2, launch.max_row - 1)
    for queue, values in enumerate(launch_rows, start=1):
        values[0] = queue
        launch.append(values)

    counts = {
        "В запуске": launch.max_row - 1,
        "Приоритет 1 / зеленые": 0,
        "Приоритет 2 / желтые": 0,
        "Приоритет 3 / голубые": 0,
        "Не запускаем / красные": no_launch.max_row - 1,
    }
    for row_idx in range(2, launch.max_row + 1):
        status = launch.cell(row_idx, 2).value
        if status == "Приоритет 1":
            counts["Приоритет 1 / зеленые"] += 1
        elif status == "Приоритет 2":
            counts["Приоритет 2 / желтые"] += 1
        elif status == "Приоритет 3":
            counts["Приоритет 3 / голубые"] += 1

    summary = workbook["Итог"]
    label_overrides = {
        "Приоритет 2 / желтые": "Приоритет 2 / синие",
        "Приоритет 3 / голубые": "Приоритет 3 / желтые",
    }
    for row_idx in range(1, summary.max_row + 1):
        label = summary.cell(row_idx, 1).value
        if label in label_overrides:
            summary.cell(row_idx, 1).value = label_overrides[label]

    summary_counts = {
        "В запуске": counts["В запуске"],
        "Приоритет 1 / зеленые": counts["Приоритет 1 / зеленые"],
        "Приоритет 2 / синие": counts["Приоритет 2 / желтые"],
        "Приоритет 3 / желтые": counts["Приоритет 3 / голубые"],
        "Не запускаем / красные": counts["Не запускаем / красные"],
    }
    for row_idx in range(1, summary.max_row + 1):
        label = summary.cell(row_idx, 1).value
        if label in summary_counts:
            summary.cell(row_idx, 2).value = summary_counts[label]
    labels = [summary.cell(row_idx, 1).value for row_idx in range(1, summary.max_row + 1)]
    if "Исключено по файлу неликвидов" not in labels:
        row_idx = summary.max_row + 1
        summary.cell(row_idx, 1).value = "Исключено по файлу неликвидов"
        summary.cell(row_idx, 2).value = len(matches)
    else:
        for row_idx in range(1, summary.max_row + 1):
            if summary.cell(row_idx, 1).value == "Исключено по файлу неликвидов":
                summary.cell(row_idx, 2).value = len(matches)

    excluded.auto_filter.ref = f"A1:P{excluded.max_row}"
    excluded.freeze_panes = "A2"
    launch.auto_filter.ref = launch.dimensions
    workbook.save(OUT)
    print(OUT)
    print(
        f"launch={launch.max_row - 1} "
        f"p1={summary_counts['Приоритет 1 / зеленые']} "
        f"p2_blue={summary_counts['Приоритет 2 / синие']} "
        f"p3_yellow={summary_counts['Приоритет 3 / желтые']} "
        f"excluded_sheet1={len(matches)}"
    )


if __name__ == "__main__":
    main()
