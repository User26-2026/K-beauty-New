import re
from pathlib import Path

import openpyxl


SRC = Path("/Users/nikita/Desktop/Новая папка 2/outputs/wb_kbeauty_launch/WB_Kbeauty_roadmap_запуск_новых_товаров_2026-06-04.xlsx")
NELIQ = Path("/Users/nikita/Downloads/Нелеквиды.xlsx")
OUT = Path("/Users/nikita/Desktop/Новая папка 2/outputs/wb_kbeauty_launch/WB_Kbeauty_roadmap_без_неликвидов_только_лист1_2026-06-04.xlsx")


def norm_id(value):
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).strip().replace("\xa0", "")
    match = re.search(r"\d{6,}", text)
    return match.group(0) if match else ""


def norm_text(value):
    if value is None:
        return ""
    text = str(value).lower().replace("\xa0", " ")
    text = re.sub(r"[^a-zа-я0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def read_sheet1_neliq():
    workbook = openpyxl.load_workbook(NELIQ, data_only=True)
    sheet = workbook["Лист1"]
    by_id = {}
    names = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not any(value not in (None, "") for value in row):
            continue
        if row and str(row[0]).strip().lower() in ("коротич", "артикул wb"):
            continue
        item_id = norm_id(row[0] if len(row) > 0 else None)
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        comment = str(row[11] or "").strip() if len(row) > 11 else ""
        if not item_id and not name:
            continue
        item = {
            "sheet": "Лист1",
            "row": row_idx,
            "id": item_id,
            "name": name,
            "comment": comment,
        }
        if item_id:
            by_id.setdefault(item_id, []).append(item)
        normalized = norm_text(name)
        if normalized:
            names.append((normalized, item))
    return by_id, names


def find_matches(launch_sheet, by_id, names):
    matches = []
    matched_rows = set()
    for row_idx in range(2, launch_sheet.max_row + 1):
        item_id = norm_id(launch_sheet.cell(row_idx, 4).value)
        if item_id and item_id in by_id:
            matches.append({
                "row": row_idx,
                "type": "Артикул WB",
                "neliq": by_id[item_id][0],
            })
            matched_rows.add(row_idx)

    for row_idx in range(2, launch_sheet.max_row + 1):
        if row_idx in matched_rows:
            continue
        launch_name = norm_text(
            (launch_sheet.cell(row_idx, 5).value or "")
            + " "
            + (launch_sheet.cell(row_idx, 6).value or "")
        )
        if not launch_name:
            continue
        for neliq_name, neliq_item in names:
            if len(neliq_name) >= 15 and (neliq_name in launch_name or launch_name in neliq_name):
                matches.append({
                    "row": row_idx,
                    "type": "Название",
                    "neliq": neliq_item,
                })
                matched_rows.add(row_idx)
                break
    return matches


def main():
    by_id, names = read_sheet1_neliq()
    workbook = openpyxl.load_workbook(SRC)
    launch_sheet = workbook["Запуск"]
    no_launch_sheet = workbook["Не запускаем"]
    matches = find_matches(launch_sheet, by_id, names)

    if "Исключено неликвиды" in workbook.sheetnames:
        del workbook["Исключено неликвиды"]
    excluded_sheet = workbook.create_sheet("Исключено неликвиды", 2)
    excluded_sheet.append([
        "Очередь",
        "Статус запуска",
        "Бренд",
        "Артикул WB",
        "Наименование",
        "Русское название",
        "Ед.изм.",
        "Себестоимость, ₽",
        "Остаток доступный",
        "Остаток WB",
        "Группа из файла",
        "Действие",
        "Источник неликвидов",
        "Строка неликвидов",
        "Совпадение",
        "Комментарий из неликвидов",
    ])

    for match in matches:
        row_idx = match["row"]
        neliq_item = match["neliq"]
        excluded_sheet.append(
            [launch_sheet.cell(row_idx, col_idx).value for col_idx in range(1, 13)]
            + [
                neliq_item["sheet"],
                neliq_item["row"],
                match["type"],
                neliq_item["comment"],
            ]
        )

    for col, width in {
        "A": 9,
        "B": 18,
        "C": 16,
        "D": 14,
        "E": 55,
        "F": 38,
        "H": 14,
        "I": 14,
        "J": 12,
        "M": 18,
        "N": 14,
        "O": 14,
        "P": 60,
    }.items():
        excluded_sheet.column_dimensions[col].width = width
    excluded_sheet.freeze_panes = "A2"
    excluded_sheet.auto_filter.ref = f"A1:P{excluded_sheet.max_row}"

    for match in matches:
        row_idx = match["row"]
        neliq_item = match["neliq"]
        no_launch_sheet.append([
            "Не запускаем",
            launch_sheet.cell(row_idx, 3).value,
            launch_sheet.cell(row_idx, 4).value,
            launch_sheet.cell(row_idx, 5).value,
            launch_sheet.cell(row_idx, 6).value,
            launch_sheet.cell(row_idx, 8).value,
            launch_sheet.cell(row_idx, 9).value,
            launch_sheet.cell(row_idx, 10).value,
            launch_sheet.cell(row_idx, 11).value,
            f"Исключено: есть в файле Нелеквиды.xlsx (Лист1, строка {neliq_item['row']})",
            launch_sheet.cell(row_idx, 13).value or "",
        ])

    for row_idx in sorted([match["row"] for match in matches], reverse=True):
        launch_sheet.delete_rows(row_idx, 1)

    for queue, row_idx in enumerate(range(2, launch_sheet.max_row + 1), start=1):
        launch_sheet.cell(row_idx, 1).value = queue

    counts = {
        "В запуске": launch_sheet.max_row - 1,
        "Приоритет 1 / зеленые": 0,
        "Приоритет 2 / желтые": 0,
        "Приоритет 3 / голубые": 0,
        "Не запускаем / красные": no_launch_sheet.max_row - 1,
    }
    for row_idx in range(2, launch_sheet.max_row + 1):
        status = launch_sheet.cell(row_idx, 2).value
        if status == "Приоритет 1":
            counts["Приоритет 1 / зеленые"] += 1
        elif status == "Приоритет 2":
            counts["Приоритет 2 / желтые"] += 1
        elif status == "Приоритет 3":
            counts["Приоритет 3 / голубые"] += 1

    summary_sheet = workbook["Итог"]
    for row_idx in range(1, summary_sheet.max_row + 1):
        label = summary_sheet.cell(row_idx, 1).value
        if label in counts:
            summary_sheet.cell(row_idx, 2).value = counts[label]
    labels = [summary_sheet.cell(row_idx, 1).value for row_idx in range(1, summary_sheet.max_row + 1)]
    if "Исключено по файлу неликвидов" not in labels:
        row_idx = summary_sheet.max_row + 1
        summary_sheet.cell(row_idx, 1).value = "Исключено по файлу неликвидов"
        summary_sheet.cell(row_idx, 2).value = len(matches)
    else:
        for row_idx in range(1, summary_sheet.max_row + 1):
            if summary_sheet.cell(row_idx, 1).value == "Исключено по файлу неликвидов":
                summary_sheet.cell(row_idx, 2).value = len(matches)

    workbook.save(OUT)
    print(OUT)
    print(
        f"matches={len(matches)} launch={launch_sheet.max_row - 1} "
        f"p1={counts['Приоритет 1 / зеленые']} "
        f"p2={counts['Приоритет 2 / желтые']} "
        f"p3={counts['Приоритет 3 / голубые']}"
    )
    for match in matches:
        row_idx = match["row"]
        neliq_item = match["neliq"]
        print(
            f"{launch_sheet.cell(row_idx, 4).value}\t"
            f"{neliq_item['row']}\t{neliq_item['name'][:100]}"
        )


if __name__ == "__main__":
    main()
