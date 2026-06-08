from __future__ import annotations

import json
from pathlib import Path

import openpyxl


INPUT = Path("/Users/nikita/Downloads/Celimax.xlsx")


def main() -> None:
    wb_formula = openpyxl.load_workbook(INPUT, data_only=False)
    ws_formula = wb_formula["Юнит"]
    print("shape", ws_formula.max_row, ws_formula.max_column)
    keywords = ["дрр", "roi", "маржа", "реклам", "себес", "цена"]
    for row in ws_formula.iter_rows():
        values = [cell.value for cell in row]
        if any(isinstance(value, str) and any(key in value.lower() for key in keywords) for value in values):
            print("FORMULA", row[0].row, json.dumps(values[:30], ensure_ascii=False, default=str))

    wb_values = openpyxl.load_workbook(INPUT, data_only=True)
    ws_values = wb_values["Юнит"]
    for row_num in range(1, min(ws_values.max_row, 40) + 1):
        print(
            "VALUE",
            row_num,
            json.dumps(
                [ws_values.cell(row_num, col).value for col in range(1, min(ws_values.max_column, 28) + 1)],
                ensure_ascii=False,
                default=str,
            ),
        )

    for row_num in range(11, 16):
        print(
            "ROW_FORMULA",
            row_num,
            json.dumps(
                [ws_formula.cell(row_num, col).value for col in range(8, 29)],
                ensure_ascii=False,
                default=str,
            ),
        )


if __name__ == "__main__":
    main()
