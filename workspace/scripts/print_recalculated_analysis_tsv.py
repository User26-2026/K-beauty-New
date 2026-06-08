from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl


INPUT = Path("/Users/nikita/Desktop/Новая папка 2/outputs/wb_product_research/Список товаров для анализа - проценты пересчитаны.xlsx")


def fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(round(value, 1)).replace(".", ",")
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def main() -> None:
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    ws = wb["Лист2"]
    lines = []
    for row in range(1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, 11)]
        if row == 1 or any(value not in (None, "") for value in values[:3]):
            lines.append("\t".join(fmt(value) for value in values))
    print("\n".join(lines))


if __name__ == "__main__":
    main()
