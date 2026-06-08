from __future__ import annotations

import json
from statistics import median
from pathlib import Path

import openpyxl


ROOT = Path("/Users/nikita/Desktop/Новая папка 2")
UNIT_JSON = ROOT / "outputs/wb_product_research/roundlab_unit_economics_from_celimax_2026-06-01.json"
SOURCE_XLSX = Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос__Общая выдача_1.xlsx")


ITEMS = [
    (212968870, "Увлажняющий тонер для лица с морской водой"),
    (221483595, "Мицеллярная вода для снятия макияжа с морской водой"),
    (212957741, "Пенка очищающая для лица от прыщей с сосной и кислотами"),
    (241746162, "Крем для рук увлажняющий для сухой кожи с берёзовым соком"),
    (212968861, "Тонер для лица увлажняющий и омолаживающий с соей"),
    (221500250, "Гель для душа от прыщей с кислотами"),
    (212962100, "Сыворотка для лица от прыщей с сосной"),
    (212806814, "Лосьон для лица мужской с березовым соком"),
    (410070711, "Крем для лица увлажняющий для проблемной кожи с сосной"),
]


# Себестоимость: supply KRW из ROUNDLAB PRICE LIST -> рубли -> * 1.4.
PRICE_MAP = {
    212968870: ("8809657114731", 484),
    221483595: ("8809738607343", 461),
    212957741: ("8809864754126", 447),
    241746162: ("8809962540713", 216),
    212968861: ("8809624726370", 655),
    221500250: ("8809849802408", 531),
    212962100: ("8809783325667", 671),
    212806814: ("8809962541239", 606),
    410070711: ("8809783325674", 783),
}


MEDIAN_GROUPS = {
    212968870: ["тонер", "морской водой"],
    221483595: ["мицелляр", "морской водой"],
    212957741: ["пенк", "сосн"],
    241746162: ["крем", "рук"],
    212968861: ["тонер", "соей"],
    221500250: ["гель для душа", "прыщ"],
    212962100: ["сыворот", "сосн"],
    212806814: ["лосьон", "муж", "берез"],
    410070711: ["крем", "проблемной", "сосн"],
}


def clean(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").strip()


def source_rows() -> list[dict]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [dict(zip(headers, values)) for values in ws.iter_rows(min_row=2, values_only=True)]


def median_price(article: int, fallback: float | None, rows: list[dict]) -> float | None:
    terms = MEDIAN_GROUPS[article]
    prices = []
    for row in rows:
        if str(row.get("Бренд", "")).lower().strip() != "round lab":
            continue
        name = clean(row.get("Товар")).lower().replace("ё", "е")
        if all(term.replace("ё", "е") in name for term in terms):
            try:
                prices.append(float(row.get("Цена (с СПП)") or 0))
            except (TypeError, ValueError):
                pass
    prices = [price for price in prices if price > 0]
    if prices:
        return float(median(prices))
    return fallback


def calc_margin_no_drr(row: dict, price_spp: float, cost: float) -> tuple[float, float, float]:
    price_wb = round(price_spp / 0.7)
    price_spp_calc = price_wb * 0.7
    margin_no_drr = (
        price_wb
        - float(row["Доп обработка склад"])
        - price_spp_calc * 0.08
        - float(row["Логистика WB"])
        - price_wb * 0.325
        - price_spp_calc * 0.0217
        - float(row["Обратная логистика"])
        - float(row["Хранение"])
        - cost
    )
    return price_wb, margin_no_drr, margin_no_drr / price_wb * 100


def main() -> None:
    data = json.loads(UNIT_JSON.read_text(encoding="utf-8"))
    by_article = {int(row["Артикул WB"]): row for row in data["rows"]}
    market_rows = source_rows()

    headers = [
        "Бренд",
        "Артикул WB",
        "Название",
        "Баркод из прайса",
        "Расч. выручка топ карточки, руб/мес",
        "Расч. заказы топ карточки, шт/мес",
        "ROI без ДРР, %",
        "Маржа/шт без ДРР, руб.",
        "Медианная цена с СПП, ₽",
        "Себестоимость, ₽",
    ]
    print("\t".join(headers))

    for article, fallback_name in ITEMS:
        row = by_article.get(article)
        barcode, cost = PRICE_MAP[article]
        if row:
            price_spp = median_price(article, float(row["Цена с СПП, ₽"]), market_rows)
            _, margin_no_drr, _ = calc_margin_no_drr(row, price_spp, float(cost))
            roi = round(margin_no_drr / float(cost) * 100, 1)
            out = [
                "Round Lab",
                article,
                row["Товар"] or fallback_name,
                barcode,
                round(float(row["Расч. выручка, ₽"])),
                round(float(row["Расч. заказы"])),
                roi,
                round(margin_no_drr),
                round(float(price_spp)),
                round(float(cost)),
            ]
        else:
            out = [
                "Round Lab",
                article,
                fallback_name,
                barcode,
                "",
                "",
                "",
                "",
                "",
                round(float(cost)),
            ]
        print("\t".join(clean(value) for value in out))


if __name__ == "__main__":
    main()
