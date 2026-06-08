from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path("/Users/nikita/Desktop/Новая папка 2")
SOURCE = Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос__Общая выдача_1.xlsx")
UNIT_SOURCE = Path("/Users/nikita/Downloads/!Корея-Китай-ЮНИТ -ШАБЛОН.xlsx")
OUTPUT = ROOT / "outputs/wb_product_research/roundlab_creamshop_candidates_2026-06-01.json"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def line_for_name(name: str) -> str:
    low = name.lower()
    if "солнцезащит" in low or "спф" in low or "spf" in low:
        return "SPF / защита от солнца"
    if "пенк" in low and "морской водой" in low:
        return "1025 Dokdo Cleanser"
    if "тонер" in low and ("морской водой" in low or "500 мл" in low):
        return "1025 Dokdo Toner"
    if "мицелляр" in low:
        return "1025 Dokdo Cleansing Water"
    if "сосн" in low or "прыщ" in low or "акне" in low:
        return "Pine / acne care"
    if "рук" in low:
        return "Hand cream"
    if "душ" in low or "тела" in low:
        return "Body care"
    if "муж" in low:
        return "Men line"
    if "pdrn" in low or "пдрн" in low:
        return "PDRN / collagen"
    if "вокруг глаз" in low or "роликом" in low:
        return "Eye care"
    return "Прочее"


def kbeauty_status(name: str) -> tuple[str, str]:
    low = name.lower()
    if "пенк" in low and "морской водой" in low:
        return (
            "Дубль подтвержден",
            "В K-Beauty найден/планировался ROUND LAB 1025 DOKDO CLEANSER 150ml, nm 437154709. Не заводить как новый SKU без причины.",
        )
    if "солнцезащитный крем" in low and ("спф" in low or "spf" in low) and "стик" not in low and "тональный" not in low and "физических" not in low:
        return (
            "Вероятный дубль/сезонный риск",
            "В юнитке K-Beauty есть ROUND LAB BIRCH JUICE MOISTURIZING SUNSCREEN 50ml, nm 437231408. Нужна проверка активной карточки.",
        )
    return ("Не найден в текущих отчетах", "В продажах/остатках/отчетах K-Beauty дубль не найден по имеющимся файлам.")


def purchase_status(row: dict) -> tuple[str, str, str]:
    name = str(row["Товар"])
    orders = row["Расч. заказы"] or 0
    revenue = row["Расч. выручка\nпо заказам (с СПП)"] or 0
    missed = row["Упущенная выручка"] or 0
    duplicate, _ = kbeauty_status(name)
    low = name.lower()

    if duplicate == "Дубль подтвержден":
        return (
            "Не заводить как новый",
            "Дубль",
            "Спрос сильный, но товар уже пересекается с K-Beauty. Лучше анализировать текущую карточку/остатки, а не закупать как новый SKU.",
        )

    if "солнцезащит" in low or "спф" in low or "spf" in low:
        return (
            "Проверить после сезонного решения",
            "Сезонный риск",
            "Продажи есть, но SPF на 01.06 уже рискованнее для первой закупки. Брать только при быстрой поставке и понятной цене.",
        )

    if orders >= 250 and revenue >= 450_000:
        return (
            "Высокая",
            "Кандидат",
            "Есть заметный спрос, выручка и упущенная выручка, дубль в текущих отчетах K-Beauty не найден.",
        )
    if orders >= 120 or missed >= 80_000:
        return (
            "Средняя",
            "Кандидат к проверке",
            "Спрос есть, но перед закупкой нужно проверить себестоимость, документы и карточки-конкуренты.",
        )
    return (
        "Низкая",
        "Второй эшелон",
        "Слабее по объему заказов/выручке. Оставить как резерв после проверки сильных SKU.",
    )


def load_roundlab_unit_rows() -> dict[int, dict]:
    wb = openpyxl.load_workbook(UNIT_SOURCE, data_only=True)
    ws = wb["КОРЕЯ"]
    rows: dict[int, dict] = {}
    for excel_row in range(1, ws.max_row + 1):
        nm = clean(ws.cell(excel_row, 3).value)
        name = clean(ws.cell(excel_row, 4).value)
        if not isinstance(nm, int) or "ROUND LAB" not in str(name).upper():
            continue
        rows[nm] = {
            "nm": nm,
            "name": name,
            "base_cost": clean(ws.cell(excel_row, 14).value),
            "cost": clean(ws.cell(excel_row, 16).value),
            "price": clean(ws.cell(excel_row, 17).value),
            "row": excel_row,
        }
    return rows


UNIT_ROWS = load_roundlab_unit_rows()


def unit_cost_match(name: str, article: int) -> dict:
    low = name.lower()
    # Strong matches only. Similar but not identical Round Lab items are left blank.
    mapping = {
        212957743: 437154709,  # 1025 Dokdo Cleanser 150 ml
        212798848: 437231408,  # Birch Juice Moisturizing Sunscreen 50 ml
        212962100: 437776835,  # Pine Calming Cica Ampoule 30 ml
    }
    quality = {
        212957743: "точное сопоставление по линейке/типу",
        212798848: "вероятное сопоставление с Birch Juice SPF",
        212962100: "вероятное сопоставление с Pine Cica Ampoule",
    }
    unit_nm = mapping.get(article)
    if not unit_nm:
        return {
            "cost": "",
            "base_cost": "",
            "source_nm": "",
            "source_name": "",
            "match": "нет надежного сопоставления в прайсе",
        }
    unit = UNIT_ROWS.get(unit_nm)
    if not unit:
        return {
            "cost": "",
            "base_cost": "",
            "source_nm": "",
            "source_name": "",
            "match": "позиция не найдена в юнит-файле",
        }
    return {
        "cost": unit["cost"],
        "base_cost": unit["base_cost"],
        "source_nm": unit["nm"],
        "source_name": unit["name"],
        "match": quality.get(article, "сопоставлено"),
    }


wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
headers = [cell.value for cell in ws[1]]

rows = []
for excel_row in range(2, ws.max_row + 1):
    values = [clean(ws.cell(excel_row, col).value) for col in range(1, ws.max_column + 1)]
    row = dict(zip(headers, values))
    if str(row.get("Бренд", "")).lower().strip() != "round lab":
        continue
    if str(row.get("Продавец", "")).strip() != "Cream.Shop":
        continue

    duplicate_status, duplicate_comment = kbeauty_status(str(row["Товар"]))
    purchase_relevance, decision, purchase_comment = purchase_status(row)
    cost_match = unit_cost_match(str(row["Товар"]), int(row["Артикул"]))
    rows.append(
        {
            "Позиция": row["Позиция"],
            "Артикул WB": row["Артикул"],
            "Товар": row["Товар"],
            "Линейка/тип": line_for_name(str(row["Товар"])),
            "Предмет": row["Предмет"],
            "Бренд": row["Бренд"],
            "Продавец": row["Продавец"],
            "Цена с СПП, ₽": row["Цена (с СПП)"],
            "Расч. заказы": row["Расч. заказы"],
            "Расч. выручка, ₽": row["Расч. выручка\nпо заказам (с СПП)"],
            "Упущенная выручка, ₽": row["Упущенная выручка"],
            "Наличие": row["Наличие"],
            "Рейтинг": row["Рейтинг"],
            "Оценки": row["Оценки"],
            "Возраст товара, дней": row["Возраст товара"],
            "Полки TOP-10": row["Кол-во полок в ТОП-10"],
            "Полки TOP-100": row["Кол-во полок в ТОП-100"],
            "Себестоимость из прайса, ₽": cost_match["cost"],
            "Базовая себестоимость, ₽": cost_match["base_cost"],
            "Источник себестоимости": cost_match["source_name"],
            "Артикул себестоимости": cost_match["source_nm"],
            "Точность сопоставления себестоимости": cost_match["match"],
            "Статус дубля с K-Beauty": duplicate_status,
            "Комментарий по дублю": duplicate_comment,
            "Актуальность закупки": purchase_relevance,
            "Итог": decision,
            "Комментарий по закупке": purchase_comment,
        }
    )

summary = {
    "source_file": str(SOURCE),
    "seller": "Cream.Shop",
    "brand": "round lab",
    "total_rows": len(rows),
    "high": sum(1 for r in rows if r["Актуальность закупки"] == "Высокая"),
    "medium": sum(1 for r in rows if r["Актуальность закупки"] == "Средняя"),
    "seasonal": sum(1 for r in rows if r["Актуальность закупки"] == "Проверить после сезонного решения"),
    "duplicates": sum(1 for r in rows if "Дубль" in r["Статус дубля с K-Beauty"]),
    "cost_matched": sum(1 for r in rows if r["Себестоимость из прайса, ₽"] != ""),
}

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps({"summary": summary, "rows": rows}, ensure_ascii=False, indent=2), encoding="utf-8")
print(OUTPUT)
