from __future__ import annotations

from pathlib import Path

import openpyxl


ROOT = Path("/Users/nikita/Desktop/Новая папка 2")
INPUT = Path("/Users/nikita/Downloads/Список товаров для анализа.xlsx")
OUTPUT = ROOT / "outputs/wb_product_research/Список товаров для анализа - проценты пересчитаны.xlsx"


def as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def main() -> None:
    wb = openpyxl.load_workbook(INPUT)
    ws = wb["Лист2"]
    ws["G1"] = "ROI без ДРР, %"

    recalculated = 0
    for row in range(2, ws.max_row + 1):
        brand = ws.cell(row, 1).value
        title = ws.cell(row, 3).value
        margin = as_number(ws.cell(row, 8).value)
        cost = as_number(ws.cell(row, 10).value)
        if not brand and not title:
            continue
        if margin is None or cost in (None, 0):
            ws.cell(row, 7).value = None
            continue
        ws.cell(row, 7).value = round(margin / cost * 100, 1)
        ws.cell(row, 7).number_format = "0.0"
        recalculated += 1

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"{OUTPUT}\nrecalculated={recalculated}")


if __name__ == "__main__":
    main()
