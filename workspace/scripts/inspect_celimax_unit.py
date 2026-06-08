from __future__ import annotations

from pathlib import Path

import openpyxl


PATH = Path("/Users/nikita/Downloads/Celimax.xlsx")


def main() -> None:
    wb = openpyxl.load_workbook(PATH, data_only=False, read_only=False)
    wb_values = openpyxl.load_workbook(PATH, data_only=True, read_only=False)
    print("sheets:", wb.sheetnames)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wsv = wb_values[sheet_name]
        print(f"\nSHEET {sheet_name} rows={ws.max_row} cols={ws.max_column}")
        print("merged:", [str(rng) for rng in list(ws.merged_cells.ranges)[:12]])
        for r in range(1, min(ws.max_row, 18) + 1):
            values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 24) + 1)]
            print(r, values)

        if sheet_name == "Юнит":
            print("wide rows 8:17")
            for r in range(8, 18):
                values = [
                    (c, ws.cell(r, c).value, wsv.cell(r, c).value)
                    for c in range(1, min(ws.max_column, 28) + 1)
                ]
                print(r, values)
            print(
                "hidden_or_width_cols:",
                [
                    (col, dim.hidden, dim.width)
                    for col, dim in ws.column_dimensions.items()
                    if dim.hidden or dim.width
                ],
            )

        print("formula sample:")
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    print(cell.coordinate, value, "=>", wsv[cell.coordinate].value)
                    count += 1
                    if count >= 120:
                        break
            if count >= 120:
                break
        print("formula_sample_count:", count)


if __name__ == "__main__":
    main()
