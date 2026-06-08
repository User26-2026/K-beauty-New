from __future__ import annotations

import json
from pathlib import Path

import openpyxl


FILES = [
    Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос__Общая выдача_1.xlsx"),
    Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_axis-y_Общая выдача_1.xlsx"),
    Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_дерма фактори_Общая выдача_1.xlsx"),
]


def main() -> None:
    for path in FILES:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print("FILE", path.name, "rows", ws.max_row, "cols", ws.max_column)
        print(json.dumps(headers, ensure_ascii=False))
        rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
        sellers = {}
        brands = {}
        for row in rows:
            sellers[str(row.get("Продавец"))] = sellers.get(str(row.get("Продавец")), 0) + 1
            brands[str(row.get("Бренд"))] = brands.get(str(row.get("Бренд")), 0) + 1
        print("BRANDS", json.dumps(sorted(brands.items(), key=lambda x: -x[1])[:20], ensure_ascii=False))
        print("SELLERS", json.dumps(sorted(sellers.items(), key=lambda x: -x[1])[:20], ensure_ascii=False))
        for row in rows[:12]:
            sample = {key: row.get(key) for key in ["Артикул", "Товар", "Бренд", "Продавец", "Цена (с СПП)", "Расч. заказы", "Расч. выручка\nпо заказам (с СПП)"]}
            print(json.dumps(sample, ensure_ascii=False, default=str))
        print()


if __name__ == "__main__":
    main()
