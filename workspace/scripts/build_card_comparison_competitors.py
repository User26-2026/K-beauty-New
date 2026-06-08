from __future__ import annotations

import csv
import io
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path("/Users/nikita/Desktop/Новая папка 2")
PYTHON = "/Users/nikita/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3"
ANALYSIS = ROOT / "outputs/wb_product_research/Список товаров для анализа - проценты пересчитаны.xlsx"

CONFIGS = {
    "AXIS-Y": {
        "source": Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_axis-y_Общая выдача_1.xlsx"),
        "script": ROOT / "scripts/match_axisy_barcodes.py",
    },
    "DERMA FACTORY": {
        "source": Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_дерма фактори_Общая выдача_1.xlsx"),
        "script": ROOT / "scripts/match_dermafactory_barcodes.py",
    },
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def num(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def max_drr(margin, price_spp, cost):
    if margin is None or price_spp in (None, 0) or cost in (None, 0):
        return None
    roi = margin / cost
    if roi <= 0.40:
        return None
    allowed_ad_spend = margin - cost * 0.40
    price_before_spp = round(price_spp / 0.70)
    return allowed_ad_spend / (price_before_spp * 0.50) * 100


def load_targets() -> pd.DataFrame:
    wb = openpyxl.load_workbook(ANALYSIS, data_only=True)
    ws = wb["Лист2"]
    headers = [ws.cell(1, col).value for col in range(1, 11)]
    rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, 11)]
        if not any(value not in (None, "") for value in values[:3]):
            continue
        item = dict(zip(headers, values))
        margin = num(item.get("Маржа/шт без ДРР, руб."))
        price = num(item.get("Цена с СПП топ карточки, ₽"))
        cost = num(item.get("Себестоимость, ₽"))
        revenue = num(item.get("Расч. выручка топ карточки, руб/мес")) or 0
        drr = max_drr(margin, price, cost)
        if drr is not None and revenue > 100_000:
            item["Артикул WB"] = str(int(num(item["Ссылка/артикул"])))
            item["Макс ДРР"] = round(drr, 1)
            rows.append(item)
    return pd.DataFrame(rows)


def load_brand_pool(brand: str) -> pd.DataFrame:
    config = CONFIGS[brand]
    raw = subprocess.check_output([PYTHON, str(config["script"])], cwd=ROOT, text=True)
    matched = pd.DataFrame(list(csv.DictReader(io.StringIO(raw), delimiter="\t")))
    source = pd.read_excel(config["source"])
    source["Артикул WB"] = source["Артикул"].astype(int).astype(str)
    source["Заказы, шт/мес"] = pd.to_numeric(source["Расч. заказы"], errors="coerce").fillna(0).astype(int)
    source["Выручка, руб/мес"] = pd.to_numeric(source["Расч. выручка\nпо заказам (с СПП)"], errors="coerce").fillna(0).astype(int)
    source["Цена с СПП, ₽"] = pd.to_numeric(source["Цена (с СПП)"], errors="coerce").fillna(0).astype(int)
    source = source[
        [
            "Артикул WB",
            "Товар",
            "Продавец",
            "Заказы, шт/мес",
            "Выручка, руб/мес",
            "Цена с СПП, ₽",
        ]
    ]
    pool = matched.merge(source, on="Артикул WB", how="left")
    pool = pool[pool["Barcode из прайса"].astype(str).str.len() > 0].copy()
    return pool


def main() -> None:
    targets = load_targets()
    pools = {brand: load_brand_pool(brand) for brand in CONFIGS}
    rows = []
    queue_rows = []

    for _, target in targets.iterrows():
        brand = clean(target["Бренд"])
        target_article = clean(target["Артикул WB"])
        target_name = clean(target["Название"])
        if brand not in pools:
            rows.append(
                {
                    "Бренд": brand,
                    "Целевой артикул": target_article,
                    "Целевой товар": target_name,
                    "Конкурент артикул": "",
                    "Конкурент товар": "Нужна отдельная выгрузка по запросу/WB-сравнение: в текущем файле нет внешних продавцов-конкурентов",
                    "Продавец": "",
                    "Заказы конкурента, шт/мес": "",
                    "Выручка конкурента, руб/мес": "",
                    "Цена конкурента с СПП, ₽": "",
                    "Тип подбора": "нет данных",
                }
            )
            continue

        pool = pools[brand]
        target_match = pool[pool["Артикул WB"].eq(target_article)]
        if target_match.empty:
            continue
        barcode = target_match.iloc[0]["Barcode из прайса"]
        group = pool[pool["Barcode из прайса"].eq(barcode)].copy()
        group = group[group["Артикул WB"].ne(target_article)]
        group = group.sort_values(["Заказы, шт/мес", "Выручка, руб/мес"], ascending=False).head(4)

        comp_articles = []
        for _, comp in group.iterrows():
            comp_articles.append(clean(comp["Артикул WB"]))
            rows.append(
                {
                    "Бренд": brand,
                    "Целевой артикул": target_article,
                    "Целевой товар": target_name,
                    "Конкурент артикул": clean(comp["Артикул WB"]),
                    "Конкурент товар": clean(comp["Товар"]),
                    "Продавец": clean(comp["Продавец"]),
                    "Заказы конкурента, шт/мес": clean(comp["Заказы, шт/мес"]),
                    "Выручка конкурента, руб/мес": clean(comp["Выручка, руб/мес"]),
                    "Цена конкурента с СПП, ₽": clean(comp["Цена с СПП, ₽"]),
                    "Тип подбора": "тот же баркод/товар из прайса",
                }
            )
        queue_rows.append(
            {
                "Бренд": brand,
                "Целевой артикул": target_article,
                "Артикулы для WB сравнения": ", ".join([target_article] + comp_articles),
                "Конкурентов найдено": len(comp_articles),
            }
        )

    detail = pd.DataFrame(rows)
    queue = pd.DataFrame(queue_rows)
    out_xlsx = ROOT / "outputs/wb_product_research/card_comparison_competitors_2026-06-01.xlsx"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        detail.to_excel(writer, index=False, sheet_name="Конкуренты")
        queue.to_excel(writer, index=False, sheet_name="Очередь WB")

    print("DETAIL_TSV")
    print("\t".join(detail.columns))
    for _, row in detail.iterrows():
        print("\t".join(clean(row[col]) for col in detail.columns))
    print("QUEUE_TSV")
    print("\t".join(queue.columns))
    for _, row in queue.iterrows():
        print("\t".join(clean(row[col]) for col in queue.columns))
    print(f"FILE\t{out_xlsx}")


if __name__ == "__main__":
    main()
