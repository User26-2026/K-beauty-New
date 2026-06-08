from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


ROOT = Path("/Users/nikita/Desktop/Новая папка 2")
SOURCE = ROOT / "outputs/wb_product_research/wb_card_comparison_source_2026-05-29.xlsx"
OUTPUT = ROOT / "outputs/wb_product_research/wb_card_comparison_template_data_2026-05-29.json"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def nm_from_header(header: str) -> str | None:
    match = re.search(r"Артикул WB\s+(\d+)", str(header))
    return match.group(1) if match else None


wb = openpyxl.load_workbook(SOURCE, data_only=True)
ws = wb["Показатели"]

current_cols: list[tuple[int, str]] = []
previous_by_nm: dict[str, int] = {}

for col in range(2, ws.max_column + 1):
    header = clean(ws.cell(2, col).value)
    if not header or "Разница" in str(header):
        continue
    nm = nm_from_header(str(header))
    if not nm:
        continue
    if "предыдущий период" in str(header):
        previous_by_nm[nm] = col
    else:
        current_cols.append((col, nm))

row_map = {
    "Название": 3,
    "Категория": 4,
    "Предмет": 5,
    "Дата создания": 6,
    "Бренд": 7,
    "Рейтинг карточки": 8,
    "Рейтинг по отзывам": 9,
    "Количество отзывов": 10,
    "Минимальная цена со скидкой (по размерам), ₽": 11,
    "Максимальная цена со скидкой (по размерам), ₽": 12,
    "Медианная цена покупателя, ₽": 13,
    "Среднее время доставки": 14,
    "Средняя позиция": 15,
    "Статус продвижения": 16,
    "Показы": 17,
    "Переходы в карточку": 18,
    "CTR, %": 19,
    "Положили в корзину": 20,
    "Конверсия в корзину, %": 21,
    "Заказали, шт за месяц": 22,
    "Заказали, шт за прошлый период": 22,
    "Конверсия в заказ, %": 23,
    "Выкупили, шт": 24,
    "Процент выкупа, %": 25,
    "Отмены, шт": 26,
}

products = []
for col, nm in current_cols:
    product = {
        "nm": nm,
        "values": {},
        "previous": {},
    }
    for label, row in row_map.items():
        if label == "Заказали, шт за прошлый период":
            prev_col = previous_by_nm.get(nm)
            product["values"][label] = clean(ws.cell(row, prev_col).value) if prev_col else ""
        else:
            product["values"][label] = clean(ws.cell(row, col).value)
    products.append(product)

data = {
    "title": "Сравнение карточек: Celimax пенки",
    "period": clean(wb["Общая информация"].cell(2, 2).value),
    "previous_period": clean(wb["Общая информация"].cell(3, 2).value),
    "products": products,
}

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
print(OUTPUT)
