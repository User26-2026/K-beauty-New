from __future__ import annotations

import argparse
import io
import re
import zipfile
from pathlib import Path

import openpyxl


METRICS = [
    "Название",
    "Категория",
    "Предмет",
    "Бренд",
    "Рейтинг карточки",
    "Рейтинг по отзывам",
    "Количество отзывов",
    "Минимальная цена со скидкой (по размерам), ₽",
    "Максимальная цена со скидкой (по размерам), ₽",
    "Медианная цена покупателя, ₽",
    "Среднее время доставки",
    "Средняя позиция",
    "Показы",
    "Переходы в карточку",
    "CTR",
    "Положили в корзину",
    "Конверсия в корзину, %",
    "Заказали, шт за месяц",
    "Заказали, шт за квартал",
    "Конверсия в заказ, %",
    "Выкупили, шт",
    "Процент выкупа",
    "Отменили, шт",
]

ALIASES = {
    "Переходы в карточку": ["Переходы в карточку", "Переход в карточку, шт"],
    "CTR": ["CTR"],
    "Положили в корзину": ["Положили в корзину", "Добавления в корзину, шт"],
    "Заказали, шт за месяц": ["Заказали, шт за месяц", "Заказы, шт"],
    "Заказали, шт за квартал": ["Заказали, шт за квартал"],
    "Выкупили, шт": ["Выкупили, шт", "Выкупы, шт"],
    "Процент выкупа": ["Процент выкупа", "Процент выкупа, %"],
    "Отменили, шт": ["Отменили, шт", "Отмены, шт"],
}


def load_workbook_from_maybe_zip(path: Path):
    data = path.read_bytes()
    if zipfile.is_zipfile(io.BytesIO(data)):
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            names = archive.namelist()
            inner_xlsx = [
                name
                for name in names
                if name.lower().endswith(".xlsx") and not name.startswith("__MACOSX/")
            ]
            if inner_xlsx and not any(name.startswith("xl/") for name in names):
                data = archive.read(inner_xlsx[0])
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def text(value) -> str:
    value = clean(value)
    if isinstance(value, float):
        return str(value).replace(".", ",")
    return str(value)


def nm_from_header(header: str) -> str | None:
    match = re.search(r"Артикул WB\s+(\d+)", header)
    return match.group(1) if match else None


def find_row_map(ws):
    row_map = {}
    for row in range(1, ws.max_row + 1):
        label = clean(ws.cell(row, 1).value)
        if label:
            row_map[str(label)] = row
    return row_map


def row_for_metric(metric: str, row_map: dict[str, int]) -> int | None:
    labels = ALIASES.get(metric, [metric])
    for label in labels:
        if label in row_map:
            return row_map[label]
    return None


def current_product_columns(ws):
    cols = []
    for col in range(2, ws.max_column + 1):
        header = clean(ws.cell(2, col).value)
        header_str = str(header)
        if not header_str or "Разница" in header_str or "предыдущий период" in header_str:
            continue
        nm = nm_from_header(header_str)
        if nm:
            cols.append((col, nm))
    return cols


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    args = parser.parse_args()

    wb = load_workbook_from_maybe_zip(Path(args.input))
    if "Показатели" not in wb.sheetnames:
        raise SystemExit("В файле нет листа `Показатели`.")

    ws = wb["Показатели"]
    row_map = find_row_map(ws)
    products = current_product_columns(ws)

    rows = [["Показатели", *[f"Артикул WB {nm}" for _, nm in products]]]
    for metric in METRICS:
        source_row = row_for_metric(metric, row_map)
        values = []
        for col, _ in products:
            values.append(text(ws.cell(source_row, col).value) if source_row else "")
        rows.append([metric, *values])

    print("\n".join("\t".join(row) for row in rows))


if __name__ == "__main__":
    main()
