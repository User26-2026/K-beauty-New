from __future__ import annotations

import csv
import io
import re
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path("/Users/nikita/Desktop/Новая папка 2")
PYTHON = "/Users/nikita/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3"
ANALYSIS = ROOT / "outputs/wb_product_research/Список товаров для анализа - проценты пересчитаны.xlsx"

SOURCES = {
    "Round Lab": Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос__Общая выдача_1.xlsx"),
    "AXIS-Y": Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_axis-y_Общая выдача_1.xlsx"),
    "DERMA FACTORY": Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_дерма фактори_Общая выдача_1.xlsx"),
}

MATCHERS = {
    "AXIS-Y": ROOT / "scripts/match_axisy_barcodes.py",
    "DERMA FACTORY": ROOT / "scripts/match_dermafactory_barcodes.py",
}

STOPWORDS = {
    "для",
    "лица",
    "крем",
    "сыворотка",
    "тонер",
    "пенка",
    "гель",
    "кожа",
    "кожи",
    "корея",
    "мл",
    "от",
    "с",
    "и",
    "на",
    "the",
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def num(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def max_drr(margin, price_spp, cost):
    if margin is None or price_spp in (None, 0) or cost in (None, 0):
        return None
    roi = margin / cost
    if roi <= 0.40:
        return None
    allowed_ad_spend = margin - cost * 0.40
    price_before_spp = round(price_spp / 0.70)
    return allowed_ad_spend / (price_before_spp * 0.50) * 100


def tokens(text: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-zа-я0-9]+", text.lower().replace("ё", "е"))
        if len(token) > 2 and token not in STOPWORDS
    }


def load_targets() -> pd.DataFrame:
    wb = openpyxl.load_workbook(ANALYSIS, data_only=True)
    ws = wb["Лист2"]
    headers = [ws.cell(1, col).value for col in range(1, 11)]
    rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, 11)]
        if not any(value not in (None, "") for value in values[:3]):
            continue
        item = dict(zip(headers, values))
        margin = num(item.get("Маржа/шт без ДРР, руб."))
        price = num(item.get("Цена с СПП топ карточки, ₽"))
        cost = num(item.get("Себестоимость, ₽"))
        revenue = num(item.get("Расч. выручка топ карточки, руб/мес")) or 0
        drr = max_drr(margin, price, cost)
        if drr is not None and revenue > 100_000:
            item["Артикул WB"] = str(int(num(item["Ссылка/артикул"])))
            item["Макс ДРР"] = round(drr, 1)
            rows.append(item)
    return pd.DataFrame(rows)


def load_source(brand: str) -> pd.DataFrame:
    source = pd.read_excel(SOURCES[brand])
    source["Артикул WB"] = source["Артикул"].astype(int).astype(str)
    source["Товар"] = source["Товар"].map(clean)
    source["Норм бренд"] = source["Бренд"].astype(str).str.lower().str.strip()
    source["Заказы"] = pd.to_numeric(source["Расч. заказы"], errors="coerce").fillna(0).astype(int)
    source["Выручка"] = pd.to_numeric(source["Расч. выручка\nпо заказам (с СПП)"], errors="coerce").fillna(0).astype(int)
    source["Цена"] = pd.to_numeric(source["Цена (с СПП)"], errors="coerce").fillna(0).astype(int)
    return source


def load_match(brand: str, source: pd.DataFrame) -> pd.DataFrame:
    if brand not in MATCHERS:
        return pd.DataFrame()
    raw = subprocess.check_output([PYTHON, str(MATCHERS[brand])], cwd=ROOT, text=True)
    matched = pd.DataFrame(list(csv.DictReader(io.StringIO(raw), delimiter="\t")))
    matched = matched[matched["Barcode из прайса"].astype(str).str.len() > 0].copy()
    return matched.merge(
        source[["Артикул WB", "Товар", "Продавец", "Предмет", "Заказы", "Выручка", "Цена"]],
        on="Артикул WB",
        how="left",
    )


def similar_rows(target, source: pd.DataFrame, exclude: set[str], limit: int) -> list[dict]:
    target_tokens = tokens(clean(target["Название"]))
    target_subject = clean(target.get("Предмет", ""))
    norm_brand = clean(target["Бренд"]).lower()
    if norm_brand == "round lab":
        brand_mask = source["Норм бренд"].eq("round lab")
    else:
        brand_mask = source["Норм бренд"].eq(norm_brand.lower())
    candidates = source[brand_mask].copy()
    candidates = candidates[~candidates["Артикул WB"].isin(exclude)].copy()
    candidates = candidates[(candidates["Заказы"] > 0) & (candidates["Выручка"] > 0)].copy()
    if candidates.empty:
        return []
    def score(row) -> tuple:
        overlap = len(tokens(row["Товар"]) & target_tokens)
        subject_bonus = 2 if target_subject and clean(row.get("Предмет")) == target_subject else 0
        return (overlap + subject_bonus, int(row["Заказы"]))
    candidates["_score"] = candidates.apply(lambda row: score(row)[0], axis=1)
    priority = candidates[candidates["_score"] > 0].sort_values(["_score", "Заказы"], ascending=False)
    fallback = candidates[candidates["_score"] <= 0].sort_values(["Заказы", "Выручка"], ascending=False)
    ordered = pd.concat([priority, fallback], ignore_index=True)
    result = []
    seen = set()
    for _, row in ordered.iterrows():
        if len(result) >= limit:
            break
        article = clean(row["Артикул WB"])
        if article in seen:
            continue
        seen.add(article)
        match_type = (
            "добор по похожему названию/предмету"
            if int(row.get("_score", 0)) > 0
            else "добор по продажам бренда из выгрузки"
        )
        result.append(
            {
                "article": article,
                "name": clean(row["Товар"]),
                "seller": clean(row["Продавец"]),
                "orders": clean(row["Заказы"]),
                "revenue": clean(row["Выручка"]),
                "price": clean(row["Цена"]),
                "type": match_type,
            }
        )
    return result


def main() -> None:
    targets = load_targets()
    sources = {brand: load_source(brand) for brand in SOURCES}
    matches = {brand: load_match(brand, sources[brand]) for brand in MATCHERS}
    detail_rows = []
    queue_rows = []
    compact_rows = []

    for _, target in targets.iterrows():
        brand = clean(target["Бренд"])
        article = clean(target["Артикул WB"])
        source = sources.get(brand)
        comps = []
        exclude = {article}

        if brand in matches:
            pool = matches[brand]
            target_match = pool[pool["Артикул WB"].eq(article)]
            if not target_match.empty:
                barcode = target_match.iloc[0]["Barcode из прайса"]
                group = pool[pool["Barcode из прайса"].eq(barcode)].copy()
                group = group[group["Артикул WB"].ne(article)]
                group = group[(group["Заказы"] > 0) & (group["Выручка"] > 0)].copy()
                group = group.sort_values(["Заказы", "Выручка"], ascending=False).head(4)
                for _, row in group.iterrows():
                    comps.append(
                        {
                            "article": clean(row["Артикул WB"]),
                            "name": clean(row["Товар"]),
                            "seller": clean(row["Продавец"]),
                            "orders": clean(row["Заказы"]),
                            "revenue": clean(row["Выручка"]),
                            "price": clean(row["Цена"]),
                            "type": "тот же товар по баркоду",
                        }
                    )
                    exclude.add(clean(row["Артикул WB"]))

        if source is not None and len(comps) < 4:
            comps.extend(similar_rows(target, source, exclude, 4 - len(comps)))

        if brand == "Round Lab":
            note = "В текущей выгрузке Round Lab нет внешних продавцов; это добор из Cream.Shop/смежных карточек."
        else:
            note = "Сначала точные карточки по баркоду, затем добор похожих."

        comp_articles = [comp["article"] for comp in comps[:4]]
        queue_rows.append(
            {
                "Бренд": brand,
                "Целевой артикул": article,
                "Артикулы для WB сравнения": ", ".join([article] + comp_articles),
                "Конкурентов найдено": len(comp_articles),
                "Комментарий": note,
            }
        )
        compact_row = {
            "Бренд": brand,
            "Целевой артикул": article,
            "Целевой товар": clean(target["Название"]),
        }
        for n, comp in enumerate(comps[:4], start=1):
            compact_row[f"Конкурент {n} артикул"] = comp["article"]
            compact_row[f"Конкурент {n} заказы/мес"] = comp["orders"]
            compact_row[f"Конкурент {n} выручка/мес"] = comp["revenue"]
        for n in range(len(comps[:4]) + 1, 5):
            compact_row[f"Конкурент {n} артикул"] = ""
            compact_row[f"Конкурент {n} заказы/мес"] = ""
            compact_row[f"Конкурент {n} выручка/мес"] = ""
        compact_row["Комментарий"] = note
        compact_rows.append(compact_row)
        for n, comp in enumerate(comps[:4], start=1):
            detail_rows.append(
                {
                    "Бренд": brand,
                    "Целевой артикул": article,
                    "Целевой товар": clean(target["Название"]),
                    "N": n,
                    "Конкурент артикул": comp["article"],
                    "Конкурент товар": comp["name"],
                    "Продавец": comp["seller"],
                    "Заказы конкурента, шт/мес": comp["orders"],
                    "Выручка конкурента, руб/мес": comp["revenue"],
                    "Цена конкурента с СПП, ₽": comp["price"],
                    "Тип подбора": comp["type"],
                }
            )

    detail = pd.DataFrame(detail_rows)
    queue = pd.DataFrame(queue_rows)
    compact = pd.DataFrame(compact_rows)
    out_xlsx = ROOT / "outputs/wb_product_research/card_comparison_queue_full_2026-06-01.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        compact.to_excel(writer, index=False, sheet_name="Для копирования")
        queue.to_excel(writer, index=False, sheet_name="Очередь WB")
        detail.to_excel(writer, index=False, sheet_name="Конкуренты")

    print("QUEUE_TSV")
    print("\t".join(queue.columns))
    for _, row in queue.iterrows():
        print("\t".join(clean(row[col]) for col in queue.columns))
    print("COMPACT_TSV")
    print("\t".join(compact.columns))
    for _, row in compact.iterrows():
        print("\t".join(clean(row[col]) for col in compact.columns))
    print("DETAIL_TSV")
    print("\t".join(detail.columns))
    for _, row in detail.iterrows():
        print("\t".join(clean(row[col]) for col in detail.columns))
    print(f"FILE\t{out_xlsx}")


if __name__ == "__main__":
    main()
