from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl


PRICE_PATH = Path(
    "/Users/nikita/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/"
    "AC94F4E0-F80E-4DED-B45E-FB6FEE8B19EE/DERMAFACTORY_2601.xlsx"
)
if not PRICE_PATH.exists():
    PRICE_PATH = Path("/Users/nikita/Downloads/DERMAFACTORY_2601.xlsx")
SOURCE_PATH = Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_дерма фактори_Общая выдача_1.xlsx")


def clean_barcode(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", "", str(value))


def num(value) -> float:
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0


def load_price() -> dict[str, dict]:
    wb = openpyxl.load_workbook(PRICE_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    items: dict[str, dict] = {}
    for r in range(3, ws.max_row + 1):
        barcode = clean_barcode(ws.cell(r, 2).value)
        en = str(ws.cell(r, 4).value or "").strip()
        if not barcode or not en:
            continue
        key = re.sub(r"^\[Derma Factory\]\s*", "", en).strip()
        items[key] = {
            "barcode": barcode,
            "price_name": en,
            "supply_krw": ws.cell(r, 7).value,
        }
    return items


def map_name(name: str) -> tuple[str | None, str]:
    low = name.lower()

    if "ниацинамид" in low and ("20%" in low or "20 %" in low):
        return "Niacinamide 20% Serum", "точное по 20% ниацинамиду"
    if "ниацинамид" in low and "порош" in low:
        return "Niacinamide 100 Powder", "точное по ниацинамид-порошку"
    if "100% ниацинамид" in low:
        return "Niacinamide 100 Powder", "точное по ниацинамид-порошку"
    if "ниацинамид" in low and ("прыщ" in low or "акне" in low):
        return "Niacinamide 20% Serum", "вероятно по ниацинамиду/акне"
    if "ниацинамид" in low and "каламин" in low:
        return "Niacinamide 10% Calamine Cream", "точное по ниацинамид + каламин"
    if "крем-дезодорант" in low or "дезодорант" in low:
        return "Niacinamide 10% Deo Cream", "точное по Deo Cream"
    if "ниацинамид" in low and "транексам" in low:
        return "Niacin Tranexamic Acid 13% Serum", "точное по ниацинамид + транексамовая кислота"
    if "ниацинамид" in low and ("пигмент" in low or "осветля" in low):
        return "Niacinamide + Arbutin 17% Serum", "вероятно по осветляющей сыворотке"

    if "ретинал" in low and "300" in low:
        return "Retinal 300ppm Cream", "точное по Retinal 300ppm"
    if "ретинол" in low and "300" in low:
        return "Retinal 300ppm Cream", "вероятно по Retinal/Retinol 300ppm"
    if "ретинал" in low and "центел" in low:
        return "Retinal Cica Ampoule", "точное по Retinal Cica"
    if "ретиноид" in low or "retinoid" in low:
        return "Retinoid 4000ppm Cream", "вероятно по Retinoid"

    if "транексам" in low and "6" in low:
        return "Tranexamic Acid 6% Cream", "точное по 6% транексамовой кислоте"
    if "транексам" in low and "крем" in low:
        return "Tranexamic Acid 6% Cream", "вероятно по транексамовому крему"

    if "pdrn 4" in low or "pdrn 4%" in low:
        return "PDRN 4% Ampoule", "точное по PDRN 4%"

    if "bakuchiol spicule" in low:
        return "Bakuchiol Spicule Cream", "точное по Bakuchiol Spicule"
    if "бакучиол" in low and ("микроиг" in low or "spicule" in low):
        return "Bakuchiol Spicule Cream", "точное по Bakuchiol Spicule"
    if "bakuchiol 1" in low or "бакучиол 1" in low:
        return "Bakuchiol 1% Cream", "точное по Bakuchiol 1%"
    if "бакучиол" in low and "сыворот" in low:
        return "Bakuchiol 5% Ampoule", "вероятно по Bakuchiol Ampoule"
    if "бакучиол" in low and "крем" in low:
        return "Bakuchiol 1% Cream", "вероятно по Bakuchiol Cream"

    if "аденозин" in low or "adenosine 7500" in low:
        return "Adenosine 7500ppm Water Essence", "точное по Adenosine 7500ppm"
    if "комбуч" in low:
        return "Kombucha 80% Treatment", "точное по Kombucha"
    if "daily pore toner" in low or "очищения пор daily pore" in low:
        return "Daily Pore Toner", "точное по Daily Pore Toner"
    if "pore clear pad" in low:
        return None, "в прайсе нет Pore Clear Pad"
    if "пэды pdrn" in low:
        return None, "в прайсе нет PDRN Toner Pad"

    if "кушон" in low and "spf" in low:
        return "Mineral Sun Cushion SPF50+/PA++++ (26g)", "точное по Sun Cushion"
    if "солнцезащит" in low and ("осветля" in low or "tone" in low):
        return "Mineral Tone-Up Sun Cream SPF50+/PA++++ (30g)", "вероятно по Tone-Up Sun Cream"
    if "солнцезащит" in low or "spf" in low:
        return "Mineral Mild Sun Cream SPF50+/PA++++ (30g)", "вероятно по SPF"

    if "центелл" in low and "крем" in low:
        return "Cica 53.2% Cream", "вероятно по Cica Cream"
    if "хауттюй" in low or "houttuynia" in low:
        return None, "в прайсе нет Houttuynia"
    if "барьер" in low or "восстановления барьера" in low:
        return "Skin Barrier Cream", "точное по Skin Barrier Cream"
    if "церамид" in low or "пантенол" in low:
        return "Cerapanthenol 8% Intensive Cream", "вероятно по Cerapanthenol"
    if "аквапорин" in low:
        return "Aqua Pore Moisture Cream", "точное по Aqua Pore/Aquaporin"

    if "волюфилин" in low and "стик" in low:
        return "Bifida Ferment 20% Ampoule Stick", "точное по стик-сыворотке Volufiline"
    if "volufiline" in low or "волюфилин" in low:
        return "Bifida Ferment 5% Ampoule", "вероятно по Volufiline Ampoule"

    if "collagen serum mist" in low or "сыворотка-мист" in low:
        return "Collagen Serum Mist", "точное по Collagen Serum Mist"
    if "коллаген" in low and "5 г" in low:
        return "Collagen 100 Powder", "вероятно по Collagen Powder"
    if "peptide eyelash" in low or "сыворотка для ресниц" in low:
        return "Peptide Eyelash Ampoule", "точное по Peptide Eyelash"
    if "матриксил" in low or "matrixyl" in low:
        return "Matrixyl 15% + Azelaiclin 2% Serum", "точное по Matrixyl 15%"
    if "пептид" in low and "крем" in low:
        return "Peptide Face & Eye Cream 20ml", "вероятно; объем в названии WB не указан"

    if "vita-c 80" in low or "vitamin c" in low:
        return "Vitamin C 80 Powder", "точное по Vita-C 80 Powder"
    if "гиалурон" in low:
        return "Hyaluronic Acid 1% Serum – Renewed (30ml → 80ml)", "вероятно по Hyaluronic Serum"
    if "увлажняющая сыворотка" in low or ("сыворотка" in low and "увлажняющая" in low):
        return "Hyaluronic Acid 1% Serum – Renewed (30ml → 80ml)", "вероятно по увлажняющей сыворотке"

    if "сыворот" in low and ("пигмент" in low or "осветля" in low):
        return "Niacinamide + Arbutin 17% Serum", "вероятно по осветляющей сыворотке"

    if "пенка" in low and "пор" in low:
        return "pH Balancing Cleansing Foam", "вероятно по пенке"
    if "тоник" in low and "пор" in low:
        return "Daily Pore Toner", "вероятно по тонику для пор"
    if "тонер" in low and "корея" in low:
        return "Bizero Mild Toner", "вероятно по базовому тонеру"
    if "тонер" in low and "увлажняющий" in low:
        return "Bizero Mild Toner", "вероятно по базовому тонеру"

    if "отбеливающий крем" in low or "пигментных пятен" in low and "крем" in low:
        return "The Whitening Cream", "вероятно по отбеливающему крему"

    return None, "нет надежного совпадения в прайсе"


def main() -> None:
    price = load_price()

    wb = openpyxl.load_workbook(SOURCE_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {name: i + 1 for i, name in enumerate(headers)}

    rows = []
    for r in range(2, ws.max_row + 1):
        brand = str(ws.cell(r, idx["Бренд"]).value or "").strip().upper()
        orders = num(ws.cell(r, idx["Расч. заказы"]).value)
        if brand != "DERMA FACTORY" or orders <= 0:
            continue
        article = int(num(ws.cell(r, idx["Артикул"]).value))
        name = str(ws.cell(r, idx["Товар"]).value or "").strip()
        price_name, confidence = map_name(name)
        barcode = ""
        supply = ""
        if price_name and price_name in price:
            barcode = price[price_name]["barcode"]
            supply = price[price_name]["supply_krw"]
            price_name_full = price[price_name]["price_name"]
        else:
            price_name_full = price_name or ""
        rows.append((orders, article, name, barcode, price_name_full, supply, confidence))

    rows.sort(reverse=True, key=lambda item: item[0])
    writer = csv.writer(sys.stdout, delimiter="\t", lineterminator="\n")
    writer.writerow(["Артикул WB", "Название товара", "Barcode из прайса", "Название в прайсе", "Supply KRW", "Сопоставление"])
    seen = set()
    for _, article, name, barcode, price_name, supply, confidence in rows:
        if article in seen:
            continue
        seen.add(article)
        writer.writerow([article, name, barcode, price_name, supply, confidence])


if __name__ == "__main__":
    main()
