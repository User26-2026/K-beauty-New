import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SOURCE = Path("/Users/nikita/Downloads/2026-06-01_Поисковый_запрос_axis-y_Общая выдача_1.xlsx")
PRICE = Path("/Users/nikita/Downloads/AXIS-Y PRICE LIST(-VAT)_26.01.xlsx")


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def norm(value):
    value = clean_text(value).lower()
    value = value.replace("ё", "е")
    value = re.sub(r"[^a-zа-я0-9.%+]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def load_price_catalog():
    ws = load_workbook(PRICE, data_only=True, read_only=True)["AXIS-Y Supply Price List"]
    catalog = {}
    for row in ws.iter_rows(min_row=3, max_row=80, values_only=True):
        barcode = row[0]
        if barcode in (None, "", "-"):
            continue
        if not str(barcode).strip().isdigit():
            continue
        # Main products use col B; samples/sets use col E.
        name = row[1] or row[4]
        supply = row[8]
        if not name or supply in (None, "", "#N/A"):
            continue
        key = clean_text(name)
        catalog[key] = {
            "barcode": str(int(barcode)),
            "name": key,
            "supply": int(float(supply)) if isinstance(supply, (int, float)) else clean_text(supply),
        }
    return catalog


CATALOG = load_price_catalog()


def find_price_name_contains(fragment):
    frag = fragment.lower()
    for name, data in CATALOG.items():
        if frag in name.lower():
            return data
    raise KeyError(fragment)


P = {
    "daily_toner_200": find_price_name_contains("Daily Purifying Treatment Toner 200ml"),
    "daily_toner_80": find_price_name_contains("Daily Purifying Treatment Toner 80ml"),
    "artichoke_ampoule": find_price_name_contains("Artichoke Intensive Skin Barrier Ampoule"),
    "quinoa_cleanser": find_price_name_contains("Quinoa One-Step Balanced Gel Cleanser"),
    "dark_spot_serum": find_price_name_contains("Dark Spot Correcting Glow Serum"),
    "glow_toner": find_price_name_contains("Dark Spot Correcting Glow Toner"),
    "glow_cream": find_price_name_contains("Dark Spot Correcting Glow Cream"),
    "txa_cream": find_price_name_contains("TXA 2.5% Intensive Brightening Cream"),
    "duo_cream": find_price_name_contains("Cera-Heart My Type Duo Cream"),
    "sunscreen": find_price_name_contains("Complete No-Stress Physical Sunscreen"),
    "mugwort_mask_100": find_price_name_contains("Mugwort Pore Clarifying Wash Off Pack 100ml"),
    "mugwort_mask_50": find_price_name_contains("Mugwort Pore Clarifying Wash Off Pack 50ml"),
    "cleansing_foam": find_price_name_contains("Sunday Morning Refreshing Cleansing Foam"),
    "gel_mask_100": find_price_name_contains("New Skin Resolution Gel Mask 100ml"),
    "gel_mask_50": find_price_name_contains("New Skin Resolution Gel Mask 50ml"),
    "heartleaf_cream": find_price_name_contains("Heartleaf My Type Calming Cream"),
    "blemish_treatment": find_price_name_contains("Spot the Difference Blemish Treatment"),
    "pha_peel": find_price_name_contains("PHA Resurfacing Glow Peel"),
    "lha_cream": find_price_name_contains("LHA Peel & Fill Pore Balancing Cream"),
    "calamine_serum": find_price_name_contains("CALAMINE Pore Control Capsule Serum"),
    "panthenol_cream": find_price_name_contains("PANTHENOL 10 Skin Smoothing Shield Cream"),
    "lip_oil": find_price_name_contains("Vita Glow Lip Oil"),
    "eye_serum": find_price_name_contains("Vegan Collagen Eye Serum"),
    "biome_toner": find_price_name_contains("Biome Comforting Infused Toner"),
    "biome_skin_lux": find_price_name_contains("Biome Skin Lux Edition"),
    "mini_glow_set": find_price_name_contains("The Mini Glow Set"),
    "mask_duo": find_price_name_contains("Mask Now Glow Later Duo Set"),
    "mini_glow_trio": find_price_name_contains("The Mini Glow Trio"),
    "glass_skin_duo": find_price_name_contains("Glass Skin Duo"),
}

# These two rows in the price list are written as shade text in the product cell.
for data in CATALOG.values():
    lname = data["name"].lower()
    if lname.startswith("chilled berry"):
        P["lip_oil_chilled"] = data
    if lname.startswith("cozy fig"):
        P["lip_oil_cozy"] = data

P["lip_oil_chilled"]["name"] = "AXIS-Y Vita Glow Lip Oil Chilled Berry"
P["lip_oil_cozy"]["name"] = "AXIS-Y Vita Glow Lip Oil Cozy Fig"


def choose(row):
    title = norm(row["Товар"])
    article = str(int(row["Артикул"]))

    def ret(key, status="точно по названию/линейке"):
        d = P[key]
        return d["barcode"], d["name"], d["supply"], status

    # Sets first, because many contain words like toner/serum/cream.
    if "glass skin duo" in title or (("для лица и век" in title or "лица и для век" in title) and "набор" in title):
        return ret("glass_skin_duo")
    if "mini glow trio" in title:
        return ret("mini_glow_trio")
    if "mini glow set" in title or "набор сывороток гель тоник" in title or "набор миниатюр" in title:
        return ret("mini_glow_set")
    if "набор масок" in title or ("глиняная" in title and "гелевая" in title):
        return ret("mask_duo")
    if "подарочный набор" in title or title.startswith("набор axis"):
        return "", "", "", "неоднозначный набор: нужно открыть карточку и сверить состав"

    # Lip oils.
    if "chilled berry" in title:
        return ret("lip_oil_chilled")
    if "cozy fig" in title:
        return ret("lip_oil_cozy")
    if "vita glow lip oil" in title or "блеск для губ" in title:
        return ret("lip_oil", "вероятно, оттенок нужно проверить")

    # Eye serum.
    if "глаз" in title or "век" in title or "collagen eye serum" in title or "коллаген" in title:
        return ret("eye_serum")

    # Toners/mists.
    if "glow toner" in title or "тонер мист" in title or "мист" in title or "тонер м..." in title:
        return ret("glow_toner")
    if "тонер" in title or "тоник" in title:
        if "80 мл" in title or "80мл" in title:
            return ret("daily_toner_80")
        if "барьер" in title or "восстанавливающий" in title:
            return ret("biome_toner", "вероятно: в прайсе это Biome Comforting Infused Toner")
        return ret("daily_toner_200", "вероятно: очищающий/проблемная кожа = Daily Purifying Treatment Toner")

    # Cleansers/peels.
    if "пенк" in title or "умывания" in title:
        return ret("cleansing_foam")
    if "пилинг" in title or "скатка" in title:
        return ret("pha_peel")

    # Masks.
    if "глиняная" in title or "полын" in title:
        if "50 мл" in title or "50мл" in title:
            return ret("mugwort_mask_50")
        return ret("mugwort_mask_100")
    if "хауттюйни" in title and "маска" in title:
        if "50 мл" in title or "50мл" in title:
            return ret("gel_mask_50")
        return ret("gel_mask_100")

    # Sunscreens before creams, because WB titles often include "крем".
    if "солнц" in title or "spf" in title or "спф" in title:
        return ret("sunscreen")

    # Creams.
    if "транексам" in title or "txa" in title:
        return ret("txa_cream")
    if "panthenol" in title or "пантенол" in title:
        return ret("panthenol_cream")
    if "хауттюйни" in title and ("крем" in title or "гель" in title):
        return ret("heartleaf_cream")
    if "duo cream" in title or "двойной" in title:
        return ret("duo_cream")
    if "сужения пор" in title and "крем" in title:
        return ret("lha_cream")
    if "крем" in title:
        if "освет" in title or "correcting" in title or "постакне" in title or "пигментац" in title or "отбел" in title:
            return ret("glow_cream")
        return "", "", "", "крем без точного указания линейки: нужно проверить карточку"

    # Serums/ampoules.
    if "каламин" in title:
        return ret("calamine_serum")
    if "артиш" in title:
        return ret("artichoke_ampoule")
    if "прыщ" in title or "акне" in title:
        return ret("dark_spot_serum", "вероятно: в названии WB сыворотка с ниацинамидом")
    if "сыворот" in title:
        if "освет" in title or "пигмент" in title or "ниацинамид" in title:
            return ret("dark_spot_serum")
        return "", "", "", "сыворотка без точного указания линейки: нужно проверить карточку"

    return "", "", "", "не найдено уверенное сопоставление"


def main():
    df = pd.read_excel(SOURCE)
    df = df[df["Бренд"].astype(str).str.upper().eq("AXIS-Y")].copy()
    df["orders"] = pd.to_numeric(df["Расч. заказы"], errors="coerce").fillna(0)
    df = df[df["orders"] > 0].drop_duplicates("Артикул").sort_values("orders", ascending=False)
    print("Артикул WB\tНазвание товара\tBarcode из прайса\tНазвание в прайсе\tSupply KRW\tСопоставление")
    for _, row in df.iterrows():
        barcode, price_name, supply, status = choose(row)
        title = clean_text(row["Товар"])
        print(f"{int(row['Артикул'])}\t{title}\t{barcode}\t{price_name}\t{supply}\t{status}")


if __name__ == "__main__":
    main()
