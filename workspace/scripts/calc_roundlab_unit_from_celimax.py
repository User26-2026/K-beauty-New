from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path("/Users/nikita/Desktop/Новая папка 2")
INPUT = Path("/Users/nikita/Downloads/Celimax.xlsx")
OUTPUT = ROOT / "outputs/wb_product_research/roundlab_unit_economics_from_celimax_2026-06-01.json"


def as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def volume_liters(name: str, line: str, subject: str) -> float:
    low = f"{name} {line} {subject}".lower()
    if "500" in low or "400" in low or "гель для душа" in low or "мицелляр" in low:
        return 1.0
    if "тонер" in low or "пенк" in low or "cleanser" in low:
        return 1.0
    return 0.4


def buyout_percent(name: str, line: str, subject: str) -> float:
    low = f"{name} {line} {subject}".lower()
    if "пенк" in low or "мицелляр" in low or "гель для душа" in low:
        return 96.0
    return 92.19


def calc_row(source: dict, constants: dict) -> dict:
    price_spp = as_number(source["Цена с СПП, ₽"])
    orders = as_number(source["Расч. заказы"]) or 0.0
    cost = as_number(source["Себестоимость в ₽ x1,4"])
    name = str(source["Товар"])
    line = str(source["Линейка/тип"])
    subject = str(source["Предмет"])

    result = {
        "Артикул WB": int(as_number(source["Артикул WB"]) or 0),
        "Товар": name,
        "Линейка/тип": line,
        "Предмет": subject,
        "Цена с СПП, ₽": price_spp,
        "Расч. заказы": orders,
        "Расч. выручка, ₽": as_number(source["Расч. выручка, ₽"]) or 0.0,
        "Упущенная выручка, ₽": as_number(source["Упущенная выручка, ₽"]) or 0.0,
        "Себ-ть": cost,
        "% СПП": constants["spp"],
        "% выкупа": buyout_percent(name, line, subject),
        "Объем, литр": volume_liters(name, line, subject),
        "Комментарий": "",
    }

    if price_spp is None or cost is None:
        result["Комментарий"] = "Не посчитано: нет себестоимости из прайса Round Lab" if cost is None else "Не посчитано: нет цены"
        numeric_blanks = [
            "Цена продажи на WB",
            "РРЦ плановая без СПП",
            "Цена с СПП расчетная",
            "Доп обработка склад",
            "Налог УСН",
            "Стоимость денег",
            "Логистика WB",
            "Комиссия WB",
            "Эквайринг",
            "Обратная логистика",
            "Хранение",
            "Рекламные расходы",
            "Брак",
            "Платная приемка",
            "МАРЖА с ед.",
            "ROI %",
            "ИТОГО себес.",
            "Итого прибыль",
            "Проходной ДРР от WB цены",
        ]
        for key in numeric_blanks:
            result[key] = None
        return result

    price_wb = round(price_spp / (1 - constants["spp"]))
    price_spp_calc = price_wb * (1 - constants["spp"])
    handling = constants["handling"]
    tax = price_spp_calc * constants["tax"]
    money_cost = 0.0
    logistics = constants["logistics_manual"]
    commission = price_wb * constants["commission"]
    acquiring = price_spp_calc * constants["acquiring"]
    reverse_logistics = constants["reverse_logistics"] * ((100 - result["% выкупа"]) / 100)
    storage = constants["storage_per_liter_day"] * constants["storage_days"]
    ads = price_wb * constants["ads"]
    defect = 0.0
    paid_acceptance = 0.0
    margin = (
        price_wb
        - handling
        - tax
        - money_cost
        - logistics
        - commission
        - acquiring
        - reverse_logistics
        - storage
        - ads
        - cost
    )
    margin_before_ads = margin + ads
    roi = margin / cost if cost else None
    total_cost = cost * orders
    total_profit = margin * orders
    pass_drr = margin_before_ads / price_wb if price_wb else None

    result.update(
        {
            "Цена продажи на WB": price_wb,
            "РРЦ плановая без СПП": price_wb,
            "Цена с СПП расчетная": price_spp_calc,
            "Доп обработка склад": handling,
            "Налог УСН": tax,
            "Стоимость денег": money_cost,
            "Логистика WB": logistics,
            "Комиссия WB": commission,
            "Эквайринг": acquiring,
            "Обратная логистика": reverse_logistics,
            "Хранение": storage,
            "Рекламные расходы": ads,
            "Брак": defect,
            "Платная приемка": paid_acceptance,
            "МАРЖА с ед.": margin,
            "ROI %": roi,
            "ИТОГО себес.": total_cost,
            "Итого прибыль": total_profit,
            "Проходной ДРР от WB цены": pass_drr,
            "Комментарий": "Посчитано по логике вкладки Юнит, реклама 10% уже вычтена",
        }
    )
    return result


def main() -> None:
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    unit = wb["Юнит"]
    source = wb["Лист10"]

    constants = {
        "handling": float(unit["B1"].value),
        "ads": float(unit["B2"].value),
        "storage_per_liter_day": float(unit["B6"].value),
        "storage_days": float(unit["B7"].value),
        "tax": float(unit["B8"].value),
        "reverse_logistics": float(unit["S3"].value),
        "commission": float(unit["S6"].value),
        "acquiring": float(unit["S7"].value),
        "spp": 0.30,
        "logistics_manual": 90.0,
    }

    headers = [source.cell(1, c).value for c in range(1, source.max_column + 1)]
    rows = []
    for r in range(2, source.max_row + 1):
        row = {headers[c - 1]: source.cell(r, c).value for c in range(1, source.max_column + 1)}
        if not row.get("Артикул WB"):
            continue
        rows.append(calc_row(row, constants))

    summary_rows = [r for r in rows if r["МАРЖА с ед."] is not None]
    summary = {
        "source": str(INPUT),
        "rows_total": len(rows),
        "rows_calculated": len(summary_rows),
        "rows_missing_cost": len(rows) - len(summary_rows),
        "total_profit": sum(r["Итого прибыль"] for r in summary_rows),
        "total_cost": sum(r["ИТОГО себес."] for r in summary_rows),
        "avg_margin": sum(r["МАРЖА с ед."] for r in summary_rows) / len(summary_rows) if summary_rows else None,
        "constants": constants,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({"summary": summary, "rows": rows}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUTPUT)


if __name__ == "__main__":
    main()
