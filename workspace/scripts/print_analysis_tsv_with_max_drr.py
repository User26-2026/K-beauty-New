from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl


INPUT = Path("/Users/nikita/Desktop/Новая папка 2/outputs/wb_product_research/Список товаров для анализа - проценты пересчитаны.xlsx")


def fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(round(value, 1)).replace(".", ",")
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def as_number(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def max_drr(margin: float | None, price_spp: float | None, cost: float | None) -> float | None:
    if margin is None or price_spp in (None, 0) or cost in (None, 0):
        return None
    roi = margin / cost
    if roi <= 0.40:
        return None
    allowed_ad_spend = margin - cost * 0.40
    # In the user's unit table, a 20% DRR setting on the first Round Lab SKU
    # lowers ROI from 90.9% to about 53.7%. That matches ad spend calculated as
    # half of the pre-SPP WB price, so we use the same base here.
    price_before_spp = round(price_spp / 0.70)
    unit_table_drr_base = price_before_spp * 0.50
    return allowed_ad_spend / unit_table_drr_base * 100


def main() -> None:
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    ws = wb["Лист2"]
    lines = []
    for row in range(1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, 11)]
        if row != 1 and not any(value not in (None, "") for value in values[:3]):
            continue
        if row == 1:
            values.append("Макс. ДРР при ROI 40%, %")
        else:
            margin = as_number(values[7])
            price_spp = as_number(values[8])
            cost = as_number(values[9])
            drr = max_drr(margin, price_spp, cost)
            values.append("" if drr is None else round(drr, 1))
        lines.append("\t".join(fmt(value) for value in values))
    print("\n".join(lines))


if __name__ == "__main__":
    main()
